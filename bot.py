import os, json, logging, random, string, threading, time, requests
import hmac, hashlib, functools, urllib.parse
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, g
import pg8000

try:
    import boto3
    from botocore.exceptions import ClientError
    S3_AVAILABLE = True
except ImportError:
    S3_AVAILABLE = False
    logging.warning("boto3 not installed — S3 disabled")

logging.basicConfig(level=logging.INFO)
app = Flask(__name__)

@app.after_request
def add_cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

# ─── TOKENS & CONFIG ────────────────────────────────────────────────
SELLER_TOKEN    = os.environ.get('SELLER_TOKEN')
BUYER_TOKEN     = os.environ.get('BUYER_TOKEN')
ADMIN_ID        = int(os.environ.get('ADMIN_ID', '0'))
PAYME_NUMBER    = os.environ.get('PAYME_NUMBER', '+998913968946')
COMMISSION_RATE = 0.05  # 5%
CLICK_TOKEN    = os.environ.get('CLICK_TOKEN', '')  # Click Terminal payment token

# Kategoriyalar — bot va sayt uchun bir xil
CATEGORIES = [
    ('Kiyim',            '👕'),
    ('Poyabzal',         '👟'),
    ('Sumka',            '👜'),
    ('Soat & Zargarlik', '⌚'),
    ('Elektronika',      '📱'),
    ('Ofis & Kompyuter', '💻'),
    ('Kantselyariya',    '✏️'),
    ('Avto',             '🚗'),
    ('Oziq-ovqat',       '🍎'),
    ('Uy-joy',           '🏠'),
    ('Parfyumeriya',     '💄'),
    ('Salomatlik',       '💊'),
    ('Sport',            '⚽'),
    ('Bolalar',          '🧸'),
    ("O'yin & Hobby",    '🎮'),
    ('Boshqa',           '📦'),
]

# AWS S3
AWS_ACCESS_KEY_ID     = os.environ.get('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.environ.get('AWS_SECRET_ACCESS_KEY')
AWS_BUCKET_NAME       = os.environ.get('AWS_BUCKET_NAME', 'joynshop-media')
AWS_REGION            = os.environ.get('AWS_REGION', 'eu-central-1')
CDN_BASE_URL          = os.environ.get('CDN_BASE_URL', '')

def get_s3():
    if not S3_AVAILABLE or not AWS_ACCESS_KEY_ID:
        return None
    return boto3.client('s3',
        region_name=AWS_REGION,
        aws_access_key_id=AWS_ACCESS_KEY_ID,
        aws_secret_access_key=AWS_SECRET_ACCESS_KEY
    )

def upload_photo_to_s3(file_id, bot_token):
    s3 = get_s3()
    if not s3:
        return None
    try:
        r = requests.get(f'https://api.telegram.org/bot{bot_token}/getFile',
                         params={'file_id': file_id}, timeout=10).json()
        if not r.get('ok'):
            return None
        file_path = r['result']['file_path']
        ext = file_path.rsplit('.', 1)[-1].lower() if '.' in file_path else 'jpg'
        tg_url = f'https://api.telegram.org/file/bot{bot_token}/{file_path}'
        img_data = requests.get(tg_url, timeout=20).content
        key = f'products/{file_id}.{ext}'
        s3.put_object(Bucket=AWS_BUCKET_NAME, Key=key, Body=img_data,
                      ContentType=f'image/{ext}')
        if CDN_BASE_URL:
            return f'{CDN_BASE_URL}/{key}'
        return f'https://{AWS_BUCKET_NAME}.s3.{AWS_REGION}.amazonaws.com/{key}'
    except Exception as e:
        logging.error(f'S3 upload error: {e}')
        return None

def upload_photo_async(file_id, bot_token, state_ref):
    def _upload():
        url = upload_photo_to_s3(file_id, bot_token)
        if url and state_ref is not None:
            if 'photo_urls' not in state_ref: state_ref['photo_urls'] = []
            if url not in state_ref['photo_urls']:
                state_ref['photo_urls'].append(url)
            logging.info(f'S3 async done: {url}')
    threading.Thread(target=_upload, daemon=True).start()

DASHBOARD_PASSWORD = os.environ.get('DASHBOARD_PASSWORD', 'joynshop2026')
BUYER_BOT_USERNAME = os.environ.get('BUYER_BOT_USERNAME', 'joynshop_bot')
APP_URL            = os.environ.get('APP_URL', '')
# Public Flask backend URL — used when a link must hit the bot's routes
# (e.g. /live/<id>) and APP_URL points to a static frontend (Vercel) that
# doesn't proxy that path. Falls back to APP_URL when not set.
BACKEND_URL        = os.environ.get('BACKEND_URL', APP_URL)

def setup_bot_ui():
    miniapp_url = f"{(BACKEND_URL or APP_URL or '').rstrip('/')}/miniapp" if (BACKEND_URL or APP_URL) else None

    if BUYER_TOKEN:
        if miniapp_url:
            requests.post(f'https://api.telegram.org/bot{BUYER_TOKEN}/setChatMenuButton', json={
                'menu_button': {'type': 'web_app', 'text': 'Joynshop', 'web_app': {'url': miniapp_url}}
            })
        requests.post(f'https://api.telegram.org/bot{BUYER_TOKEN}/setMyCommands', json={
            'commands': [
                {'command': 'start',     'description': '🏠 Bosh sahifa'},
                {'command': 'mystatus',  'description': '📋 Mening buyurtmalarim'},
                {'command': 'myprofile', 'description': '👤 Profilim'},
                {'command': 'feedback',  'description': '✍️ Fikr bildirish'},
                {'command': 'settings',  'description': '⚙️ Sozlamalar'},
            ]
        })

    if SELLER_TOKEN:
        requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/setChatMenuButton', json={
            'menu_button': {'type': 'commands'}
        })
        requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/setMyCommands', json={
            'commands': [
                {'command': 'start',      'description': '🏠 Bosh sahifa'},
                {'command': 'addproduct', 'description': '➕ Mahsulot qo\'shish'},
                {'command': 'myproducts', 'description': '📦 Mahsulotlarim'},
                {'command': 'myorders',   'description': '📋 Buyurtmalar'},
                {'command': 'mystats',    'description': '📊 Statistika'},
                # LIVE FROZEN: traction olganidan keyin uncomment qilinadi
                # {'command': 'golive',     'description': '🔴 Live boshlash'},
                # {'command': 'mylive',     'description': '📺 Live dashboard'},
                {'command': 'mychannels', 'description': '📢 Kanallarim'},
                {'command': 'help',       'description': 'ℹ️ Yordam'},
            ]
        })
    logging.info("Bot UI setup done.")

# ─── SHARED STORAGE ─────────────────────────────────────────────────
products        = {}
groups          = {}
orders          = {}
wishlists       = {}
buyer_profiles  = {}
refund_requests = {}
seller_state    = {}
customers       = {}  # {seller_id: {user_id: {...}}}
lives           = {}  # {live_id: {...}} - Live Commerce streams
_photo_url_cache = {}
seller_shops    = {}
seller_products = {}
seller_profiles = {}  # {uid: {legal_status, stir, bank_account, bank_name, bank_mfo, director_name, legal_completed_at}}
verified_channels       = {}
pending_moderator_codes = {}
referrals               = {}
referral_map            = {}

# ─── PERSISTENCE (PostgreSQL) ───────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL')

def get_db():
    import urllib.parse, ssl
    r = urllib.parse.urlparse(DATABASE_URL)
    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
    return pg8000.connect(
        host=r.hostname, port=r.port or 5432,
        database=r.path.lstrip('/'),
        user=r.username, password=r.password,
        ssl_context=ssl_ctx
    )

def init_db():
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS joynshop_data "
            "(key TEXT PRIMARY KEY, value TEXT)"
        )
        conn.commit()
        cur.close(); conn.close()
        logging.info("DB initialized")
    except Exception as e:
        logging.error(f"init_db error: {e}", exc_info=True)

def save_data():
    if not DATABASE_URL:
        logging.warning("No DATABASE_URL")
        return
    for attempt in range(3):
        try:
            data = {
                'products':               products,
                'groups':                 groups,
                'orders':                 orders,
                'wishlists':              wishlists,
                'buyer_profiles':         buyer_profiles,
                'refund_requests':        refund_requests,
                'seller_products':        {str(k): v for k, v in seller_products.items()},
                'seller_shops':           {str(k): v for k, v in seller_shops.items()},
                'seller_profiles':        {str(k): v for k, v in seller_profiles.items()},
                'verified_channels':      verified_channels,
                'pending_moderator_codes':pending_moderator_codes,
                'referrals':              referrals,
                'referral_map':           {str(k): v for k, v in referral_map.items()},
                'customers':              {str(k): v for k, v in customers.items()},
                'lives':                  {str(k): v for k, v in lives.items()},
            }
            payload = json.dumps(data, ensure_ascii=False, default=str)
            conn    = get_db()
            cur     = conn.cursor()
            cur.execute(
                "INSERT INTO joynshop_data (key, value) VALUES ('main', %s) "
                "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
                (payload,)
            )
            conn.commit()
            cur.close(); conn.close()
            logging.info(f"Data saved: {len(products)} products, {len(orders)} orders")
            return
        except Exception as e:
            logging.error(f"save_data error (attempt {attempt+1}): {e}")
            if attempt == 2:
                logging.error("save_data failed 3 times!", exc_info=True)

def load_data():
    global products, groups, orders, wishlists, buyer_profiles
    global refund_requests, seller_products, verified_channels
    global pending_moderator_codes, referrals, referral_map, seller_shops, customers, lives
    global seller_profiles
    if not DATABASE_URL:
        logging.warning("No DATABASE_URL — starting fresh")
        return
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT value FROM joynshop_data WHERE key = 'main'")
        row  = cur.fetchone()
        cur.close(); conn.close()
        if not row:
            logging.info("No data in DB — starting fresh")
            return
        data = json.loads(row[0]) if isinstance(row[0], str) else row[0]
        products               = data.get('products', {})
        # Migration: ensure is_active default for products saved before the field existed
        for _p in products.values():
            _p.setdefault('is_active', True)
            _p.setdefault('mxik_code', None)
            _p.setdefault('mxik_name', None)
        groups                 = data.get('groups', {})
        orders                 = data.get('orders', {})
        wishlists              = data.get('wishlists', {})
        buyer_profiles         = data.get('buyer_profiles', {})
        refund_requests        = data.get('refund_requests', {})
        verified_channels      = data.get('verified_channels', {})
        raw_ss = data.get('seller_shops', {})
        seller_shops = {int(k) if str(k).isdigit() else k: v for k, v in raw_ss.items()}
        # Seller profiles (legal info)
        raw_sp = data.get('seller_profiles', {})
        seller_profiles = {int(k) if str(k).isdigit() else k: v for k, v in raw_sp.items()}
        for _prof in seller_profiles.values():
            _prof.setdefault('legal_status',       None)
            _prof.setdefault('stir',               None)
            _prof.setdefault('bank_account',       None)
            _prof.setdefault('bank_name',          None)
            _prof.setdefault('bank_mfo',           None)
            _prof.setdefault('director_name',      None)
            _prof.setdefault('legal_completed_at', None)
        # Migration: ensure onboarding_status default for shops saved before the field existed
        for _shops in seller_shops.values():
            for _shop in _shops:
                _shop.setdefault('onboarding_status', 'active')
                # Billz fields (Phase 1 onboarding)
                _shop.setdefault('billz_secret_token', None)
                _shop.setdefault('billz_shop_id', '')
                _shop.setdefault('billz_shop_name', '')
                _shop.setdefault('billz_connected_at', None)
                _shop.setdefault('billz_global_solo_discount', 10)
                _shop.setdefault('billz_global_group_discount', 20)
        pending_moderator_codes= data.get('pending_moderator_codes', {})
        referrals              = data.get('referrals', {})
        raw_rm                 = data.get('referral_map', {})
        referral_map           = {int(k) if str(k).isdigit() else k: v for k, v in raw_rm.items()}
        raw_sp                 = data.get('seller_products', {})
        seller_products        = {int(k) if k.isdigit() else k: v for k, v in raw_sp.items()}
        raw_cu = data.get('customers', {})
        customers = {int(k) if str(k).isdigit() else k: v for k, v in raw_cu.items()}
        raw_lv = data.get('lives', {})
        lives = {k: v for k, v in raw_lv.items()}
        logging.info(f"Data loaded: {len(products)} products, {len(orders)} orders")
        print(f"[JOYNSHOP] Data loaded: {len(products)} products, {len(seller_shops)} shops, {len(orders)} orders")
    except Exception as e:
        logging.error(f"load_data error: {e}", exc_info=True)
        print(f"[JOYNSHOP] load_data ERROR: {e}")

# ─── CLICK PAYMENT HELPERS ──────────────────────────────────────────
def send_invoice(cid, title, description, payload, amount, photo_url=None):
    """Telegram Bot Payments orqali Click invoice yuborish"""
    if not CLICK_TOKEN:
        return None
    data = {
        'chat_id':          cid,
        'title':            title,
        'description':      description,
        'payload':          payload,
        'provider_token':   CLICK_TOKEN,
        'currency':         'UZS',
        'prices':           json.dumps([{'label': title, 'amount': amount * 100}]),
        'need_name':        True,
        'need_phone_number': True,
        'need_shipping_address': False,
        'is_flexible':      False,
    }
    if photo_url:
        data['photo_url'] = photo_url
        data['photo_size'] = 512
    return api('sendInvoice', data, BUYER_TOKEN)

def answer_pre_checkout(query_id, ok=True, error=None, token=None):
    """PreCheckoutQuery ga javob berish"""
    data = {'pre_checkout_query_id': query_id, 'ok': ok}
    if error:
        data['error_message'] = error
    return api('answerPreCheckoutQuery', data, token or SELLER_TOKEN)

# ─── CRM HELPER ─────────────────────────────────────────────────────
def update_customer(seller_id, user_id, user_name, amount, product_name,
                    source='order', phone='', username=''):
    """Sotuvchining CRM bazasini yangilash.
    phone/username — yangi mijozlarda saqlanadi. Eski mijozlar uchun bo'sh string
    bo'lsa, mavjud qiymat saqlanadi.
    """
    sid = str(seller_id)
    uid = str(user_id)
    if sid not in customers:
        customers[sid] = {}
    if uid not in customers[sid]:
        customers[sid][uid] = {
            'name':         user_name,
            'user_id':      user_id,
            'total_orders': 0,
            'total_spent':  0,
            'orders':       [],
            'first_order':  datetime.now().strftime('%d.%m.%Y'),
            'last_order':   datetime.now().strftime('%d.%m.%Y'),
            'source':       source,
            'tags':         [],
            'phone':        '',
            'username':     '',
        }
    cust = customers[sid][uid]
    cust['name']         = user_name
    cust['total_orders'] += 1
    cust['total_spent']  += amount
    cust['last_order']   = datetime.now().strftime('%d.%m.%Y')
    # Yangi qiymat berilgan bo'lsa yangilanadi, aks holda mavjud saqlanadi
    if phone:
        cust['phone'] = phone
    if username:
        cust['username'] = username.lstrip('@')
    cust['orders'].append({
        'product': product_name,
        'amount':  amount,
        'date':    datetime.now().strftime('%d.%m.%Y %H:%M')
    })
    # Faqat oxirgi 20 ta buyurtmani saqlash
    if len(cust['orders']) > 20:
        cust['orders'] = cust['orders'][-20:]
    save_data()

# ─── HELPERS ────────────────────────────────────────────────────────
def api(method, data, token=None):
    url = f'https://api.telegram.org/bot{token or BUYER_TOKEN}/{method}'
    return requests.post(url, json=data).json()

def send(cid, text, kb=None, parse_mode='HTML', token=None):
    d = {'chat_id': cid, 'text': text, 'parse_mode': parse_mode}
    if kb: d['reply_markup'] = json.dumps(kb)
    return api('sendMessage', d, token)

def send_seller(cid, text, kb=None):
    return send(cid, text, kb, token=SELLER_TOKEN)

def send_buyer(cid, text, kb=None):
    return send(cid, text, kb, token=BUYER_TOKEN)

def edit_message(cid, mid, text, kb=None, token=None):
    token = token or SELLER_TOKEN
    d = {'chat_id': cid, 'message_id': mid, 'text': text, 'parse_mode': 'HTML'}
    if kb: d['reply_markup'] = json.dumps(kb)
    return requests.post(f'https://api.telegram.org/bot{token}/editMessageText', json=d).json()

def send_or_edit_seller(cid, text, kb=None, state=None):
    mid = state.get('ob_msg_id') if state else None
    if mid:
        r = edit_message(cid, mid, text, kb)
        if r and r.get('ok'):
            return r
    r = send(cid, text, kb)
    if r and state is not None:
        result = r.get('result', {})
        if result.get('message_id'):
            state['ob_msg_id'] = result['message_id']
    return r

def edit_caption(cid, mid, caption, kb=None):
    d = {'chat_id': cid, 'message_id': mid, 'caption': caption, 'parse_mode': 'HTML'}
    if kb: d['reply_markup'] = json.dumps(kb)
    api('editMessageCaption', d)

def answer_cb(cbid, text='', alert=False, token=None):
    api('answerCallbackQuery', {'callback_query_id': cbid, 'text': text, 'show_alert': alert}, token)

def fmt(n):
    return f"{int(n):,}"

def gen_code():
    return 'JS-' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def bar(count, min_g):
    return '🟢' * count + '⚪️' * (min_g - count)

# ─── POST CAPTION ────────────────────────────────────────────────────
def strip_html(text):
    """HTML teglarini olib tashlash — invoice description uchun"""
    import re
    return re.sub(r'<[^>]+>', '', text)

def invoice_description(p, pid):
    """Invoice uchun toza matn (HTML yo'q)"""
    count     = len(groups.get(pid, []))
    min_g     = p['min_group']
    orig      = p['original_price']
    solo      = p.get('solo_price', 0)
    group     = p['group_price']
    sale_type = p.get('sale_type', 'both')
    solo_disc = round((orig - solo) / orig * 100) if solo and orig else 0
    grp_disc  = round((orig - group) / orig * 100) if orig else 0
    lines = []
    if sale_type == 'solo':
        if p.get('description'):
            lines.append(p['description'][:100])
        lines.append(f"🏪 {p.get('shop_name','')} | {p.get('contact','')}")
        return "\n".join(lines)[:255]
    elif sale_type == 'group':
        lines.append(f"👥 Guruh narxi: {fmt(group)} so'm (-{grp_disc}%)")
        lines.append(f"👥 Guruh: {count}/{min_g} • Kerak: {max(0, min_g-count)} kishi")
    else:
        if solo:
            lines.append(f"👤 Yakka: {fmt(solo)} so'm (-{solo_disc}%)")
        lines.append(f"👥 Guruh: {fmt(group)} so'm (-{grp_disc}%)")
        lines.append(f"👥 Guruh: {count}/{min_g} • Kerak: {max(0, min_g-count)} kishi")
    lines.append(f"🕐 {p.get('deadline','')}")
    if p.get('description'):
        lines.append(p['description'][:100])
    lines.append(f"🏪 {p.get('shop_name','')} | {p.get('contact','')}")
    return "\n".join(lines)[:255]

def post_caption(p, pid):
    count     = len(groups.get(pid, []))
    min_g     = p['min_group']
    orig      = p['original_price']
    solo      = p.get('solo_price', 0)
    group     = p['group_price']
    solo_disc = round((orig - solo) / orig * 100) if solo and orig else 0
    grp_disc  = round((orig - group) / orig * 100) if orig else 0
    status    = '\U0001f525' if count < min_g else '\u2705'
    sale_type = p.get('sale_type', 'both')

    lines = [f"<b>{p['name']}</b>\n"]

    if sale_type == 'solo':
        if orig and orig != solo:
            lines.append(f"\U0001f4b0 Asl narx: <s>{fmt(orig)} so'm</s>")
        lines.append(f"\U0001f6d2 <b>{fmt(solo)} so'm</b>")
    elif sale_type == 'group':
        if orig and orig != group:
            lines.append(f"\U0001f4b0 Asl narx: <s>{fmt(orig)} so'm</s>")
        lines.append(f"\U0001f465 Guruh narxi: <b>{fmt(group)} so'm</b>  <i>(-{grp_disc}%)</i>")
        lines.append(f"\U0001f465 Guruh: {count}/{min_g} {status}")
        lines.append(f"\u23f3 Kerak: {max(0, min_g - count)} kishi")
    else:
        if orig:
            lines.append(f"\U0001f4b0 Asl narx: <s>{fmt(orig)} so'm</s>")
        if solo:
            lines.append(f"\U0001f464 Yakka: <b>{fmt(solo)} so'm</b>  <i>(-{solo_disc}%)</i>")
        lines.append(f"\U0001f465 Guruh: <b>{fmt(group)} so'm</b>  <i>(-{grp_disc}%)</i>")
        lines.append(f"\U0001f465 Guruh: {count}/{min_g} {status}")
        lines.append(f"\u23f3 Kerak: {max(0, min_g - count)} kishi")

    if sale_type != 'solo':
        lines.append(f"\U0001f550 {p.get('deadline','')}")
    if p.get('description'):
        lines.append(f"\n📝 {p['description']}")
    if p.get('variants'):
        lines.append(f"🎨 {', '.join(p['variants'])}")

    contact = p.get('contact', '')
    phone2  = p.get('phone2', '')
    address = p.get('address', '')
    social  = p.get('social', {})

    shop_line = f"\n🏪 <b>{p['shop_name']}</b>"
    if contact:
        shop_line += f"  |  📞 {contact}"
    lines.append(shop_line)
    if phone2:
        lines.append(f"📱 {phone2}")
    if address:
        lines.append(f"📍 {address}")
    if social:
        icons = {'instagram':'📸','telegram':'✈️','youtube':'▶️','website':'🌐','tiktok':'🎵'}
        for k, v in social.items():
            icon = icons.get(k.lower(), '🔗')
            lines.append(f"{icon} {k.capitalize()}: {v}")

    return "\n".join(lines)

def join_kb(pid, count, min_g, has_solo=False, sale_type='both'):
    """Kanal post tugmalari — Joynshop sayt to'lov sahifasiga yo'naltiradi."""
    if count >= min_g:
        return {'inline_keyboard': [[{'text': "✅ Guruh to'ldi!", 'url': APP_URL or f'https://t.me/{BUYER_BOT_USERNAME}'}]]}
    kb = []
    # Sayt to'lov sahifasi — joynshop.uz/pay/PID?type=solo|group
    base = (APP_URL or '').rstrip('/') + '/pay'
    if sale_type in ('solo', 'both') and has_solo:
        kb.append([{'text': "🛒 Sotib olish (yakka)",
                    'url': f'{base}/{pid}?type=solo'}])
    if sale_type in ('group', 'both'):
        kb.append([{'text': f"👥 Guruhga qo'shilish ({count}/{min_g})",
                    'url': f'{base}/{pid}?type=group'}])
    if not kb:
        kb.append([{'text': "🛍 Xarid qilish",
                    'url': f'{base}/{pid}?type=group'}])
    return kb_inline(kb)

def kb_inline(rows):
    return {'inline_keyboard': rows}

# ─── CHEK ────────────────────────────────────────────────────────────
def build_check(order_code, order):
    p            = products.get(order['product_id'], {})
    sale_type    = '👤 Yakka' if order.get('type') == 'solo' else '👥 Guruh'
    variant_line = f"\n🎨 {order['variant']}" if order.get('variant') else ''
    return (
        f"🧾 <b>JOYNSHOP CHEKI</b>\n"
        f"━━━━━━━━━━━━━━━\n"
        f"🏪 {p.get('shop_name', 'Sotuvchi')}\n"
        f"📦 {p.get('name', '')}{variant_line}\n"
        f"🛒 {sale_type} sotuv\n"
        f"💰 {fmt(order['amount'])} so'm\n"
        f"📅 {order.get('created', '')}\n"
        f"🆔 #{order_code}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"✅ <b>Tasdiqlandi</b>\n"
        f"🔒 Joynshop kafolati ostida"
    )

# ─── XARIDOR PROFILI ─────────────────────────────────────────────────
def get_profile(uid):
    if uid not in buyer_profiles:
        buyer_profiles[uid] = {
            'total_orders': 0, 'total_saved': 0,
            'groups_joined': 0, 'cashback': 0, 'referrals': 0
        }
    return buyer_profiles[uid]

def update_profile(uid, amount, original_price, is_group=False):
    p = get_profile(uid)
    p['total_orders'] += 1
    p['total_saved']  += (original_price - amount)
    if is_group: p['groups_joined'] += 1
    p['cashback'] += int(amount * 0.02)

# ─── EXPIRE ──────────────────────────────────────────────────────────
def notify_group_filled(pid):
    """Guruh to'lganda barcha a'zolarga, sotuvchiga va kanalga xabar yuboradi."""
    p = products.get(pid)
    if not p: return
    count   = len(groups.get(pid, []))
    min_g   = p.get('min_group', 3)
    sid     = p.get('seller_id')
    name    = p.get('name', '')
    shop    = p.get('shop_name', '')
    contact = p.get('contact', '')
    total      = count * p.get('group_price', 0)
    commission = int(total * COMMISSION_RATE)
    payout     = total - commission

    # 1. Barcha guruh a'zolariga
    for wuid in groups.get(pid, []):
        try:
            send_buyer(wuid,
                f"\U0001f525 <b>GURUH TO'LDI!</b>\n\n"
                f"\U0001f3ea {shop}\n"
                f"\U0001f4e6 {name}\n"
                f"\U0001f4de Sotuvchi: {contact}\n\n"
                f"\u2705 Buyurtmangiz tasdiqlandi! Sotuvchi siz bilan bog'lanadi.",
                {'inline_keyboard': [[
                    {'text': '\u2b50 Baho bering', 'callback_data': f'rate_start_{pid}'},
                    {'text': '\u21a9\ufe0f Qaytarish', 'callback_data': f'refund_{pid}'},
                ]]}
            )
        except: pass

    # 2. Wishlist da bor lekin guruhga qo'shilmagan odamlarga
    group_members = set(groups.get(pid, []))
    for wl_uid, wl in wishlists.items():
        if pid in wl and wl_uid not in group_members:
            try:
                send_buyer(wl_uid,
                    f"\U0001f614 <b>Siz qiziqgan guruh to'ldi!</b>\n\n"
                    f"\U0001f4e6 {name}\n"
                    f"\U0001f3ea {shop}\n\n"
                    f"Keyingi guruhga qo'shilish uchun kuzatib boring:",
                    {'inline_keyboard': [[
                        {'text': "\U0001f514 Kuzatish",
                         'url': f'https://t.me/{BUYER_BOT_USERNAME}?start=join_{pid}'}
                    ]]}
                )
            except: pass

    # 3. Sotuvchiga moliyaviy xabar
    if sid:
        send_seller(sid,
            f"\U0001f389 <b>GURUH TO'LDI!</b>\n\n"
            f"\U0001f4e6 {name}\n"
            f"\U0001f465 {count}/{min_g} kishi qo'shildi!\n\n"
            f"\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\n"
            f"\U0001f4b0 Jami sotuv: <b>{fmt(total)} so'm</b>\n"
            f"\U0001f4ca Komissiya (5%): <b>{fmt(commission)} so'm</b>\n"
            f"\u2705 Sizga to'lanadi: <b>{fmt(payout)} so'm</b>\n"
            f"\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\n\n"
            f"\U0001f4b3 Karta raqamingizni yuboring,\n"
            f"pul 24 soat ichida o'tkaziladi."
        )
        if ADMIN_ID:
            send_seller(ADMIN_ID,
                f"\U0001f4b0 <b>To'lov kerak!</b>\n\n"
                f"\U0001f4e6 {name}\n"
                f"\U0001f464 Sotuvchi ID: <code>{sid}</code>\n"
                f"\U0001f4b5 O'tkazish kerak: <b>{fmt(payout)} so'm</b>\n"
                f"\U0001f4ca Komissiya: <b>{fmt(commission)} so'm</b>"
            )

    # 4. Kanal postini yangilash
    channel_cid = p.get('channel_chat_id')
    channel_mid = p.get('channel_message_id')
    if channel_cid and channel_mid:
        try:
            edit_caption(channel_cid, channel_mid,
                post_caption(p, pid),
                {'inline_keyboard': [[
                    {'text': "\u2705 Guruh to'ldi!",
                     'url': f'https://t.me/{BUYER_BOT_USERNAME}'}
                ]]}
            )
        except: pass


def expire_product(pid):
    p = products.get(pid)
    if not p or p.get('status') == 'closed': return
    products[pid]['status'] = 'closed'
    count = len(groups.get(pid, []))
    sid   = p.get('seller_id')

    for uid in groups.get(pid, []):
        try:
            if count >= p['min_group']:
                send_buyer(uid,
                    f"🎉 <b>Guruh to'ldi!</b>\n\n"
                    f"<b>{p['name']}</b>\n"
                    f"📞 Sotuvchi: {p.get('contact')}"
                )
            else:
                send_buyer(uid,
                    f"😔 <b>Guruh to'lmadi</b>\n\n"
                    f"<b>{p['name']}</b>\n"
                    f"💰 To'lovingiz 24 soat ichida qaytariladi."
                )
        except: pass

    if sid:
        if count >= p['min_group']:
            # notify_group_filled barcha xabarlarni yuboradi
            notify_group_filled(pid)
        else:
            send_seller(sid,
                f"😔 <b>Guruh to'lmadi</b>\n\n"
                f"<b>{p['name']}</b>\n"
                f"👥 {count}/{p['min_group']} kishi\n\n"
                f"Yangi mahsulot qo'shib ko'ring.",
                {'inline_keyboard': [[{'text': "➕ Yangi mahsulot qo'shish", 'callback_data': 'menu_addproduct'}]]}
            )

# ─── REMINDER & LIVE UPDATE ──────────────────────────────────────────
def reminder_loop():
    while True:
        time.sleep(1800)
        try:
            now = datetime.now()
            for pid, p in list(products.items()):
                if p.get('status') == 'closed': continue
                ddt = p.get('deadline_dt')
                if not ddt: continue
                deadline  = datetime.strptime(ddt, '%Y-%m-%d %H:%M')
                remaining = (deadline - now).total_seconds()
                count     = len(groups.get(pid, []))
                needed    = p['min_group'] - count
                if remaining <= 0:
                    expire_product(pid)
                    continue
                hours = remaining / 3600
                if needed > 0 and (11.5 <= hours <= 12.5 or 1.5 <= hours <= 2.5):
                    for uid in groups.get(pid, []):
                        try:
                            send_buyer(uid,
                                f"⚡️ <b>SHOSHILING!</b>\n\n"
                                f"<b>{p['name']}</b>\n"
                                f"{needed} kishi kerak!\n"
                                f"⏰ {int(hours)} soat qoldi!"
                            )
                        except: pass
                    sid = p.get('seller_id')
                    if sid:
                        send_seller(sid,
                            f"📢 <b>Eslatma!</b>\n\n"
                            f"<b>{p['name']}</b>\n"
                            f"👥 {count}/{p['min_group']} kishi\n"
                            f"⏰ {int(hours)} soat qoldi\n\n"
                            f"Kanalda qayta e'lon qiling:",
                            {'inline_keyboard': [[{'text': "📢 Qayta e'lon qilish", 'callback_data': f'boost_{pid}'}]]}
                        )
        except Exception as e:
            logging.error(f"Reminder error: {e}")

def live_update_loop():
    while True:
        time.sleep(30)
        try:
            for pid, p in list(products.items()):
                if p.get('status') == 'closed': continue
                cid = p.get('channel_chat_id')
                mid = p.get('channel_message_id')
                if not cid or not mid: continue
                count = len(groups.get(pid, []))
                try:
                    edit_caption(cid, mid,
                        post_caption(p, pid),
                        join_kb(pid, count, p['min_group'], has_solo=bool(p.get('solo_price')))
                    )
                except: pass
        except Exception as e:
            logging.error(f"Live update: {e}")

threading.Thread(target=reminder_loop, daemon=True).start()
threading.Thread(target=live_update_loop, daemon=True).start()

# ─── SPAM ────────────────────────────────────────────────────────────
PROD_ALLOWED_CBS = {
    'prod_photo_done', 'prod_skip_desc', 'prod_add_desc', 'prod_add_solo', 'prod_add_variants',
    'prod_mxik_manual_btn', 'prod_mxik_again', 'prod_mxik_confirm', 'prod_mxik_skip',
    'prod_confirm_publish', 'prod_continue', 'prod_restart',
    'prod_deadline_24', 'prod_deadline_48', 'prod_deadline_72', 'prod_deadline_168',
    'ob_skip_phone2', 'ob_skip_address', 'ob_skip_social', 'ob_keep_phone',
    'ob_delivery_deliver', 'ob_delivery_pickup', 'ob_delivery_both',
    'edit_shop_0', 'edit_shop_1', 'edit_shop_2',
    'back_menu', 'noop', 'menu_mycustomers', 'menu_export', 'live_cancel', 'live_start',
}

PROD_BLOCKED_TEXTS = {
    "➕ Mahsulot qo'shish", '/addproduct',
    '📦 Mahsulotlarim', '/myproducts',
    '📋 Buyurtmalar', '/myorders',
    '📊 Statistika', '/mystats',
    "📢 Do'konlarim", '/mychannels',
    '❓ Yordam', '/help',
}

def is_prod_in_progress(uid):
    s = seller_state.get(uid)
    if not s: return False
    prod_steps = {'prod_name','prod_category','prod_sale_type','prod_photo','prod_price','prod_min_group',
                  'prod_desc','prod_confirm','prod_edit_desc','prod_edit_solo','prod_edit_variants',
                  'prod_mxik_search','prod_mxik_manual','prod_mxik_confirm_state'}
    return s.get('step') in prod_steps

def get_prod_progress_text(uid):
    s = seller_state.get(uid, {})
    step = s.get('step','')
    name = s.get('name', s.get('prod_name', '—'))
    step_names = {
        'prod_name':      '1/7 — Nom',
        'prod_category':  '2/7 — Kategoriya',
        'prod_sale_type': '3/7 — Sotuv turi',
        'prod_photo':     '2/5 — Rasmlar',
        'prod_price':     '3/5 — Narx',
        'prod_min_group': '4/5 — Guruh soni',
        'prod_desc':      '5/5 — Tavsif',
        'prod_confirm':   '6/6 — Tasdiqlash',
        'prod_edit_desc': '5/5 — Tavsif',
        'prod_edit_solo': '5/5 — Yakka narx',
        'prod_edit_variants': '5/5 — Variantlar',
    }
    step_label = step_names.get(step, step)
    return (
        f"📋 Siz allaqachon mahsulot qo'shmoqdasiz:\n\n"
        f"📦 <b>{name}</b>\n"
        f"📍 Holat: {step_label}\n\n"
        f"Davom etasizmi yoki yangi boshlamoqchimisiz?"
    )

def is_spam(text):
    if not text: return False
    lower = text.lower()
    return lower.count('http') + lower.count('t.me/') > 1

def moderate_chat(msg):
    cid = msg['chat']['id']
    mid = msg['message_id']
    if is_spam(msg.get('text', '')):
        api('deleteMessage', {'chat_id': cid, 'message_id': mid})
        send_buyer(cid, "⚠️ Spam xabar o'chirildi.")
        return True
    return False

# ══════════════════════════════════════════════════════════════════════
#  SELLER WEBHOOK
# ══════════════════════════════════════════════════════════════════════
@app.route('/seller/webhook', methods=['POST'])
def seller_webhook():
    data = request.json
    if 'callback_query'   in data: seller_handle_cb(data['callback_query'])
    elif 'pre_checkout_query' in data: handle_pre_checkout(data['pre_checkout_query'])
    elif 'message'        in data:
        msg = data['message']
        if 'successful_payment' in msg:
            handle_successful_payment(msg)
        else:
            seller_handle_msg(msg)
    return 'ok'

def seller_handle_cb(cb):
    cbid = cb['id']
    uid  = cb['from']['id']
    d    = cb['data']

    if d == 'noop':
        answer_cb(cbid); return

    if is_prod_in_progress(uid) and d not in PROD_ALLOWED_CBS and not d.startswith('prod_') and not d.startswith('ob_') and not d.startswith('edit_shop_') and not d.startswith('cat_') and not d.startswith('sale_type_') and not d.startswith('crm_') and not d.startswith('live_'):
        answer_cb(cbid)
        send_seller(uid, get_prod_progress_text(uid),
            {'inline_keyboard': [
                [{'text': "▶️ Davom etish", 'callback_data': 'prod_continue'}],
                [{'text': "🗑 Bekor qilish", 'callback_data': 'prod_restart'}],
            ]}
        )
        return

    if d == 'add_new_shop':
        answer_cb(cbid)
        seller_state[uid] = {'step': 'ob_shop_name', 'adding_shop': True}
        send_seller(uid, "🏪 <b>Yangi do'kon</b>\n\n<b>1/4</b> Do'kon nomini yozing:")
        return

    if d.startswith('ob_delivery_'):
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        delivery_map = {'ob_delivery_deliver': 'deliver', 'ob_delivery_pickup': 'pickup', 'ob_delivery_both': 'both'}
        s['ob_delivery'] = delivery_map.get(d, 'pickup')
        answer_cb(cbid)
        # Yuridik ma'lumot mavjud bo'lsa to'g'ridan kanalga, aks holda yuridik flow'ni boshlaymiz
        if seller_has_legal(uid):
            s['step'] = 'ob_channel'
            send_or_edit_seller(uid,
                "📢 Telegram kanal username:\n<i>@mening_kanalim</i>\n\n"
                "⚠️ Seller bot kanalga <b>admin</b> sifatida qo'shilgan bo'lishi kerak!",
                state=s)
        else:
            # Mavjud onboarding state'ni saqlab leg_after='channel' bilan flow'ga o'tamiz
            s['step']      = 'leg_status'
            s['leg_after'] = 'channel'
            s.pop('ob_msg_id', None)
            send_seller(uid,
                "📋 <b>Yuridik ma'lumotlar — qadam 1/6</b>\n\n"
                "Yuridik statusingiz qanday?\n\n"
                "<i>Bu Payme split to'lov va fiskal chek uchun kerak. "
                "Bekor qilish: /cancel</i>",
                {'inline_keyboard': [
                    [{'text': "👤 YaTT (yakka tartibdagi tadbirkor)", 'callback_data': 'leg_pick_yatt'}],
                    [{'text': "🏢 MChJ (mas'uliyati cheklangan jamiyat)", 'callback_data': 'leg_pick_mchj'}],
                ]})
        return

    if d in ('ob_skip_phone2', 'ob_keep_phone'):
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        if d == 'ob_keep_phone':
            idx = s.get('edit_shop_idx', 0)
            shops = seller_shops.get(uid, [])
            s['ob_phone'] = shops[idx].get('phone', '') if idx < len(shops) else s.get('ob_phone', '')
        else:
            s['ob_phone2'] = ''
        s['step'] = 'ob_address'; answer_cb(cbid)
        send_or_edit_seller(uid,
            "📍 Do'kon manzili (ixtiyoriy):\n<i>Toshkent, Chilonzor, 3-mavze</i>",
            {'inline_keyboard': [[{'text': "⏭ O'tkazib yuborish", 'callback_data': 'ob_skip_address'}]]},
            state=s)
        return

    if d == 'ob_skip_address':
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        s['ob_address'] = ''; s['step'] = 'ob_social'; answer_cb(cbid)
        send_or_edit_seller(uid,
            "🌐 Ijtimoiy tarmoqlar (ixtiyoriy):\n"
            "<code>instagram: @dokon_uz\ntelegram: @kanal\nwebsite: dokon.uz</code>",
            {'inline_keyboard': [[{'text': "⏭ O'tkazib yuborish", 'callback_data': 'ob_skip_social'}]]},
            state=s)
        return

    if d == 'ob_skip_social':
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        s['ob_social'] = {}; s['step'] = 'ob_delivery'; answer_cb(cbid)
        send_or_edit_seller(uid,
            "🚚 Yetkazib berish turini tanlang:",
            {'inline_keyboard': [
                [{'text': "🚚 Yetkazib beraman",   'callback_data': 'ob_delivery_deliver'}],
                [{'text': "🏪 Xaridor olib ketadi", 'callback_data': 'ob_delivery_pickup'}],
                [{'text': "🚚🏪 Ikkalasi ham",       'callback_data': 'ob_delivery_both'}],
            ]},
            state=s)
        return

    if d.startswith('edit_shop_'):
        idx = int(d.split('_')[2])
        shops = seller_shops.get(uid, [])
        if idx >= len(shops): answer_cb(cbid); return
        shop = shops[idx]; answer_cb(cbid)
        social_text = '\n'.join(f"🔗 {k}: {v}" for k, v in shop.get('social', {}).items())
        send_seller(uid,
            f"✏️ <b>Do'kon tahrirlash</b>\n\n"
            f"🏪 {shop['name']}\n📞 {shop['phone']}"
            f"{chr(10)+'📱 '+shop.get('phone2','') if shop.get('phone2') else ''}"
            f"{chr(10)+'📍 '+shop.get('address','') if shop.get('address') else ''}"
            f"{chr(10)+social_text if social_text else ''}\n📢 {shop.get('channel','')}",
            {'inline_keyboard': [
                [{'text': "✏️ Qayta to'ldirish",          'callback_data': f'edit_shop_full_{idx}'}],
                [{'text': "📱 Tel qo'shish/o'zgartirish", 'callback_data': f'edit_shop_phone_{idx}'}],
                [{'text': "📍 Manzil",                    'callback_data': f'edit_shop_address_{idx}'}],
                [{'text': "🌐 Ijtimoiy tarmoqlar",         'callback_data': f'edit_shop_social_{idx}'}],
                [{'text': "❌ Bekor",                      'callback_data': 'noop'}],
            ]})
        return

    if d.startswith('edit_shop_full_'):
        idx = int(d.split('_')[3])
        answer_cb(cbid)
        seller_state[uid] = {'step': 'ob_shop_name', 'edit_shop_idx': idx}
        shops = seller_shops.get(uid, [])
        shop = shops[idx] if idx < len(shops) else {}
        send_seller(uid,
            f"✏️ Do'kon nomini kiriting:\n<i>Hozir: {shop.get('name', '')}</i>")
        return

    if d.startswith('edit_shop_phone_'):
        idx = int(d.split('_')[3])
        answer_cb(cbid)
        seller_state[uid] = {'step': 'edit_phone_direct', 'edit_shop_idx': idx,
                             'ob_shop_name': seller_shops[uid][idx]['name'],
                             'ob_delivery':  seller_shops[uid][idx].get('delivery','pickup'),
                             'ob_channel':   seller_shops[uid][idx].get('channel',''),
                             'ob_address':   seller_shops[uid][idx].get('address',''),
                             'ob_social':    seller_shops[uid][idx].get('social',{})}
        send_seller(uid, "📞 Yangi telefon raqam:\n<i>+998XXXXXXXXX</i>")
        return

    if d.startswith('edit_shop_address_'):
        idx = int(d.split('_')[3])
        answer_cb(cbid)
        seller_state[uid] = {'step': 'edit_address_direct', 'edit_shop_idx': idx,
                             'ob_shop_name': seller_shops[uid][idx]['name'],
                             'ob_phone':     seller_shops[uid][idx].get('phone',''),
                             'ob_phone2':    seller_shops[uid][idx].get('phone2',''),
                             'ob_delivery':  seller_shops[uid][idx].get('delivery','pickup'),
                             'ob_channel':   seller_shops[uid][idx].get('channel',''),
                             'ob_social':    seller_shops[uid][idx].get('social',{})}
        send_seller(uid, "📍 Do'kon manzili:\n<i>Toshkent, Chilonzor, 3-mavze</i>")
        return

    if d.startswith('edit_shop_social_'):
        idx = int(d.split('_')[3])
        answer_cb(cbid)
        seller_state[uid] = {'step': 'edit_social_direct', 'edit_shop_idx': idx,
                             'ob_shop_name': seller_shops[uid][idx]['name'],
                             'ob_phone':     seller_shops[uid][idx].get('phone',''),
                             'ob_phone2':    seller_shops[uid][idx].get('phone2',''),
                             'ob_address':   seller_shops[uid][idx].get('address',''),
                             'ob_delivery':  seller_shops[uid][idx].get('delivery','pickup'),
                             'ob_channel':   seller_shops[uid][idx].get('channel','')}
        send_seller(uid,
            "🌐 Ijtimoiy tarmoqlar:\n\n"
            "<code>instagram: @dokon_uz\ntelegram: @kanal\nwebsite: dokon.uz\nyoutube: @kanal</code>\n\n"
            "<i>Faqat mavjudlarini yozing</i>")
        return

    if d.startswith('sel_shop_'):
        idx = int(d.split('_')[2])
        shops = seller_shops.get(uid, [])
        if idx >= len(shops): answer_cb(cbid, '❌'); return
        shop = shops[idx]; answer_cb(cbid)
        seller_state[uid] = {
            'step': 'prod_name', 'shop_idx': idx,
            'shop_name':      shop['name'],
            'contact':        shop['phone'],
            'delivery_type':  shop.get('delivery','pickup'),
            'seller_channel': shop.get('channel',''),
        }
        send_seller(uid,
            f"📦 <b>{shop['name']}</b> uchun yangi mahsulot\n\n"
            "<b>1/4</b> Mahsulot nomini yozing:")
        return

    if d.startswith('cat_'):
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s: return
        cat = d[4:]
        s['category'] = cat; s['step'] = 'prod_sale_type'
        send_seller(uid,
            f"✅ Kategoriya: <b>{cat}</b>\n\n"
            "<b>3/7</b> Sotuv turini tanlang:",
            {'inline_keyboard': [
                [{'text': "👥 Faqat guruhli sotuv", 'callback_data': 'sale_type_group'}],
                [{'text': "👤 Faqat yakka sotuv",   'callback_data': 'sale_type_solo'}],
                [{'text': "👥+👤 Ikkalasi ham",      'callback_data': 'sale_type_both'}],
            ]}
        )
        return

    if d in ('sale_type_group', 'sale_type_solo', 'sale_type_both'):
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s: return
        s['sale_type'] = d.replace('sale_type_', '')
        s['step'] = 'prod_photo'; s['photo_ids'] = []; s['photo_urls'] = []
        type_label = {'group': '👥 Guruhli', 'solo': '👤 Yakka', 'both': '👥+👤 Ikkalasi'}
        send_seller(uid,
            f"✅ {type_label.get(s['sale_type'])} sotuv tanlandi\n\n"
            "<b>4/7</b> Mahsulot rasmini yuboring 📸\n<i>1-5 ta rasm yuborishingiz mumkin</i>"
        )
        return

    if d == 'prod_photo_done':
        s = seller_state.get(uid)
        if not s or not s.get('photo_ids'):
            answer_cb(cbid, '❌ Rasm yo\'q'); return
        s['step'] = 'prod_price'; answer_cb(cbid)
        sale_type = s.get('sale_type', 'both')
        if sale_type == 'solo':
            price_hint = "<b>4/5</b> Yakka sotuv narxini kiriting (so'm):\n<code>850000</code>"
        else:
            price_hint = "<b>4/5</b> Narxlarni kiriting:\n<code>850000 / 550000</code>\n<i>asl narx / guruh narxi</i>"
        send_seller(uid, f"✅ {len(s['photo_ids'])} ta rasm.\n\n{price_hint}")
        return

    if d == 'prod_confirm_publish':
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        answer_cb(cbid)
        publish_product(uid, uid, s); return

    if d == 'prod_add_desc':
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        s['step'] = 'prod_edit_desc'; answer_cb(cbid)
        _send_or_edit_prod(uid, s,
            "📝 <b>Yangi tavsif yozing</b>\n\n"
            "Maksimum 300 belgi.\n\n"
            "<i>Bekor qilish: /cancel</i>", None)
        return

    if d == 'prod_add_solo':
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        s['step'] = 'prod_edit_solo'; answer_cb(cbid)
        _send_or_edit_prod(uid, s,
            "💰 <b>Yakka sotuv narxini yozing</b>\n\n"
            "Faqat raqam (so'm).\n\n"
            "<i>Bekor qilish: /cancel</i>", None)
        return

    if d.startswith('prod_deadline_'):
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        hours_map = {'prod_deadline_24': 24, 'prod_deadline_48': 48,
                     'prod_deadline_72': 72, 'prod_deadline_168': 168}
        h = hours_map.get(d, 48)
        s['deadline_hours'] = h
        answer_cb(cbid)
        shop = seller_shops.get(uid, [{}])[s.get('shop_idx', 0)]
        show_prod_confirm(uid, s, shop)
        return

    if d == 'prod_add_variants':
        s = seller_state.get(uid)
        if not s: answer_cb(cbid); return
        s['step'] = 'prod_edit_variants'; answer_cb(cbid)
        _send_or_edit_prod(uid, s,
            "🎨 <b>Variantlarni vergul bilan yozing</b>\n\n"
            "<i>Masalan: 38, 39, 40 yoki Qizil, Ko'k, Yashil</i>\n\n"
            "<i>Bekor qilish: /cancel</i>", None)
        return

    if d in ('start_addproduct', 'menu_addproduct'):
        answer_cb(cbid)
        if is_prod_in_progress(uid):
            send_seller(uid, get_prod_progress_text(uid),
                {'inline_keyboard': [
                    [{'text': "▶️ Davom etish",              'callback_data': 'prod_continue'}],
                    [{'text': "🗑 Bekor qilib, yangi boshlash", 'callback_data': 'prod_restart'}],
                ]}
            )
            return
        # Sotuvchining do'konlarini tekshirish
        shops = seller_shops.get(uid, [])
        if not shops:
            send_seller(uid, "❌ Avval do'kon yarating.\n\n/start yozing.")
            return
        if len(shops) == 1:
            # Bitta do'kon — to'g'ridan-to'g'ri
            seller_state[uid] = {'step': 'prod_name', 'shop_idx': 0}
            send_seller(uid, "📦 <b>Yangi mahsulot</b>\n\n1️⃣ Mahsulot nomini yozing:")
        else:
            # Bir nechta do'kon — tanlash
            kb = []
            for i, sh in enumerate(shops):
                shop_name = sh.get('name', f"Do'kon #{i+1}")
                ch        = sh.get('channel', '')
                label     = f"🏪 {shop_name}" + (f" ({ch})" if ch else "")
                kb.append([{'text': label, 'callback_data': f'prod_shop_{i}'}])
            kb.append([{'text': "❌ Bekor", 'callback_data': 'back_menu'}])
            send_seller(uid,
                "🏪 <b>Qaysi do'konga mahsulot qo'shasiz?</b>\n\n"
                f"Sizda {len(shops)} ta do'kon bor:",
                {'inline_keyboard': kb}
            )
        return

    if d.startswith('prod_shop_'):
        answer_cb(cbid)
        try:
            idx = int(d.split('_')[2])
        except:
            return
        shops = seller_shops.get(uid, [])
        if idx < 0 or idx >= len(shops):
            send_seller(uid, "❌ Noto'g'ri do'kon.")
            return
        seller_state[uid] = {'step': 'prod_name', 'shop_idx': idx}
        shop_name = shops[idx].get('name', '')
        send_seller(uid,
            f"📦 <b>Yangi mahsulot — {shop_name}</b>\n\n"
            f"1️⃣ Mahsulot nomini yozing:"
        )
        return

    if d == 'prod_continue':
        answer_cb(cbid)
        s = seller_state.get(uid, {})
        step = s.get('step','')
        step_msgs = {
            'prod_name':      "Mahsulot nomini yozing:",
            'prod_photo':     "Rasmlarni yuboring (1-5 ta):",
            'prod_price':     "Narxni yozing (asl/guruh):",
            'prod_min_group': "Minimal guruh sonini yozing (2-100):",
        }
        msg = step_msgs.get(step, "Davom eting:")
        send_seller(uid, f"✅ Davom etmoqdasiz\n\n{msg}")
        return

    if d == 'prod_skip_desc':
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s: return
        s['description'] = ''
        s['step'] = 'prod_mxik_search'
        send_seller(uid,
            "🔍 <b>MXIK kodi (ixtiyoriy)</b>\n\n"
            "Mahsulot nomini yoki kalit so'z kiriting yoki o'tkazib yuboring:\n"
            "<i>Masalan: krem, ko'ylak paxta, telefon</i>\n\n"
            "Bekor qilish: /cancel",
            {'inline_keyboard': [
                [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
            ]})
        return

    # ─── MXIK CALLBACKS ───
    if d == 'prod_mxik_again':
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s: return
        s['step'] = 'prod_mxik_search'
        s.pop('mxik_results', None)
        send_seller(uid,
            "🔍 Yangi qidiruv. Kalit so'zni kiriting:",
            {'inline_keyboard': [
                [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
            ]})
        return

    if d == 'prod_mxik_manual_btn':
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s: return
        s['step'] = 'prod_mxik_manual'
        send_seller(uid,
            "🔢 17 raqamli MXIK kodni kiriting:\n\n"
            "<i>Masalan: 03304011003000000</i>")
        return

    if d.startswith('prod_mxik_page_'):
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s or 'mxik_results' not in s: return
        try:
            pg = int(d[len('prod_mxik_page_'):])
        except ValueError:
            return
        render_mxik_results(uid, uid, s.get('mxik_keyword',''), s['mxik_results'], page=pg)
        return

    if d.startswith('prod_mxik_pick_'):
        s = seller_state.get(uid)
        if not s or 'mxik_results' not in s:
            answer_cb(cbid, "❌ Holat yo'qoldi"); return
        try:
            idx = int(d[len('prod_mxik_pick_'):])
        except ValueError:
            answer_cb(cbid); return
        results = s['mxik_results']
        if idx < 0 or idx >= len(results):
            answer_cb(cbid, "❌ Topilmadi"); return
        item = results[idx]
        s['mxik_code'] = item['code']
        s['mxik_name'] = item['name']
        s['step'] = 'prod_mxik_confirm_state'
        answer_cb(cbid)
        render_mxik_confirm(uid, uid, item['code'], item['name'], item.get('classify',''))
        return

    if d == 'prod_mxik_skip':
        s = seller_state.get(uid)
        if not s:
            answer_cb(cbid); return
        answer_cb(cbid, "⏭ O'tkazib yuborildi")
        s.pop('mxik_results', None)
        s.pop('mxik_keyword', None)
        after = s.get('mxik_after', 'create')

        if after == 'edit_pp':
            # Tahrirlash rejimi — joriy MXIK saqlanadi, hech narsa o'zgarmaydi
            pid = s.get('pp_pid')
            seller_state.pop(uid, None)
            send_seller(uid,
                "⏭ O'tkazib yuborildi. MXIK keyinroq qo'shilishi mumkin.",
                {'inline_keyboard': [
                    [{'text': "👁 Mahsulotni ko'rish", 'callback_data': f'mp_view_{pid}'}],
                    [{'text': "📦 Mahsulotlarim",     'callback_data': 'menu_myproducts'}],
                ]})
            return

        if after == 'bz_act':
            # Billz activation — MXIK siz solo narxga o'tamiz
            pid = s.get('bz_pid')
            p = products.get(pid) or {}
            new_state = {k: v for k, v in s.items()
                         if k.startswith('bz_') or k == 'bz_pid'}
            new_state['step'] = 'bz_act_solo'
            seller_state[uid] = new_state
            orig = int(p.get('original_price', 0) or 0)
            solo_disc = new_state.get('bz_suggested_solo') or max(1, int(orig * 0.9))
            send_seller(uid,
                f"⏭ MXIK o'tkazib yuborildi (keyinroq qo'shish mumkin).\n\n"
                f"💰 Asl narx: <b>{fmt(orig)} so'm</b>\n\n"
                f"<b>1/4</b> Yakka narxni yozing (so'm).\n"
                f"💡 Tavsiya: <b>{fmt(solo_disc)}</b> so'm\n\n"
                f"Bekor qilish: /cancel")
            return

        # Default — yangi mahsulot yaratish: prod_confirm'ga MXIK siz o'tamiz
        s.pop('mxik_code', None)
        s.pop('mxik_name', None)
        s['step'] = 'prod_confirm'
        shop = seller_shops.get(uid,[{}])[s.get('shop_idx',0)]
        show_prod_confirm(uid, s, shop)
        return

    if d == 'prod_mxik_confirm':
        s = seller_state.get(uid)
        if not s or s.get('step') != 'prod_mxik_confirm_state':
            answer_cb(cbid, "❌ Holat yo'qoldi"); return
        if not s.get('mxik_code'):
            answer_cb(cbid, "❌ MXIK tanlanmagan"); return
        answer_cb(cbid, "✅ Saqlandi")
        s.pop('mxik_results', None)
        s.pop('mxik_keyword', None)
        after = s.get('mxik_after', 'create')

        # ─── pp_edit rejimida: products[pid] ga to'g'ridan saqlaymiz ───
        if after == 'edit_pp':
            pid = s.get('pp_pid')
            p = products.get(pid)
            if p and p.get('seller_id') == uid:
                p['mxik_code'] = s['mxik_code']
                p['mxik_name'] = s['mxik_name']
                save_data()
                # Kanal post caption'ini yangilash shart emas — MXIK caption'da ko'rsatilmaydi
            seller_state.pop(uid, None)
            send_seller(uid,
                f"✅ MXIK saqlandi: <code>{s['mxik_code']}</code>",
                {'inline_keyboard': [
                    [{'text': "👁 Mahsulotni ko'rish", 'callback_data': f'mp_view_{pid}'}],
                    [{'text': "📦 Mahsulotlarim",     'callback_data': 'menu_myproducts'}],
                ]})
            return

        # ─── Billz activation rejimida: products[pid] ga saqlab bz_act_solo ga o'tamiz ───
        if after == 'bz_act':
            pid = s.get('bz_pid')
            p = products.get(pid)
            if p and p.get('seller_id') == uid:
                p['mxik_code'] = s['mxik_code']
                p['mxik_name'] = s['mxik_name']
                save_data()
            # Bz state'ni qayta tiklab bz_act_solo'ga o'tamiz
            new_state = {k: v for k, v in s.items()
                         if k.startswith('bz_') or k == 'bz_pid'}
            new_state['step'] = 'bz_act_solo'
            seller_state[uid] = new_state
            orig = int((p or {}).get('original_price', 0) or 0)
            solo_disc = new_state.get('bz_suggested_solo')
            if not solo_disc:
                # Re-derive — bz_activate_<pid> chaqirilganda saqlangan, lekin yo'q bo'lsa
                solo_disc = max(1, int(orig * 0.9))
            send_seller(uid,
                f"✅ MXIK saqlandi.\n\n"
                f"💰 Asl narx: <b>{fmt(orig)} so'm</b>\n\n"
                f"<b>1/4</b> Yakka narxni yozing (so'm).\n"
                f"💡 Tavsiya: <b>{fmt(solo_disc)}</b> so'm\n\n"
                f"Bekor qilish: /cancel")
            return

        # ─── Default (create rejimida): prod_confirm'ga o'tish ───
        s['step'] = 'prod_confirm'
        shop = seller_shops.get(uid,[{}])[s.get('shop_idx',0)]
        show_prod_confirm(uid, s, shop)
        return

    if d == 'prod_restart':
        answer_cb(cbid)
        seller_state.pop(uid, None)
        send_seller(uid, "🗑 Bekor qilindi. Yangi mahsulot boshlang:",
            {'inline_keyboard': [[{'text': "➕ Mahsulot qo'shish", 'callback_data': 'menu_addproduct'}]]}
        )
        return

    if d == 'menu_mystats':
        answer_cb(cbid)
        my = seller_products.get(uid, [])
        if not my:
            send_seller(uid, "📊 Statistika yo'q.\n\n/addproduct — mahsulot qo'shing!"); return
        revenue    = sum(o['amount'] for o in orders.values() if o.get('product_id') in my and o['status'] == 'confirmed')
        commission = int(revenue * COMMISSION_RATE)
        send_seller(uid,
            f"📊 <b>Sizning statistikangiz:</b>\n\n"
            f"📦 Jami mahsulot: {len(my)}\n"
            f"🔥 Aktiv: {sum(1 for pid in my if products.get(pid,{}).get('status')!='closed')}\n"
            f"✅ Muvaffaqiyatli guruh: {sum(1 for pid in my if len(groups.get(pid,[]))>=products.get(pid,{}).get('min_group',99))}\n"
            f"👥 Jami qo'shilgan: {sum(len(groups.get(pid,[])) for pid in my)}\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"💰 Jami sotuv: {fmt(revenue)} so'm\n"
            f"📊 Komissiya (5%): {fmt(commission)} so'm\n"
            f"✅ Sof daromad: {fmt(revenue-commission)} so'm",
            {'inline_keyboard': [
                [{'text': "📑 Excel eksport", 'callback_data': 'menu_export'}],
                [{'text': "🔙 Menyu",         'callback_data': 'back_menu'}],
            ]}
        )
        return

    if d == 'menu_mycustomers' or d.startswith('crm_'):
        answer_cb(cbid)
        sid = str(uid)
        my_customers = customers.get(sid, {})

        # Filter va sahifa
        per_page    = 7
        page        = 1
        cur_filter  = 'all'
        if d.startswith('crm_page_'):
            try: page = int(d.split('_')[2])
            except: page = 1
        elif d.startswith('crm_filter_'):
            cur_filter = d[11:]

        # ─── MIJOZ KARTASINI KO'RISH ───
        if d.startswith('crm_view_'):
            cuid = d[9:]
            cust = my_customers.get(cuid, {})
            if not cust:
                send_seller(uid, "❌ Mijoz topilmadi."); return
            avg  = cust['total_spent'] // cust['total_orders'] if cust['total_orders'] > 0 else 0
            tags = ', '.join(cust.get('tags', [])) or '\u2014'
            note = cust.get('note', '')
            phone    = cust.get('phone', '') or '\u2014'
            username = cust.get('username', '')
            username_line = f"@{username}" if username else '\u2014'
            # Faollik holati
            from datetime import datetime as _dt, timedelta
            try:
                last_dt = _dt.strptime(cust.get('last_order','01.01.2020'), '%d.%m.%Y')
                days_ago = (_dt.now() - last_dt).days
                if days_ago < 7:    activity = "🟢 Faol"
                elif days_ago < 30: activity = "🟡 O'rtacha"
                else:               activity = "🔴 Yo'qotilgan"
            except:
                activity = "—"

            send_seller(uid,
                f"👤 <b>{cust['name']}</b>\n"
                "━━━━━━━━━━━━━━━\n"
                "📊 <b>Statistika:</b>\n"
                f"🛒 Jami xaridlar: {cust['total_orders']} ta\n"
                f"💰 Jami sarflagan: {fmt(cust['total_spent'])} so'm\n"
                f"📈 O'rtacha check: {fmt(avg)} so'm\n"
                f"📅 Birinchi xarid: {cust.get('first_order','—')}\n"
                f"📅 Oxirgi xarid: {cust.get('last_order','—')}\n"
                f"⚡ Holati: {activity}\n"
                f"🏷 Teglar: {tags}\n"
                f"📞 Telefon: {phone}\n"
                f"👤 Username: {username_line}"
                + (f"\n\n📝 <b>Izoh:</b> {note}" if note else ""),
                {'inline_keyboard': [
                    [{'text': "⭐ VIP",       'callback_data': 'crm_tag_'+cuid+'_vip'},
                     {'text': "🔴 Muammoli", 'callback_data': 'crm_tag_'+cuid+'_problem'},
                     {'text': "💎 Doimiy",   'callback_data': 'crm_tag_'+cuid+'_loyal'}],
                    [{'text': "💬 Xabar yuborish",   'callback_data': 'crm_msg_'+cuid}],
                    [{'text': "📝 Izoh qo'shish",     'callback_data': 'crm_note_'+cuid}],
                    [{'text': "📊 Xaridlar tarixi",   'callback_data': 'crm_history_'+cuid}],
                    [{'text': "⬅️ Orqaga",           'callback_data': 'menu_mycustomers'}],
                ]}
            )
            return

        # ─── XARIDLAR TARIXI ───
        if d.startswith('crm_history_'):
            cuid = d[len('crm_history_'):]
            cust = my_customers.get(cuid, {})
            if not cust:
                send_seller(uid, "❌ Mijoz topilmadi."); return
            orders_list = list(reversed(cust.get('orders', [])))  # eng yangi tepada
            if not orders_list:
                txt = (f"📊 <b>Xaridlar tarixi — {cust['name']}</b>\n"
                       f"━━━━━━━━━━━━━━━\n\nHali xarid yo'q.")
            else:
                lines = [f"📊 <b>Xaridlar tarixi — {cust['name']}</b>",
                         "━━━━━━━━━━━━━━━"]
                for i, o in enumerate(orders_list, 1):
                    lines.append(f"{i}. {o.get('product','—')}")
                    lines.append(f"   💰 {fmt(o.get('amount',0))} so'm · {o.get('date','—')}")
                lines.append("━━━━━━━━━━━━━━━")
                lines.append(f"💰 Jami: {fmt(cust['total_spent'])} so'm")
                lines.append("<i>(oxirgi 20 ta xarid ko'rsatilgan)</i>")
                txt = "\n".join(lines)
            send_seller(uid, txt, {'inline_keyboard': [
                [{'text': "⬅️ Mijozga qaytish", 'callback_data': 'crm_view_'+cuid}],
            ]})
            return

        # ─── TEG QO'YISH ───
        if d.startswith('crm_tag_'):
            parts = d.split('_')
            cuid, tag = parts[2], parts[3]
            if cuid in my_customers:
                tags = my_customers[cuid].get('tags', [])
                if tag in tags:
                    tags.remove(tag); msg = f"🏷 Teg olib tashlandi: {tag}"
                else:
                    tags.append(tag); msg = f"✅ Teg qo'shildi: {tag}"
                my_customers[cuid]['tags'] = tags
                customers[sid] = my_customers
                save_data()
                send_seller(uid, msg)
            return

        # ─── XABAR YUBORISH ───
        if d.startswith('crm_msg_'):
            cuid = d[8:]
            cust = my_customers.get(cuid, {})
            if not cust:
                send_seller(uid, "❌ Mijoz topilmadi."); return
            seller_state[uid] = {'step': 'crm_send_msg', 'target_uid': cust.get('user_id'), 'target_name': cust['name']}
            send_seller(uid,
                f"💬 <b>{cust['name']}</b> ga xabar yuborish\n\n"
                f"Xabar matnini yozing (yoki /cancel):"
            )
            return

        # ─── IZOH QO'SHISH ───
        if d.startswith('crm_note_'):
            cuid = d[9:]
            cust = my_customers.get(cuid, {})
            if not cust:
                send_seller(uid, "❌ Mijoz topilmadi."); return
            seller_state[uid] = {'step': 'crm_add_note', 'target_cuid': cuid, 'target_name': cust['name']}
            current_note = cust.get('note', '')
            send_seller(uid,
                f"📝 <b>{cust['name']}</b> uchun izoh\n\n"
                + (f"Joriy izoh: <i>{current_note}</i>\n\n" if current_note else "") +
                f"Yangi izoh yozing (yoki /cancel):"
            )
            return

        # ─── QIDIRUV ───
        if d == 'crm_search':
            seller_state[uid] = {'step': 'crm_search_query'}
            send_seller(uid,
                "🔍 <b>Mijoz qidirish</b>\n\n"
                "Ism yoki telefon raqamini yozing:"
            )
            return

        # ─── RO'YXAT (filter bilan) ───
        if not my_customers:
            send_seller(uid,
                "👥 <b>Mijozlar bazasi</b>\n\nHali mijoz yo'q.\n"
                "Buyurtmalar tasdiqlanganidan keyin mijozlar bu yerda ko'rinadi.",
                {'inline_keyboard': [[{'text': "⬅️ Menyu", 'callback_data': 'back_menu'}]]}
            )
            return

        # Filter qo'llash
        from datetime import datetime as _dt
        def days_since_last(c):
            try:
                last_dt = _dt.strptime(c.get('last_order','01.01.2020'), '%d.%m.%Y')
                return (_dt.now() - last_dt).days
            except: return 999

        all_items = list(my_customers.items())
        if cur_filter == 'vip':
            filtered = [(k,v) for k,v in all_items if 'vip' in v.get('tags', [])]
        elif cur_filter == 'active':
            filtered = [(k,v) for k,v in all_items if days_since_last(v) < 7]
        elif cur_filter == 'lost':
            filtered = [(k,v) for k,v in all_items if days_since_last(v) >= 30]
        elif cur_filter == 'new':
            filtered = [(k,v) for k,v in all_items if v.get('total_orders', 0) == 1]
        elif cur_filter == 'repeat':
            filtered = [(k,v) for k,v in all_items if v.get('total_orders', 0) > 1]
        else:
            filtered = all_items

        # Saralash
        filtered.sort(key=lambda x: x[1]['total_spent'], reverse=True)
        total = len(filtered)
        start = (page - 1) * per_page
        page_custs = filtered[start:start + per_page]

        # Umumiy stat
        total_revenue = sum(v['total_spent'] for v in my_customers.values())
        repeat = sum(1 for v in my_customers.values() if v['total_orders'] > 1)
        vip    = sum(1 for v in my_customers.values() if 'vip' in v.get('tags', []))

        filter_labels = {
            'all':    'Hammasi', 'vip': 'VIP', 'active': 'Faol',
            'lost':   "Yo'qotilgan", 'new': 'Yangi', 'repeat': 'Qaytib kelgan'
        }
        text = (
            "👥 <b>Mijozlar bazasi</b>\n"
            "━━━━━━━━━━━━━━━\n"
            f"📊 Jami: {len(my_customers)} ta • VIP: {vip} • Qaytib kelgan: {repeat}\n"
            f"💰 Jami daromad: {fmt(total_revenue)} so'm\n"
            f"🔍 Filter: <b>{filter_labels.get(cur_filter, 'Hammasi')}</b> ({total} ta)\n\n"
        )

        if not page_custs:
            text += "Bu filterda mijoz topilmadi.\n"
        else:
            for i, (cuid, cust) in enumerate(page_custs, start=start+1):
                medal = ['🥇','🥈','🥉'][i-1] if i <= 3 else str(i) + "."
                badges = ""
                if 'vip' in cust.get('tags', []):     badges += " ⭐"
                if 'problem' in cust.get('tags', []): badges += " 🔴"
                if 'loyal' in cust.get('tags', []):   badges += " 💎"
                d_ago = days_since_last(cust)
                act_icon = "🟢" if d_ago < 7 else "🟡" if d_ago < 30 else "🔴"
                text += (
                    f"{medal} <b>{cust['name']}{badges}</b>\n"
                    f"   🛒 {cust['total_orders']} ta • 💰 {fmt(cust['total_spent'])} so'm\n"
                    f"   {act_icon} Oxirgi: {cust.get('last_order','—')}\n\n"
                )

        # Klaviatura
        kb_rows = []
        # Filter qatori
        kb_rows.append([
            {'text': ('✅ ' if cur_filter=='all' else '') + 'Hammasi',  'callback_data': 'crm_filter_all'},
            {'text': ('✅ ' if cur_filter=='vip' else '⭐ ') + 'VIP',    'callback_data': 'crm_filter_vip'},
            {'text': ('✅ ' if cur_filter=='active' else '🟢 ') + 'Faol','callback_data': 'crm_filter_active'},
        ])
        kb_rows.append([
            {'text': ('✅ ' if cur_filter=='lost' else '🔴 ') + "Yo'qotilgan", 'callback_data': 'crm_filter_lost'},
            {'text': ('✅ ' if cur_filter=='new' else '🆕 ') + 'Yangi',         'callback_data': 'crm_filter_new'},
            {'text': ('✅ ' if cur_filter=='repeat' else '🔄 ') + 'Qaytma',     'callback_data': 'crm_filter_repeat'},
        ])
        # Qidiruv
        kb_rows.append([{'text': "🔍 Qidirish", 'callback_data': 'crm_search'}])
        # Mijozlar — yangi format: "👤 {ism} · {N} ta · {summa_qisqa}"
        for cuid, cust in page_custs:
            cust_name = cust.get('name', '—')
            if len(cust_name) > 22:
                cust_name = cust_name[:21].rstrip() + '…'
            label = (f"👤 {cust_name} · {cust['total_orders']} ta · "
                     f"{format_price_short(cust['total_spent'])}")
            kb_rows.append([{'text': label, 'callback_data': 'crm_view_' + cuid}])
        # Pagination
        nav = []
        if page > 1:
            nav.append({'text': "◀️", 'callback_data': f'crm_page_{page-1}'})
        if start + per_page < total:
            nav.append({'text': "▶️", 'callback_data': f'crm_page_{page+1}'})
        if nav: kb_rows.append(nav)
        kb_rows.append([{'text': "⬅️ Menyu", 'callback_data': 'back_menu'}])

        send_seller(uid, text, {'inline_keyboard': kb_rows})
        return

    # ─── LIVE COMMERCE ───────────────────────────────────────────
    if d == 'live_cancel':
        answer_cb(cbid, "Bekor qilindi")
        if uid in seller_state and seller_state[uid].get('step','').startswith('live_'):
            del seller_state[uid]
        return

    if d.startswith('live_pick_'):
        answer_cb(cbid)
        pid = d[10:]
        if pid not in products:
            send_seller(uid, "❌ Mahsulot topilmadi.")
            return
        seller_state[uid] = {
            'step': 'live_video',
            'product_id': pid,
        }
        p = products[pid]
        send_seller(uid,
            f"📦 <b>{p.get('name','')}</b>\n\n"
            f"🎥 <b>Live videoni yuboring</b>\n\n"
            f"⏱ Davomiyligi: 15 soniyadan 5 daqiqagacha\n"
            f"📐 Vertikal video tavsiya etiladi (9:16)\n\n"
            f"<i>Videoni shu chatga yuboring...</i>",
            {'inline_keyboard': [[{'text': "❌ Bekor", 'callback_data': 'live_cancel'}]]}
        )
        return

    if d.startswith('live_dur_'):
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s or s.get('step') != 'live_duration':
            return
        try:
            hours = int(d.split('_')[2])
        except:
            return
        s['duration_hours'] = hours
        s['step'] = 'live_discount'
        labels = {1: '1 soat', 3: '3 soat', 24: '24 soat'}
        send_seller(uid,
            f"⏰ Davomiyligi: {labels.get(hours, str(hours)+' soat')}\n\n"
            f"💸 <b>Qo'shimcha chegirma %?</b>\n\n"
            f"<i>Oddiy guruh narxi ustiga qo'shimcha chegirma.\n"
            f"Misol: oddiy 30%, live qo'shimcha 10% = jami 40% tejash</i>",
            {'inline_keyboard': [
                [{'text': "5%",  'callback_data': 'live_disc_5'},
                 {'text': "10%", 'callback_data': 'live_disc_10'},
                 {'text': "15%", 'callback_data': 'live_disc_15'},
                 {'text': "20%", 'callback_data': 'live_disc_20'}],
                [{'text': "❌ Bekor", 'callback_data': 'live_cancel'}],
            ]}
        )
        return

    if d.startswith('live_disc_'):
        answer_cb(cbid)
        s = seller_state.get(uid)
        if not s or s.get('step') != 'live_discount':
            return
        try:
            pct = int(d.split('_')[2])
        except:
            return
        s['discount_pct'] = pct
        s['step'] = 'live_confirm'
        # Tasdiq xabari
        p = products.get(s['product_id'], {})
        group_price = p.get('group_price', 0)
        live_price  = int(group_price * (100 - pct) / 100)
        s['live_price'] = live_price
        send_seller(uid,
            f"📋 <b>LIVE tasdiqlash</b>\n\n"
            f"📦 {p.get('name','')}\n"
            f"💰 Oddiy guruh: {fmt(group_price)} so'm\n"
            f"🔴 Live narx: <b>{fmt(live_price)} so'm</b> (-{pct}% qo'shimcha)\n"
            f"⏰ Davomiyligi: {s['duration_hours']} soat\n\n"
            f"Live boshlangach kanaliga avtomatik post yuboramiz.",
            {'inline_keyboard': [
                [{'text': "🚀 LIVE BOSHLASH", 'callback_data': 'live_start'}],
                [{'text': "❌ Bekor",         'callback_data': 'live_cancel'}],
            ]}
        )
        return

    if d == 'live_start':
        answer_cb(cbid, "🔴 Live boshlanmoqda...", token=SELLER_TOKEN)
        s = seller_state.get(uid)
        if not s or s.get('step') != 'live_confirm':
            return
        # Live yaratish
        from datetime import datetime as _dt, timedelta
        live_id = 'live_' + ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
        now = _dt.now()
        ends = now + timedelta(hours=s['duration_hours'])
        live_data = {
            'id':             live_id,
            'product_id':     s['product_id'],
            'seller_id':      uid,
            'video_file_id':  s.get('video_file_id', ''),
            'video_duration': s.get('video_duration', 0),
            'duration_hours': s['duration_hours'],
            'discount_pct':   s['discount_pct'],
            'live_price':     s['live_price'],
            'status':         'live',
            'started_at':     now.strftime('%Y-%m-%d %H:%M'),
            'ends_at':        ends.strftime('%Y-%m-%d %H:%M'),
            'viewers':        [],
            'joiners':        [],
            'viewer_count':   0,
            'questions':      [],
            'channel_msg_id': None,
        }
        lives[live_id] = live_data
        del seller_state[uid]
        save_data()

        # Sotuvchi kanaliga post yuborish
        p = products.get(s['product_id'], {})
        channel = p.get('seller_channel', '')
        if channel:
            try:
                video_id = live_data['video_file_id']
                caption = (
                    f"🔴 <b>LIVE — {p.get('name','')}</b>\n\n"
                    f"💰 Oddiy: {fmt(p.get('group_price',0))} so'm\n"
                    f"🔥 LIVE narx: <b>{fmt(s['live_price'])} so'm</b>\n"
                    f"⏰ {s['duration_hours']} soat ichida tugaydi!\n\n"
                    f"🎬 Tomosha qiling va guruhga qo'shiling 👇"
                )
                live_url = f"{(BACKEND_URL or '').rstrip('/')}/live/{live_id}"
                kb_live = {'inline_keyboard': [[
                    {'text': "▶️ LIVE TOMOSHA QILISH", 'url': live_url}
                ]]}
                if video_id:
                    r = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendVideo', json={
                        'chat_id':    channel,
                        'video':      video_id,
                        'caption':    caption,
                        'parse_mode': 'HTML',
                        'reply_markup': kb_live,
                    }).json()
                else:
                    r = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendMessage', json={
                        'chat_id':    channel,
                        'text':       caption,
                        'parse_mode': 'HTML',
                        'reply_markup': kb_live,
                    }).json()
                if r.get('ok'):
                    lives[live_id]['channel_msg_id'] = r['result'].get('message_id')
                    save_data()
            except Exception as e:
                logging.error(f"Live channel post error: {e}")

        send_seller(uid,
            f"🔴 <b>LIVE BOSHLANDI!</b>\n\n"
            f"📦 {p.get('name','')}\n"
            f"⏰ Tugaydi: {ends.strftime('%H:%M')}\n\n"
            f"Real-time dashboard:",
            {'inline_keyboard': [[
                {'text': "📊 Dashboard", 'callback_data': f'live_dash_{live_id}'},
            ]]}
        )
        return

    if d.startswith('live_dash_'):
        answer_cb(cbid)
        live_id = d[10:]
        lv = lives.get(live_id)
        if not lv:
            send_seller(uid, "❌ Live topilmadi.")
            return
        p = products.get(lv.get('product_id',''), {})
        from datetime import datetime as _dt
        try:
            ends = _dt.strptime(lv['ends_at'], '%Y-%m-%d %H:%M')
            remaining = ends - _dt.now()
            if remaining.total_seconds() > 0:
                h = int(remaining.total_seconds() // 3600)
                m = int((remaining.total_seconds() % 3600) // 60)
                time_str = f"{h} soat {m} daqiqa qoldi"
            else:
                time_str = "TUGADI"
                lv['status'] = 'ended'
                save_data()
        except:
            time_str = "—"

        joiners = lv.get('joiners', [])
        joined_amount = sum(orders.get(j.get('order_code',''), {}).get('amount', 0)
                            for j in joiners if isinstance(j, dict))
        send_seller(uid,
            f"📊 <b>LIVE Dashboard</b>\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📦 {p.get('name','')}\n"
            f"⏰ {time_str}\n\n"
            f"👀 Ko'rdi: <b>{lv.get('viewer_count', 0)}</b>\n"
            f"👥 Qo'shildi: <b>{len(joiners)}</b>\n"
            f"💰 Sotuv: {fmt(joined_amount)} so'm\n"
            f"🔴 Status: {lv.get('status','—').upper()}",
            {'inline_keyboard': [
                [{'text': "🔄 Yangilash",  'callback_data': f'live_dash_{live_id}'}],
                [{'text': "🛑 Tugatish",   'callback_data': f'live_end_{live_id}'}],
                [{'text': "⬅️ Menyu",      'callback_data': 'back_menu'}],
            ]}
        )
        return

    if d.startswith('live_end_'):
        answer_cb(cbid, "Live tugatildi")
        live_id = d[9:]
        lv = lives.get(live_id)
        if not lv or lv.get('seller_id') != uid:
            return
        lv['status'] = 'ended'
        save_data()
        p = products.get(lv.get('product_id',''), {})
        send_seller(uid,
            f"🛑 <b>Live tugatildi</b>\n\n"
            f"📦 {p.get('name','')}\n"
            f"👀 Ko'rdi: {lv.get('viewer_count', 0)}\n"
            f"👥 Qo'shildi: {len(lv.get('joiners', []))}"
        )
        return
    # ─── /LIVE COMMERCE ──────────────────────────────────────────

    if d == 'menu_export' or d.startswith('export_'):
        answer_cb(cbid)

        # Tanlash menyusi
        if d == 'menu_export':
            send_seller(uid,
                "📑 <b>Eksport — Excel</b>\n\n"
                "Nimani yuklab olasiz?",
                {'inline_keyboard': [
                    [{'text': "🛒 Buyurtmalar",     'callback_data': 'export_orders'}],
                    [{'text': "👥 Mijozlar",        'callback_data': 'export_customers'}],
                    [{'text': "📦 Mahsulotlar",     'callback_data': 'export_products'}],
                    [{'text': "💰 Moliyaviy",       'callback_data': 'export_finance'}],
                    [{'text': "⬅️ Menyu",           'callback_data': 'back_menu'}],
                ]}
            )
            return

        # Excel yaratish
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import io
            wb = Workbook()
            ws = wb.active
            header_fill = PatternFill(start_color="FF6A1A", end_color="FF6A1A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center      = Alignment(horizontal="center", vertical="center")

            export_type = d[7:]  # orders / customers / products / finance
            my_pids = set(seller_products.get(uid, []))
            sid = str(uid)
            filename = f"joynshop_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            if export_type == 'orders':
                ws.title = "Buyurtmalar"
                headers = ["№", "Kod", "Sana", "Mahsulot", "Mijoz", "Telefon", "Summa", "Tur", "Status", "Manzil"]
                ws.append(headers)
                for i, (code, o) in enumerate([(c,o) for c,o in orders.items() if o.get('product_id') in my_pids], 1):
                    p = products.get(o.get('product_id',''), {})
                    ws.append([
                        i, code, o.get('created',''), p.get('name',''),
                        o.get('user_name',''), o.get('user_phone',''),
                        o.get('amount',0),
                        'Yakka' if o.get('type')=='solo' else 'Guruh',
                        o.get('status',''),
                        o.get('address','')
                    ])

            elif export_type == 'customers':
                ws.title = "Mijozlar"
                headers = ["№", "Ism", "Telegram ID", "Buyurtmalar", "Jami", "O'rtacha", "Birinchi", "Oxirgi", "Teglar", "Izoh"]
                ws.append(headers)
                my_custs = customers.get(sid, {})
                sorted_c = sorted(my_custs.items(), key=lambda x: x[1].get('total_spent',0), reverse=True)
                for i, (cuid, c) in enumerate(sorted_c, 1):
                    avg = c['total_spent']//c['total_orders'] if c.get('total_orders') else 0
                    ws.append([
                        i, c.get('name',''), c.get('user_id',''),
                        c.get('total_orders',0), c.get('total_spent',0), avg,
                        c.get('first_order',''), c.get('last_order',''),
                        ', '.join(c.get('tags',[])), c.get('note','')
                    ])

            elif export_type == 'products':
                ws.title = "Mahsulotlar"
                headers = ["№", "ID", "Nom", "Kategoriya", "Asl narx", "Guruh narx", "Yakka narx", "Min guruh", "Qoldiq", "Sotildi", "Status"]
                ws.append(headers)
                my_prods = [(pid, p) for pid, p in products.items() if pid in my_pids]
                for i, (pid, p) in enumerate(my_prods, 1):
                    sold = p.get('stock_initial', 0) - p.get('stock', 0) if p.get('stock', 9999) < 9999 else len(groups.get(pid, []))
                    ws.append([
                        i, pid, p.get('name',''), p.get('category',''),
                        p.get('original_price',0), p.get('group_price',0), p.get('solo_price',0),
                        p.get('min_group',0),
                        p.get('stock','♾') if p.get('stock', 9999) < 9999 else '♾',
                        sold,
                        p.get('status','active')
                    ])

            elif export_type == 'finance':
                ws.title = "Moliya"
                headers = ["Sana", "Buyurtmalar soni", "Jami summa", "Komissiya (5%)", "Toza daromad"]
                ws.append(headers)
                # Kunlar bo'yicha
                from collections import defaultdict
                by_day = defaultdict(lambda: {'count': 0, 'sum': 0})
                for code, o in orders.items():
                    if o.get('product_id') in my_pids and o.get('status') == 'confirmed':
                        date = o.get('created', '').split(' ')[0]
                        by_day[date]['count'] += 1
                        by_day[date]['sum']   += o.get('amount', 0)
                for date in sorted(by_day.keys(), reverse=True):
                    s = by_day[date]['sum']
                    comm = int(s * COMMISSION_RATE)
                    ws.append([date, by_day[date]['count'], s, comm, s - comm])

            # Header style
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            # Auto-width
            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

            # Faylni saqlash va yuborish
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            files = {'document': (filename, buf.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data  = {'chat_id': uid, 'caption': f"📑 {export_type.capitalize()} eksporti tayyor"}
            requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendDocument', data=data, files=files, timeout=30)
        except Exception as e:
            logging.error(f"Export error: {e}")
            send_seller(uid, f"❌ Eksport xatosi: {e}")
        return

    if d == 'menu_myorders':
        answer_cb(cbid)
        my_pids = seller_products.get(uid, [])
        pending = {k:v for k,v in orders.items() if v.get('product_id') in my_pids and v['status']=='confirming'}
        if not pending:
            send_seller(uid, "📋 Tasdiqlanmagan buyurtma yo'q.",
                {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'back_menu'}]]}
            ); return
        for code, o in list(pending.items())[-10:]:
            p = products.get(o['product_id'], {})
            send_seller(uid,
                f"🔔 <b>YANGI TO'LOV!</b>\n\n"
                f"📦 {p.get('name','')}\n👤 {o['user_name']}\n"
                f"💰 {fmt(o['amount'])} so'm\n"
                f"🛒 {'Yakka' if o.get('type')=='solo' else 'Guruh'}\n🆔 #{code}",
                {'inline_keyboard': [[
                    {'text': '✅ Tasdiqlash', 'callback_data': f'seller_ac_{code}'},
                    {'text': '❌ Rad',        'callback_data': f'seller_ar_{code}'}
                ]]}
            )
        return

    if d == 'menu_myproducts':
        answer_cb(cbid)
        render_myproducts(uid, uid, page=0)
        return

    if d.startswith('mp_page_'):
        try:
            page = int(d[8:])
        except ValueError:
            answer_cb(cbid); return
        answer_cb(cbid)
        render_myproducts(uid, uid, page=page)
        return

    if (d.startswith('mp_edit_') and not d.startswith('mp_edit_field_')) or d.startswith('edit_prod_'):
        pid = d[8:] if d.startswith('mp_edit_') else d[10:]
        p = products.get(pid)
        if not p or p.get('seller_id') != uid:
            answer_cb(cbid, "❌ Topilmadi"); return
        answer_cb(cbid)
        kb = {'inline_keyboard': [
            [{'text': "📝 Nom",            'callback_data': f'mp_edit_field_name_{pid}'}],
            [{'text': "💰 Asl narx",       'callback_data': f'mp_edit_field_orig_{pid}'},
             {'text': "👥 Guruh narxi",    'callback_data': f'mp_edit_field_grp_{pid}'}],
            [{'text': "👤 Yakka narx",     'callback_data': f'mp_edit_field_solo_{pid}'},
             {'text': "👥 Min guruh",      'callback_data': f'mp_edit_field_min_{pid}'}],
            [{'text': "⏰ Deadline",       'callback_data': f'mp_edit_field_deadline_{pid}'},
             {'text': "📝 Tavsif",         'callback_data': f'mp_edit_field_desc_{pid}'}],
            [{'text': "🎨 Variantlar",     'callback_data': f'mp_edit_field_variants_{pid}'},
             {'text': "📸 Rasm",           'callback_data': f'mp_edit_field_photo_{pid}'}],
            [{'text': "🏷 MXIK",           'callback_data': f'mp_edit_field_mxik_{pid}'}],
            [{'text': "⬅️ Orqaga",         'callback_data': f'mp_view_{pid}'}],
        ]}
        send_seller(uid,
            f"✏️ <b>{p.get('name','')}</b> ni tahrirlash\n\n"
            f"Qaysi maydonni o'zgartirasiz?", kb)
        return

    if d.startswith('mp_edit_field_'):
        rest = d[len('mp_edit_field_'):]
        # field_<pid> formati
        try:
            field, pid = rest.split('_', 1)
        except ValueError:
            answer_cb(cbid); return
        p = products.get(pid)
        if not p or p.get('seller_id') != uid:
            answer_cb(cbid, "❌ Topilmadi"); return
        answer_cb(cbid)
        # MXIK — alohida search flow (prod_mxik_* step'larini qayta ishlatadi, lekin edit rejimida)
        if field == 'mxik':
            seller_state[uid] = {
                'step': 'prod_mxik_search',
                'pp_pid': pid,
                'mxik_after': 'edit_pp',
            }
            send_seller(uid,
                f"🏷 <b>MXIK qayta tanlash (ixtiyoriy)</b>\n\n"
                f"📦 {p.get('name','')[:50]}\n\n"
                f"Mahsulot nomini yoki kalit so'z kiriting yoki o'tkazib yuboring:",
                {'inline_keyboard': [
                    [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                    [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
                ]})
            return
        prompts = {
            'name':      ("📝 Yangi nom yozing:", 'pp_edit_name'),
            'orig':      ("💰 Yangi asl narxni yozing (so'm):", 'pp_edit_orig'),
            'grp':       ("👥 Yangi guruh narxini yozing (so'm):", 'pp_edit_grp'),
            'solo':      ("👤 Yangi yakka narxni yozing (so'm):", 'pp_edit_solo'),
            'min':       ("👥 Yangi minimal guruh sonini yozing (2-100):", 'pp_edit_min'),
            'deadline':  ("⏰ Yangi muddat (soat, masalan: 24, 48, 72, 168):", 'pp_edit_deadline'),
            'desc':      ("📝 Yangi tavsifni yozing:", 'pp_edit_desc'),
            'variants':  ("🎨 Yangi variantlar (vergul bilan ajrating):", 'pp_edit_variants'),
            'photo':     ("📸 Yangi rasm(lar)ni yuboring (bitta yoki bir nechta):", 'pp_edit_photo'),
        }
        if field not in prompts:
            send_seller(uid, "❌ Noma'lum maydon"); return
        prompt, step = prompts[field]
        seller_state[uid] = {'step': step, 'pp_pid': pid, 'pp_photo_ids': []}
        send_seller(uid, prompt + "\n\n<i>Bekor qilish: /cancel</i>")
        return

    if d.startswith('mp_view_'):
        pid = d[8:]
        p = products.get(pid)
        if not p or p.get('seller_id') != uid:
            answer_cb(cbid, "❌ Topilmadi"); return
        answer_cb(cbid)
        cls   = _classify_product_status(p)
        count = len(groups.get(pid, []))
        is_billz_draft = (p.get('source') == 'billz' and not p.get('is_active', True))
        # Narx liniyasi — draft uchun asl narx, aks holda guruh narxi
        if is_billz_draft:
            price_line = f"💰 {fmt(p.get('original_price',0))} so'm  <i>(asl narx)</i>"
        else:
            grp = p.get('group_price', 0) or p.get('original_price', 0)
            price_line = f"💰 {fmt(grp)} so'm"
        source_label = "Billz" if p.get('source') == 'billz' else "Manual"
        mxik_code = p.get('mxik_code')
        mxik_line = f"\n🏷 MXIK: <code>{mxik_code}</code>" if mxik_code else "\n🏷 MXIK: ⚠️ Yo'q (kiritish kerak)"
        txt = (
            f"{cls['emoji']} <b>{p.get('name','')}</b>\n\n"
            f"{price_line}\n"
            f"👥 Guruh: {count}/{p.get('min_group',0)}\n"
            f"🛍 {source_label}\n"
            f"📅 Tugash: {p.get('deadline','—') or '—'}\n"
            f"📊 Status: {cls['label']}"
            f"{mxik_line}"
        )
        # Tugma layouti — holatga qarab
        kb_rows = []
        is_closed  = (p.get('status') == 'closed')
        is_expired = (cls['label'] == 'Muddati tugagan')
        if is_billz_draft:
            kb_rows.append([
                {'text': "▶️ Yoqish",   'callback_data': f'bz_activate_{pid}'},
                {'text': "🗑 O'chirish", 'callback_data': f'mp_del_{pid}'},
            ])
        elif is_expired:
            kb_rows.append([
                {'text': "♻️ Qayta yoqish", 'callback_data': f'mp_renew_{pid}'},
                {'text': "🗑 O'chirish",    'callback_data': f'mp_del_{pid}'},
            ])
        elif not is_closed:
            kb_rows.append([
                {'text': "✏️ Tahrirlash", 'callback_data': f'mp_edit_{pid}'},
                {'text': "🗑 O'chirish",  'callback_data': f'mp_del_{pid}'},
            ])
        # Closed (faqat o'qish) — yuqori qator umuman yo'q
        kb_rows.append([
            {'text': "📊 Statistika", 'callback_data': f'mp_stats_{pid}'},
            {'text': "🔙 Orqaga",     'callback_data': 'menu_myproducts'},
        ])
        send_seller(uid, txt, {'inline_keyboard': kb_rows})
        return

    if d.startswith('mp_stats_'):
        pid = d[9:]
        if pid not in products:
            answer_cb(cbid, "❌ Topilmadi"); return
        answer_cb(cbid, "📊 Tez orada qo'shiladi", alert=True)
        return

    if d.startswith('mp_renew_') and not d.startswith('mp_renewh_'):
        pid = d[9:]
        p = products.get(pid)
        if not p or p.get('seller_id') != uid:
            answer_cb(cbid, "❌ Topilmadi"); return
        answer_cb(cbid)
        send_seller(uid,
            f"♻️ <b>Qayta yoqish — {p.get('name','')}</b>\n\n"
            f"Yangi muddat tanlang. Boshqa maydonlar saqlanadi (narx, min guruh va h.k.).",
            {'inline_keyboard': [
                [{'text': "24 soat",  'callback_data': f'mp_renewh_{pid}_24'},
                 {'text': "2 kun",    'callback_data': f'mp_renewh_{pid}_48'}],
                [{'text': "3 kun",    'callback_data': f'mp_renewh_{pid}_72'},
                 {'text': "1 hafta",  'callback_data': f'mp_renewh_{pid}_168'}],
                [{'text': "🔙 Bekor", 'callback_data': f'mp_view_{pid}'}],
            ]})
        return

    if d.startswith('mp_renewh_'):
        # mp_renewh_<pid>_<hours>
        rest = d[len('mp_renewh_'):]
        try:
            pid, hours_str = rest.rsplit('_', 1)
            hours = int(hours_str)
        except (ValueError, IndexError):
            answer_cb(cbid); return
        p = products.get(pid)
        if not p or p.get('seller_id') != uid:
            answer_cb(cbid, "❌ Topilmadi"); return
        if hours < 1 or hours > 720:
            answer_cb(cbid, "❌ Noto'g'ri muddat"); return
        new_dt = datetime.now() + timedelta(hours=hours)
        p['deadline']    = new_dt.strftime('%d.%m.%Y %H:%M')
        p['deadline_dt'] = new_dt.strftime('%Y-%m-%d %H:%M')
        p['status']      = 'active'
        p['is_active']   = True
        save_data()
        answer_cb(cbid, "✅ Qayta yoqildi")
        send_seller(uid,
            f"✅ <b>{p.get('name','')}</b> qayta yoqildi.\n\n"
            f"📅 Yangi muddat: {p['deadline']}\n\n"
            f"<i>Kanal posti caption'i 30 soniya ichida avtomatik yangilanadi.</i>",
            {'inline_keyboard': [
                [{'text': "👁 Mahsulotni ko'rish", 'callback_data': f'mp_view_{pid}'}],
                [{'text': "📦 Mahsulotlarim",     'callback_data': 'menu_myproducts'}],
            ]})
        return

    if d == 'menu_help':
        answer_cb(cbid)
        send_seller(uid,
"📘 <b>Sotuvchi yordam</b>\n\n"
            "/start       — 🏠 Bosh sahifa\n"
            "/addproduct  — ➕ Mahsulot qo'shish\n"
            "/myproducts  — 📦 Mahsulotlarim\n"
            "/myorders    — 📋 Buyurtmalar\n"
            "/mystats     — 📊 Statistika\n"
            "/billz       — 🔌 Billz integratsiyasi\n"
            "/legal       — 📋 Yuridik ma'lumotlar\n"
            "/menu        — 📱 Bosh menyu\n"
            "/cancel      — ❌ Bekor qilish\n"
            "/help        — ℹ️ Yordam\n\n"
            "💬 Yordam: @joynshop_support",
            {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'back_menu'}]]}
        )
        return

    if d == 'back_menu':
        answer_cb(cbid)
        send_seller(uid,
            "🏪 <b>Joynshop Sotuvchi Paneli</b>\n\nGuruh savdosi orqali ko'proq soting!",
            {'inline_keyboard': [
                [
                    {'text': "📦 Mahsulotlarim",      'callback_data': 'menu_myproducts'},
                    {'text': "📋 Buyurtmalar",        'callback_data': 'menu_myorders'},
                ],
                [
                    {'text': "➕ Mahsulot qo'shish",  'callback_data': 'menu_addproduct'},
                    {'text': "👥 Mijozlar",            'callback_data': 'menu_mycustomers'},
                ],
                [
                    {'text': "📊 Statistika",         'callback_data': 'menu_mystats'},
                    {'text': "🔌 Integratsiyalar",    'callback_data': 'menu_integrations'},
                ],
            ]}
        )
        return

    # ─── INTEGRATIONS DISPATCH ───
    if d == 'menu_integrations':
        answer_cb(cbid)
        render_integrations_menu(uid, uid)
        return

    if d.startswith('integ_'):
        integ_id = d[6:]
        entry = next((e for e in INTEGRATIONS if e['id'] == integ_id), None)
        if not entry:
            answer_cb(cbid, "❌ Topilmadi"); return
        if entry['status'] != 'active':
            answer_cb(cbid, f"🔒 {entry['name']} tez orada qo'shiladi", alert=True)
            return
        answer_cb(cbid)
        # Billz uchun maxsus mantiq — ulangan bo'lsa boshqaruv, aks holda onboarding
        if integ_id == 'billz':
            connected = seller_billz_connected_shops(uid)
            if len(connected) == 1:
                # Bitta do'kon ulangan — to'g'ridan boshqaruv menyusiga
                _open_billz_management(uid, uid, connected[0])
            elif len(connected) > 1:
                # Ko'p do'kon ulangan — qaysi birini boshqarishni tanlash
                kb = []
                for idx in connected:
                    sh = seller_shops[uid][idx]
                    kb.append([{
                        'text': f"✅ {sh.get('name','—')[:20]} → {sh.get('billz_shop_name','—')[:15]}",
                        'callback_data': f'billz_view_{idx}',
                    }])
                kb.append([{'text': "🔙 Integratsiyalar", 'callback_data': 'menu_integrations'}])
                send_seller(uid,
                    "✅ <b>Billz POS — boshqaruv</b>\n\n"
                    "Sizda bir nechta do'kon Billz'ga ulangan. Qaysi birini boshqarmoqchisiz?",
                    {'inline_keyboard': kb})
            else:
                # Hech qaysi ulanmagan — onboarding boshlash
                _start_billz_onboarding(uid, uid)
            return
        # Boshqa active integratsiyalar uchun standart handler
        handler_name = entry.get('handler')
        handler = globals().get(handler_name) if handler_name else None
        if handler:
            handler(uid, uid)
        else:
            send_seller(uid, f"⚠️ {entry['name']} handler topilmadi")
        return

    # ─── BILLZ INTEGRATION CALLBACKS ───
    if d.startswith('billz_view_'):
        try:
            idx = int(d[11:])
        except ValueError:
            answer_cb(cbid); return
        shops = seller_shops.get(uid, [])
        if idx >= len(shops):
            answer_cb(cbid, "❌ Do'kon topilmadi"); return
        shop = shops[idx]
        answer_cb(cbid)
        connected = bool(shop.get('billz_secret_token'))
        if connected:
            billz_count = sum(1 for p in products.values()
                              if p.get('seller_id') == uid and p.get('source') == 'billz')
            txt = (
                f"✅ <b>{shop.get('name','')}</b>\n\n"
                f"Billz: ulangan\n"
                f"🏬 Billz do'koni: <b>{shop.get('billz_shop_name','—')}</b>\n"
                f"📅 Ulangan: {shop.get('billz_connected_at','—')}\n"
                f"📦 Import qilingan: {billz_count} ta\n"
            )
            kb = [
                [{'text': "📥 Mahsulotlarni import/yangilash", 'callback_data': f'billz_import_{idx}'}],
                [{'text': "⚙️ Global chegirma sozlamalari",   'callback_data': f'billz_disc_{idx}'}],
                [{'text': "🔌 Uzish (Phase 4)",                'callback_data': f'billz_disconnect_{idx}'}],
                [{'text': "⬅️ Orqaga",                         'callback_data': 'billz_menu'}],
            ]
        else:
            txt = (
                f"⚪️ <b>{shop.get('name','')}</b>\n\n"
                f"Billz hali ulanmagan.\n\n"
                f"<b>Ulash uchun:</b>\n"
                f"1. Billz UI → Sozlamalar → API → <b>Создать ключ</b>\n"
                f"2. Yaratilgan secret token'ni nusxalang\n"
                f"3. Quyidagi tugmani bosing va token'ni shu chatga yuboring"
            )
            kb = [
                [{'text': "🔌 Billz ni ulash", 'callback_data': f'billz_connect_{idx}'}],
                [{'text': "⬅️ Orqaga",        'callback_data': 'billz_menu'}],
            ]
        send_seller(uid, txt, {'inline_keyboard': kb})
        return

    if d == 'billz_menu':
        answer_cb(cbid)
        render_billz_menu(uid, uid)
        return

    if d.startswith('bz_activate_'):
        pid = d[12:]
        p = products.get(pid)
        if not p or p.get('seller_id') != uid:
            answer_cb(cbid, "❌ Topilmadi"); return
        if p.get('source') != 'billz':
            answer_cb(cbid, "❌ Faqat Billz mahsulotlari"); return
        answer_cb(cbid)
        # Sotuvchi shop'idan global discount
        shop = next((sh for sh in seller_shops.get(uid, [])
                     if sh.get('billz_secret_token')), None)
        solo_disc  = (shop or {}).get('billz_global_solo_discount', 10)
        group_disc = (shop or {}).get('billz_global_group_discount', 20)
        orig = int(p.get('original_price', 0) or 0)
        suggested_solo  = max(1, int(orig * (100 - solo_disc) / 100))
        suggested_group = max(1, int(orig * (100 - group_disc) / 100))

        # Agar mahsulotda MXIK yo'q bo'lsa — avval MXIK qadami
        if not p.get('mxik_code'):
            seller_state[uid] = {
                'step': 'prod_mxik_search',
                'bz_pid': pid,
                'mxik_after': 'bz_act',
                'bz_suggested_solo':  suggested_solo,
                'bz_suggested_group': suggested_group,
            }
            # Mahsulot nomi bilan boshlovchi taklif — sotuvchi xohlasa darhol qidiradi
            send_seller(uid,
                f"▶️ <b>Yoqish — {p.get('name','')[:40]}</b>\n\n"
                f"🏷 <b>MXIK kodi (ixtiyoriy)</b>\n\n"
                f"Mahsulot nomi yoki kalit so'z kiriting yoki o'tkazib yuboring:\n"
                f"<i>Taklif: {p.get('name','')[:30]}</i>\n\n"
                f"Bekor qilish: /cancel",
                {'inline_keyboard': [
                    [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                    [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
                ]})
            return

        # MXIK bor — to'g'ridan narx qadamiga
        seller_state[uid] = {
            'step': 'bz_act_solo', 'bz_pid': pid,
            'bz_suggested_solo':  suggested_solo,
            'bz_suggested_group': suggested_group,
        }
        send_seller(uid,
            f"▶️ <b>Yoqish — {p.get('name','')}</b>\n\n"
            f"💰 Asl narx: <b>{fmt(orig)} so'm</b>\n\n"
            f"<b>1/4</b> Yakka narxni yozing (so'm).\n"
            f"💡 Tavsiya: <b>{fmt(suggested_solo)}</b> so'm  ({solo_disc}% chegirma)\n\n"
            f"Bekor qilish: /cancel")
        return

    if d.startswith('bz_deadline_'):
        # bz_deadline_24, _48, _72, _168
        try:
            hours = int(d[12:])
        except ValueError:
            answer_cb(cbid); return
        s = seller_state.get(uid)
        if not s or s.get('step') != 'bz_act_deadline':
            answer_cb(cbid); return
        pid = s.get('bz_pid')
        p = products.get(pid)
        if not p:
            seller_state.pop(uid, None)
            answer_cb(cbid, "❌ Topilmadi"); return
        answer_cb(cbid, f"✅ {hours} soat")
        deadline_dt = datetime.now() + timedelta(hours=hours)
        p['deadline']    = deadline_dt.strftime('%d.%m.%Y %H:%M')
        p['deadline_dt'] = deadline_dt.strftime('%Y-%m-%d %H:%M')
        p['solo_price']  = s.get('bz_solo', 0)
        p['group_price'] = s.get('bz_group', 0)
        p['min_group']   = s.get('bz_min', 3)
        p['solo_available'] = bool(p['solo_price'])
        p['is_active']   = True
        p['status']      = 'active'
        save_data()
        seller_state.pop(uid, None)
        send_seller(uid, "📤 Kanalga e'lon qilinmoqda...")
        ok, err = post_to_channel(uid, pid)
        save_data()
        if ok:
            send_seller(uid,
                f"✅ <b>Yoqildi va e'lon qilindi!</b>\n\n"
                f"📦 {p.get('name','')}\n"
                f"💰 {fmt(p.get('group_price',0))} so'm\n"
                f"⏰ {p.get('deadline','')}",
                {'inline_keyboard': [
                    [{'text': "📦 Mahsulotlarim", 'callback_data': 'menu_myproducts'}],
                ]})
        else:
            # Kanal post fail — is_active'ni qaytaramiz
            p['is_active'] = False
            p['status']    = 'draft'
            save_data()
            send_seller(uid,
                f"⚠️ Yoqildi, lekin kanalga post qo'yib bo'lmadi:\n{err}\n\n"
                f"Bot kanalga admin sifatida qo'shilganmi? "
                f"Tekshiring va /myproducts → Yoqish bosib qaytadan urining.")
        return

    if d.startswith('billz_disc_') and not d.startswith('billz_disc_set_'):
        try:
            idx = int(d[11:])
        except ValueError:
            answer_cb(cbid); return
        shops = seller_shops.get(uid, [])
        if idx >= len(shops):
            answer_cb(cbid, "❌ Topilmadi"); return
        shop = shops[idx]
        answer_cb(cbid)
        send_seller(uid,
            f"⚙️ <b>Global chegirma sozlamalari</b>\n\n"
            f"Bu qiymatlar yangi Billz mahsulotini yoqishda <b>tavsiya narx</b> sifatida ishlatiladi.\n\n"
            f"👤 Solo chegirma:  <b>{shop.get('billz_global_solo_discount', 10)}%</b>\n"
            f"👥 Guruh chegirma: <b>{shop.get('billz_global_group_discount', 20)}%</b>",
            {'inline_keyboard': [
                [{'text': "👤 Solo chegirma o'zgartirish", 'callback_data': f'billz_disc_set_solo_{idx}'}],
                [{'text': "👥 Guruh chegirma o'zgartirish", 'callback_data': f'billz_disc_set_grp_{idx}'}],
                [{'text': "⬅️ Orqaga",                     'callback_data': f'billz_view_{idx}'}],
            ]})
        return

    if d.startswith('billz_disc_set_'):
        # billz_disc_set_solo_<idx> | billz_disc_set_grp_<idx>
        rest = d[len('billz_disc_set_'):]
        try:
            kind, idx_str = rest.rsplit('_', 1)
            idx = int(idx_str)
        except (ValueError, IndexError):
            answer_cb(cbid); return
        if kind not in ('solo', 'grp'):
            answer_cb(cbid); return
        answer_cb(cbid)
        seller_state[uid] = {'step': 'bz_set_disc', 'bz_disc_kind': kind, 'bz_disc_idx': idx}
        label = "Solo" if kind == 'solo' else "Guruh"
        send_seller(uid,
            f"⚙️ Yangi <b>{label}</b> chegirma foizini yozing (0-90):\n\n"
            f"Masalan: <code>15</code> — 15%\n\n"
            f"Bekor qilish: /cancel")
        return

    if d.startswith('billz_import_'):
        try:
            idx = int(d[13:])
        except ValueError:
            answer_cb(cbid); return
        shops = seller_shops.get(uid, [])
        if idx >= len(shops) or not shops[idx].get('billz_secret_token'):
            answer_cb(cbid, "❌ Billz ulanmagan"); return
        answer_cb(cbid, "📥 Boshlanmoqda...")
        import_billz_products(uid, uid, idx)
        return

    if d.startswith('billz_connect_'):
        try:
            idx = int(d[14:])
        except ValueError:
            answer_cb(cbid); return
        shops = seller_shops.get(uid, [])
        if idx >= len(shops):
            answer_cb(cbid, "❌ Do'kon topilmadi"); return
        if not get_fernet():
            answer_cb(cbid, "❌ Encryption sozlanmagan", alert=True); return
        answer_cb(cbid)
        seller_state[uid] = {'step': 'billz_secret_token', 'billz_shop_idx': idx}
        send_seller(uid,
            f"🔌 <b>Billz ulash</b>\n\n"
            f"Billz secret token'ingizni shu chatga yuboring.\n\n"
            f"⚠️ Token shifrlanib saqlanadi. Bekor qilish: /cancel")
        return

    if d.startswith('billz_pickshop_'):
        # billz_pickshop_<seller_shop_idx>_<billz_shop_id>
        rest = d[len('billz_pickshop_'):]
        try:
            seller_idx_str, billz_shop_id = rest.split('_', 1)
            seller_idx = int(seller_idx_str)
        except (ValueError, IndexError):
            answer_cb(cbid); return
        s = seller_state.get(uid)
        if not s or s.get('step') != 'billz_shop_select':
            answer_cb(cbid, "❌ Holat topilmadi"); return
        candidates = s.get('billz_candidates', [])
        chosen = next((c for c in candidates if c['shop_id'] == billz_shop_id), None)
        if not chosen:
            answer_cb(cbid, "❌ Do'kon topilmadi"); return
        plain_token = s.get('billz_pending_token')
        if not plain_token:
            seller_state.pop(uid, None)
            answer_cb(cbid, "❌ Token yo'qoldi"); return
        encrypted = encrypt_token(plain_token)
        if not encrypted:
            answer_cb(cbid, "❌ Shifrlash xatosi", alert=True); return
        shops = seller_shops.get(uid, [])
        if seller_idx >= len(shops):
            answer_cb(cbid, "❌ Do'kon topilmadi"); return
        shops[seller_idx]['billz_secret_token']  = encrypted
        shops[seller_idx]['billz_shop_id']       = chosen['shop_id']
        shops[seller_idx]['billz_shop_name']     = chosen['shop_name']
        shops[seller_idx]['billz_connected_at']  = datetime.now().strftime('%Y-%m-%d %H:%M')
        save_data()
        seller_state.pop(uid, None)
        answer_cb(cbid, "✅ Ulandi!")
        send_seller(uid,
            f"✅ <b>Billz ulandi!</b>\n\n"
            f"🏪 {shops[seller_idx].get('name','')}\n"
            f"🏬 Billz do'koni: <b>{chosen['shop_name']}</b>\n\n"
            f"Keyingi qadam: mahsulotlarni import qilish (Faza 2 — keyingi deploy).",
            {'inline_keyboard': [[{'text': "🔌 Billz menyu", 'callback_data': 'billz_menu'}]]})
        return

    # ─── LEGAL INFO CALLBACKS ───
    if d == 'leg_start':
        answer_cb(cbid)
        start_legal_flow(uid, uid, after='menu')
        return

    if d in ('leg_pick_yatt', 'leg_pick_mchj'):
        s = seller_state.get(uid)
        if not s or s.get('step') != 'leg_status':
            answer_cb(cbid); return
        answer_cb(cbid, "✅")
        status = 'yatt' if d == 'leg_pick_yatt' else 'mchj'
        prof = seller_profiles.setdefault(uid, {})
        prof['legal_status'] = status
        # Edit rejimi bo'lsa — leg_confirm'ga qaytamiz
        if s.get('leg_editing'):
            s.pop('leg_editing', None)
            s['step'] = 'leg_confirm'
            render_legal_confirm(uid, uid)
            return
        s['step'] = 'leg_stir'
        send_seller(uid,
            "📋 <b>Qadam 2/6 — STIR</b>\n\n"
            "Soliq raqami (STIR / INN) ni kiriting (9 raqam):\n\n"
            "<i>Bekor qilish: /cancel</i>")
        return

    if d == 'leg_confirm':
        s = seller_state.get(uid)
        if not s or s.get('step') != 'leg_confirm':
            answer_cb(cbid); return
        prof = seller_profiles.get(uid, {})
        # Yakuniy tekshiruv — barcha kerakli field'lar borligini
        required = ['legal_status', 'stir', 'bank_account', 'bank_name', 'bank_mfo']
        if prof.get('legal_status') == 'mchj':
            required.append('director_name')
        missing = [f for f in required if not prof.get(f)]
        if missing:
            answer_cb(cbid, f"❌ Yetishmayapti: {', '.join(missing)}", alert=True)
            return
        answer_cb(cbid, "✅ Saqlandi")
        prof['legal_completed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        save_data()
        after = s.get('leg_after', 'menu')
        seller_state.pop(uid, None)
        if after == 'channel':
            # Onboarding davom etadi — kanal kiritishga o'tamiz
            seller_state[uid] = {**s, 'step': 'ob_channel'}
            seller_state[uid].pop('leg_after', None)
            seller_state[uid].pop('leg_editing', None)
            send_seller(uid,
                "✅ Yuridik ma'lumotlar saqlandi.\n\n"
                "📢 Endi Telegram kanal username:\n<i>@mening_kanalim</i>\n\n"
                "⚠️ Seller bot kanalga <b>admin</b> sifatida qo'shilgan bo'lishi kerak!")
        else:
            send_seller(uid,
                "✅ <b>Yuridik ma'lumotlar saqlandi!</b>\n\n"
                "Payme split to'lov va fiskal chek ulansa, avtomatik faollashtiriladi.",
                {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'back_menu'}]]})
        return

    if d == 'leg_edit_menu':
        s = seller_state.get(uid)
        if not s:
            # Summary ekranidan kelgan — yangi state ochamiz, leg_confirm'ga qaytaramiz
            seller_state[uid] = {'step': 'leg_confirm', 'leg_after': 'menu'}
        else:
            s['step'] = 'leg_confirm'
        answer_cb(cbid)
        is_mchj = seller_profiles.get(uid, {}).get('legal_status') == 'mchj'
        kb = [
            [{'text': "📌 Status", 'callback_data': 'leg_edit_field_status'},
             {'text': "🏛 STIR",   'callback_data': 'leg_edit_field_stir'}],
            [{'text': "🏦 Bank",   'callback_data': 'leg_edit_field_bank_name'},
             {'text': "💳 Hisob",  'callback_data': 'leg_edit_field_bank_account'}],
            [{'text': "🔢 MFO",    'callback_data': 'leg_edit_field_bank_mfo'}],
        ]
        if is_mchj:
            kb.append([{'text': "👔 Direktor", 'callback_data': 'leg_edit_field_director_name'}])
        kb.append([{'text': "⬅️ Bekor", 'callback_data': 'leg_edit_cancel'}])
        send_seller(uid, "✏️ Qaysi ma'lumotni o'zgartirasiz?", {'inline_keyboard': kb})
        return

    if d == 'leg_edit_cancel':
        answer_cb(cbid)
        # Edit menyu'dan bekor — confirm ekraniga qaytamiz (agar kontekst bo'lsa)
        if seller_has_legal(uid):
            render_legal_confirm(uid, uid)
        else:
            send_seller(uid, "❌ Bekor qilindi.",
                {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'back_menu'}]]})
        return

    if d.startswith('leg_edit_field_'):
        field = d[len('leg_edit_field_'):]
        s = seller_state.setdefault(uid, {'step': 'leg_confirm', 'leg_after': 'menu'})
        s['leg_editing'] = True
        answer_cb(cbid)
        prompts = {
            'status': ("📌 Status — qaytadan tanlang:", 'leg_status'),
            'stir':   ("🏛 STIR — 9 raqam:", 'leg_stir'),
            'bank_name':    ("🏦 Bank nomini kiriting:", 'leg_bank_name'),
            'bank_account': ("💳 Hisob raqami — 20 raqam:", 'leg_account'),
            'bank_mfo':     ("🔢 MFO — 5 raqam:", 'leg_mfo'),
            'director_name':("👔 Direktor F.I.O. (3 ta so'z):", 'leg_director'),
        }
        if field not in prompts:
            send_seller(uid, "❌ Noma'lum maydon"); return
        prompt, step = prompts[field]
        s['step'] = step
        if field == 'status':
            send_seller(uid, prompt, {'inline_keyboard': [
                [{'text': "👤 YaTT", 'callback_data': 'leg_pick_yatt'}],
                [{'text': "🏢 MChJ", 'callback_data': 'leg_pick_mchj'}],
            ]})
        else:
            send_seller(uid, prompt + "\n\n<i>Bekor qilish: /cancel</i>")
        return

    if d.startswith('boost_') and not d.startswith('boost_confirm_'):
        pid = d[6:]
        p = products.get(pid)
        if not p: answer_cb(cbid, '❌ Topilmadi!'); return
        answer_cb(cbid)
        send_seller(uid,
            f"📢 <b>{p['name']}</b> ni qayta e'lon qilmoqchimisiz?\n\n"
            f"🆔 <code>{pid}</code>",
            {'inline_keyboard': [[
                {'text': "✅ Ha, e'lon qil", 'callback_data': f'boost_confirm_{pid}'},
                {'text': '❌ Bekor',         'callback_data': 'noop'},
            ]]}
        )
        return

    if d.startswith('mp_del_') or d.startswith('delete_prod_'):
        pid = d[7:] if d.startswith('mp_del_') else d[12:]
        p = products.get(pid)
        if not p: answer_cb(cbid, '❌ Topilmadi!'); return
        if p.get('seller_id') != uid and uid != ADMIN_ID:
            answer_cb(cbid, "❌ Ruxsat yo'q!"); return
        answer_cb(cbid)
        send_seller(uid,
            f"🗑 <b>{p['name']}</b> ni o'chirasizmi?\n\n"
            f"⚠️ Mahsulot ro'yxatdan olib tashlanadi va kanal post o'chiriladi.\n"
            f"Bu amalni qaytarib bo'lmaydi.",
            {'inline_keyboard': [[
                {'text': "✅ Ha, o'chir",   'callback_data': f'mp_delok_{pid}'},
                {'text': "❌ Yo'q",         'callback_data': f'mp_view_{pid}'},
            ]]}
        )
        return

    if d.startswith('mp_delok_') or d.startswith('delete_confirm_'):
        pid = d[9:] if d.startswith('mp_delok_') else d[15:]
        p = products.get(pid)
        if not p: answer_cb(cbid, '❌ Topilmadi!'); return
        if p.get('seller_id') != uid and uid != ADMIN_ID:
            answer_cb(cbid, "❌ Ruxsat yo'q!"); return
        # Soft delete
        p['status']    = 'closed'
        p['is_active'] = False
        # Kanal post o'chirish
        ch_cid = p.get('channel_chat_id')
        ch_mid = p.get('channel_message_id')
        if ch_cid and ch_mid:
            try:
                requests.post(
                    f'https://api.telegram.org/bot{SELLER_TOKEN}/deleteMessage',
                    json={'chat_id': ch_cid, 'message_id': ch_mid}, timeout=5
                )
            except Exception as e:
                logging.error(f"deleteMessage error: {e}")
        # seller_products dan olib tashlash
        if uid in seller_products and pid in seller_products[uid]:
            seller_products[uid].remove(pid)
        save_data()
        answer_cb(cbid, "✅ O'chirildi!")
        send_seller(uid,
            f"🗑 <b>{p.get('name','')}</b> o'chirildi.",
            {'inline_keyboard': [[{'text': "📦 Mahsulotlarim",
                                    'callback_data': 'menu_myproducts'}]]})
        return

    if d.startswith('boost_confirm_'):
        pid = d[14:]
        if pid not in products:
            answer_cb(cbid, '❌ Topilmadi!'); return
        p = products[pid]
        if p.get('seller_id') != uid and uid != ADMIN_ID:
            answer_cb(cbid, "❌ Ruxsat yo'q!"); return
        count    = len(groups.get(pid, []))
        channel  = p.get('seller_channel')
        caption  = post_caption(p, pid)
        kb       = json.dumps(join_kb(pid, count, p['min_group'], has_solo=bool(p.get('solo_price')), sale_type=p.get('sale_type','both')))
        photo_ids = p.get('photo_ids') or ([p['photo_id']] if p.get('photo_id') else [])

        if len(photo_ids) > 1:
            media = [{'type': 'photo', 'media': fid} for fid in photo_ids]
            media[0]['caption'] = caption
            media[0]['parse_mode'] = 'HTML'
            result = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendMediaGroup', json={
                'chat_id': channel, 'media': media,
            }).json()
            if result.get('ok') and result.get('result'):
                requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendMessage', json={
                    'chat_id': channel, 'text': caption,
                    'parse_mode': 'HTML', 'reply_markup': kb,
                })
                products[pid]['channel_message_id'] = result['result'][0].get('message_id')
                products[pid]['channel_chat_id']    = channel
                answer_cb(cbid, "✅ Qayta e'lon qilindi!")
                send_seller(uid, f"📢 <b>{p['name']}</b> qayta e'lon qilindi!")
            else:
                answer_cb(cbid, '❌ Xato!')
        else:
            result = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendPhoto', json={
                'chat_id': channel, 'photo': photo_ids[0] if photo_ids else p.get('photo_id'),
                'caption': caption, 'parse_mode': 'HTML', 'reply_markup': kb,
            }).json()
            if result.get('ok'):
                products[pid]['channel_message_id'] = result['result']['message_id']
                products[pid]['channel_chat_id']    = channel
                answer_cb(cbid, "✅ Qayta e'lon qilindi!")
                send_seller(uid, f"📢 <b>{p['name']}</b> qayta e'lon qilindi!")
            else:
                answer_cb(cbid, "❌ Xato! Bot kanalga admin sifatida qo'shilganmi?")
        return

    if d.startswith('seller_ac_'):
        code     = d[10:]
        if code not in orders:
            answer_cb(cbid, '❌'); return
        o        = orders[code]
        pid      = o['product_id']
        buyer_id = o['user_id']
        p        = products.get(pid, {})
        orders[code]['status'] = 'confirmed'
        # CRM yangilash
        seller_id = p.get('seller_id')
        if seller_id:
            update_customer(seller_id, buyer_id, o.get('user_name',''), o['amount'], p.get('name',''),
                            phone=o.get('user_phone',''), username=o.get('username',''))

        # ─── INVENTAR: stock kamaytirish ───
        if pid in products and products[pid].get('stock', 9999) < 9999:
            products[pid]['stock'] = max(0, products[pid].get('stock', 0) - 1)
            current_stock = products[pid]['stock']
            stock_init    = products[pid].get('stock_initial', current_stock)
            # Stock 0 — mahsulotni yopish
            if current_stock == 0:
                products[pid]['status'] = 'closed'
                if seller_id:
                    send_seller(seller_id,
                        f"🔴 <b>QOLDIQ TUGADI!</b>\n\n"
                        f"📦 {p.get('name','')}\n"
                        f"Mahsulot avtomatik yopildi.\n\n"
                        f"Qoldiq qo'shish uchun mahsulotni tahrirlang.",
                        {'inline_keyboard': [[{'text': "✏️ Tahrirlash", 'callback_data': f'edit_prod_{pid}'}]]}
                    )
            # Stock 5 yoki kam — ogohlantirish (faqat 5,4,3,2,1 da bir marta)
            elif current_stock <= 5 and stock_init > 5:
                if seller_id:
                    send_seller(seller_id,
                        f"⚠️ <b>QOLDIQ KAM!</b>\n\n"
                        f"📦 {p.get('name','')}\n"
                        f"📊 Qoldiq: {current_stock} ta\n\n"
                        f"Tez orada qo'shimcha qo'shing.",
                        {'inline_keyboard': [[{'text': "✏️ Tahrirlash", 'callback_data': f'edit_prod_{pid}'}]]}
                    )

        save_data()

        if o.get('type') == 'group':
            if pid not in groups: groups[pid] = []
            if buyer_id not in groups[pid]: groups[pid].append(buyer_id)
            count = len(groups[pid])
            min_g = p.get('min_group', 3)
            save_data()
            answer_cb(cbid, f'✅ {count}/{min_g}')
            update_profile(buyer_id, o['amount'], p.get('original_price', o['amount']), True)
            send_buyer(buyer_id, build_check(code, o))
            dtype = p.get('delivery_type', 'pickup')
            if dtype == 'deliver':
                send_buyer(buyer_id,
                    f"🚚 <b>Yetkazib berish uchun manzil yuboring</b>\n\n"
                    f"Shahar, tuman, ko'cha, uy raqami\n"
                    f"<i>Masalan: Toshkent, Yunusobod, Amir Temur 108, 15-xonadon</i>"
                )
                get_profile(buyer_id)['awaiting_address'] = code
            ref_link = f"https://t.me/{BUYER_BOT_USERNAME}?start=ref_{buyer_id}"
            send_buyer(buyer_id,
                f"🎉 <b>Guruhga qo'shildingiz!</b>\n\n"
                f"👥 Guruh: {count}/{min_g}\n\nGuruh to'lganda xabar beramiz! 🔔\n\n"
                f"👫 Do'stingizni taklif qiling — +10,000 so'm cashback!",
                {'inline_keyboard': [
                    [{'text': "🔗 Do'stni taklif qilish", 'url': f"https://t.me/share/url?url={ref_link}&text=🛍%20Do'stlarim%20bilan%20birgalikda%20xarid%20qilib%2040%25%20gacha%20tejayapman!%20Sen%20ham%20ulab%20ko'r%20👇"}],
                    [{'text': "↩️ Qaytarish so'rash", 'callback_data': f'refund_{code}'}]
                ]}
            )
            if count >= min_g:
                notify_group_filled(pid)
        else:
            answer_cb(cbid, '✅ Tasdiqlandi!')
            update_profile(buyer_id, o['amount'], p.get('original_price', o['amount']), False)
            send_buyer(buyer_id, build_check(code, o))
            dtype = p.get('delivery_type', 'pickup')
            if dtype == 'deliver':
                send_buyer(buyer_id,
                    f"🚚 <b>Yetkazib berish uchun manzil yuboring</b>\n\n"
                    f"Shahar, tuman, ko'cha, uy raqami\n"
                    f"<i>Masalan: Toshkent, Yunusobod, Amir Temur 108, 15-xonadon</i>"
                )
                get_profile(buyer_id)['awaiting_address'] = code
            send_buyer(buyer_id,
                f"✅ <b>Buyurtma tasdiqlandi!</b>\n\n📞 Sotuvchi: {p.get('contact','')}\n\nMahsulot yetkazilgandan so'ng:",
                {'inline_keyboard': [[
                    {'text': '⭐ Baho bering', 'callback_data': f'rate_start_{pid}'},
                    {'text': '↩️ Qaytarish',   'callback_data': f'refund_{code}'}
                ]]}
            )
        return

    if d.startswith('seller_ar_'):
        code = d[10:]
        if code in orders:
            orders[code]['status'] = 'rejected'
            save_data()
            send_buyer(orders[code]['user_id'],
                f"❌ <b>To'lov tasdiqlanmadi</b>\n\n#{code}\n\nIzohda kodni tekshiring."
            )
        answer_cb(cbid, '❌ Rad'); return

    if d.startswith('seller_approve_refund_'):
        code = d[21:]
        if code in refund_requests:
            refund_requests[code]['status'] = 'approved'
            save_data()
            o = orders.get(code, {})
            send_buyer(refund_requests[code]['user_id'],
                f"✅ <b>Qaytarish tasdiqlandi!</b>\n\n#{code}\n"
                f"💰 {fmt(o.get('amount',0))} so'm 24 soat ichida qaytariladi.\nPayme: {PAYME_NUMBER}"
            )
        answer_cb(cbid, '✅ Tasdiqlandi'); return

    if d.startswith('seller_deny_refund_'):
        code = d[19:]
        if code in refund_requests:
            refund_requests[code]['status'] = 'denied'
            save_data()
            send_buyer(refund_requests[code]['user_id'],
                f"❌ <b>Qaytarish rad etildi</b>\n\n#{code}\nSotuvchi bilan bog'laning."
            )
        answer_cb(cbid, '❌ Rad'); return

    if d in ('variants_yes', 'variants_no'):
        s = seller_state.get(uid)
        if not s:
            answer_cb(cbid, '❌ Jarayon topilmadi!'); return
        answer_cb(cbid)
        if d == 'variants_yes':
            s['step'] = 'variants_input'
            send_seller(uid,
                "Variantlarni vergul bilan yozing:\n\n"
                "<i>O'lcham uchun: 38, 39, 40, 41, 43</i>\n"
                "<i>Rang uchun: Qizil, Ko'k, Yashil</i>\n"
                "<i>Aralash: S, M, L, XL</i>"
            )
        else:
            s['variants'] = []
            s['step'] = 'min_group'
            send_seller(uid, "6️⃣ Minimal guruh soni (2-100):")
        return

    if d.startswith('delivery_'):
        s = seller_state.get(uid)
        if not s:
            answer_cb(cbid, '❌ Jarayon topilmadi!'); return
        dtype = 'deliver' if d == 'delivery_deliver' else 'pickup'
        s['delivery_type'] = dtype
        s['step'] = 'seller_channel'
        answer_cb(cbid)
        send_seller(uid,
            f"{'🚚 Sotuvchi yetkazadi' if dtype == 'deliver' else '🏪 Xaridor olib ketadi'} ✅\n\n"
            "9️⃣ Kanalingiz username ini yozing:\n"
            "<i>Masalan: @mening_kanalim</i>\n\n"
            "⚠️ Sotuvchi bot kanalga <b>admin</b> sifatida qo'shilgan bo'lishi kerak!"
        )
        return

    if d.startswith('addmod_ch_'):
        channel = d[10:]
        answer_cb(cbid)
        if verified_channels.get(channel, {}).get('owner_id') != uid:
            send_seller(uid, "❌ Bu kanal egasi emassiz!"); return
        seller_state[uid] = {'step': 'add_mod_user', 'mod_channel': channel}
        send_seller(uid,
            f"🛡 <b>{channel}</b> uchun moderator qo'shish\n\n"
            f"Moderatorning Telegram @username ini yozing:\n"
            f"<i>Masalan: @username</i>\n\n"
            f"⚠️ U avval seller botni ishga tushirgan bo'lishi kerak!"
        )
        return

    if d == 'confirm_product':
        s = seller_state.get(uid)
        if not s:
            answer_cb(cbid, '❌ Jarayon topilmadi!'); return
        answer_cb(cbid)
        publish_product(uid, uid, s)
        return

    edit_map = {
        'edit_name':           ('name',           '1️⃣ Yangi mahsulot nomini yozing:'),
        'edit_shop_name':      ('shop_name',       "2️⃣ Yangi do'kon nomini yozing:"),
        'edit_description':    ('description',     '3️⃣ Yangi tavsifni yozing:'),
        'edit_original_price': ('original_price',  '4️⃣ Yangi asl narxni yozing (so\'m):'),
        'edit_group_price':    ('group_price',     '5️⃣ Yangi guruh narxini yozing (so\'m):'),
        'edit_min_group':      ('min_group',       '6️⃣ Yangi minimal guruh sonini yozing (2-100):'),
        'edit_photo':          ('photo',           '7️⃣ Yangi rasmni yuboring 📸'),
        'edit_contact':        ('contact',         "8️⃣ Yangi aloqa ma'lumotini yozing:"),
        'edit_seller_channel': ('seller_channel',  '9️⃣ Yangi kanal username ini yozing:'),
    }
    if d in edit_map:
        s = seller_state.get(uid)
        if not s:
            answer_cb(cbid, '❌ Jarayon topilmadi!'); return
        field, prompt = edit_map[d]
        s['step']       = 'editing'
        s['edit_field'] = field
        seller_state[uid] = s
        answer_cb(cbid)
        send_seller(uid, prompt)
        return

# ─── VALIDATION HELPERS ─────────────────────────────────────────────
def parse_price(text):
    """Bo'sh joy va vergullarni olib tashlab int ga o'giradi. Xato bo'lsa None."""
    try:
        return int(str(text).replace(' ', '').replace(',', '').replace('.', ''))
    except (ValueError, TypeError):
        return None

def validate_prices(orig, group, solo, sale_type='both'):
    """Narx mantiq tekshiruvi.
    Returns (ok: bool, error_msg: str). solo=0 bo'lsa solo tekshirilmaydi.
    """
    if orig is None or orig <= 0:
        return False, "❌ Asl narx 0 dan katta bo'lishi kerak"
    if sale_type != 'solo':
        if group is None or group <= 0:
            return False, "❌ Guruh narxi 0 dan katta bo'lishi kerak"
        if group >= orig:
            return False, "❌ Guruh narxi asl narxdan past bo'lishi kerak"
    if solo and solo > 0:
        if solo >= orig:
            return False, "❌ Yakka narx asl narxdan past bo'lishi kerak"
        if sale_type == 'both' and group and solo < group:
            # Both: yakka narx odatda guruh narxidan yuqori
            return False, "❌ Yakka narx guruh narxidan yuqori bo'lishi kerak"
    return True, ''

def validate_stir(text):
    """STIR / INN: aynan 9 ta raqam, 1-chi raqam 1-6.
    Returns (ok, value, err)."""
    import re
    s = (text or '').strip().replace(' ', '')
    if not re.fullmatch(r'[1-6]\d{8}', s):
        return False, '', "❌ STIR — 9 ta raqam, 1-chi raqam 1-6 oralig'ida"
    return True, s, ''

def validate_bank_account(text):
    """Hisob raqami: aynan 20 ta raqam.
    Returns (ok, value, err)."""
    import re
    s = (text or '').strip().replace(' ', '')
    if not re.fullmatch(r'\d{20}', s):
        return False, '', "❌ Hisob raqami — aynan 20 ta raqam"
    return True, s, ''

def validate_mfo(text):
    """MFO bank kodi: aynan 5 ta raqam.
    Returns (ok, value, err)."""
    import re
    s = (text or '').strip().replace(' ', '')
    if not re.fullmatch(r'\d{5}', s):
        return False, '', "❌ MFO — aynan 5 ta raqam"
    return True, s, ''

def validate_bank_name(text):
    """Bank nomi: kamida 3 belgi.
    Returns (ok, value, err)."""
    s = (text or '').strip()
    if len(s) < 3:
        return False, '', "❌ Bank nomi juda qisqa (kamida 3 belgi)"
    return True, s[:100], ''

def validate_director_name(text):
    """Direktor F.I.O.: kamida 3 ta so'z.
    Returns (ok, value, err)."""
    s = (text or '').strip()
    if len(s.split()) < 3:
        return False, '', "❌ To'liq F.I.O. kiriting (familiya, ism, sharif — 3 ta so'z)"
    return True, s[:150], ''

# ─── MXIK (tasnif.soliq.uz) ─────────────────────────────────────────
# TODO: To'lov tizimi (Paylov) ulanganda MXIK ni majburiy qilish kerak.
# Hozir Render -> tasnif.soliq.uz bloklangani uchun optional —
# har MXIK promptida "⏭ O'tkazib yuborish" tugmasi mavjud.
# Qaytarish: prod_desc/prod_skip_desc/mp_edit_field_mxik/bz_activate
# joylaridagi "⏭ O'tkazib yuborish" tugmalarini olib tashlash + skip handler'ni
# faqat draft saqlash uchun moslashtirish.
MXIK_BASE_URL    = 'https://tasnif.soliq.uz/api/cls-api'
MXIK_CACHE_TTL   = 300        # 5 daqiqa
MXIK_PAGE_SIZE   = 10         # API'dan har bir sahifa uchun
MXIK_TIMEOUT_PRIMARY = 10     # birinchi urinish
MXIK_TIMEOUT_RETRY   = 5      # qayta urinish
_mxik_search_cache = {}       # {keyword.lower(): (fetched_at: datetime, results: list)}

# Session — DNS, TLS connection reuse (Render → Toshkent geographik latency uchun muhim)
_mxik_session = None
def _get_mxik_session():
    global _mxik_session
    if _mxik_session is None:
        _mxik_session = requests.Session()
        _mxik_session.headers.update({
            'User-Agent': 'Mozilla/5.0 (compatible; JoynshopBot/1.0; +https://joynshop.uz)',
            'Accept':     'application/json',
            'Accept-Language': 'ru,uz;q=0.9,en;q=0.8',
        })
    return _mxik_session

def mxik_validate_code(text):
    """17 raqamli MXIK kod validatsiyasi.
    Returns (ok, code, err)."""
    import re
    s = (text or '').strip().replace(' ', '').replace('-', '')
    if not re.fullmatch(r'\d{17}', s):
        return False, '', "❌ MXIK kod aynan 17 raqam bo'lishi kerak"
    return True, s, ''

def mxik_simplify_item(item):
    """API response item'idan kerakli field'larni ajratadi."""
    parts = [item.get('groupName', ''), item.get('className', '')]
    classify = ' → '.join(p for p in parts if p)[:80]
    return {
        'code':     item.get('mxikCode', ''),
        'name':     item.get('subPositionName') or item.get('positionName') or item.get('name', ''),
        'classify': classify,
        'brand':    item.get('brandName') or '',
        'units':    item.get('unitsName', '') or '',
    }

def _mxik_do_request(keyword, timeout):
    """Bitta MXIK so'rov — exception'larni propagate qiladi."""
    return _get_mxik_session().get(
        f'{MXIK_BASE_URL}/elasticsearch/search',
        params={'search': keyword, 'size': MXIK_PAGE_SIZE, 'page': 0, 'lang': 'ru'},
        timeout=timeout,
    )

def mxik_search(keyword):
    """Tasnif.soliq.uz dan kalit so'z bo'yicha qidirish.
    5 daqiqalik cache. 10s primary + 5s retry timeout.
    Session reuse orqali DNS/TLS overhead kamayadi.
    Returns (results: list[dict] | None, error: str | None).
    """
    key = (keyword or '').lower().strip()
    if not key:
        return None, "Bo'sh so'rov"
    # Cache check
    if key in _mxik_search_cache:
        fetched_at, results = _mxik_search_cache[key]
        if (datetime.now() - fetched_at).total_seconds() < MXIK_CACHE_TTL:
            return results, None

    last_err = None
    for attempt, timeout in enumerate([MXIK_TIMEOUT_PRIMARY, MXIK_TIMEOUT_RETRY], 1):
        try:
            r = _mxik_do_request(keyword, timeout=timeout)
            if r.status_code != 200:
                last_err = f"HTTP {r.status_code}"
                logging.warning(f"MXIK HTTP non-200 (attempt {attempt}, query={key!r}): {r.status_code}")
                continue
            data = r.json() or {}
            if not data.get('success'):
                last_err = "javob success=false"
                logging.warning(f"MXIK success=false (attempt {attempt}, query={key!r}): {data}")
                continue
            items = data.get('data') or []
            results = [mxik_simplify_item(it) for it in items]
            _mxik_search_cache[key] = (datetime.now(), results)
            return results, None
        except requests.Timeout:
            last_err = f"timeout {timeout}s"
            logging.warning(f"MXIK timeout (attempt {attempt}, query={key!r}, t={timeout}s)")
        except requests.ConnectionError as e:
            last_err = "connection error"
            logging.error(f"MXIK ConnectionError (attempt {attempt}, query={key!r}): {e}")
        except Exception as e:
            last_err = "unknown error"
            logging.error(f"MXIK unknown exception (attempt {attempt}, query={key!r}): {type(e).__name__}: {e}")

    # Ikkala urinish ham fail
    return None, ("⚠️ tasnif.soliq.uz hozir javob bermayapti.\n"
                  "Kodni qo'lda kiriting yoki keyinroq urinib ko'ring.")

def validate_min_group_text(text):
    """Min guruh: butun son, 2-100 oralig'ida.
    Returns (ok, value, error_msg).
    """
    try:
        mg = int(str(text).strip())
    except (ValueError, TypeError):
        return False, 0, "❌ Butun son kiriting (masalan: 5)"
    if mg < 2 or mg > 100:
        return False, 0, "❌ Minimal guruh 2 dan 100 gacha bo'lishi kerak"
    return True, mg, ''

# ─── CHANNEL HELPERS ────────────────────────────────────────────────
def can_manage_channel(uid, channel):
    ch = verified_channels.get(channel)
    if not ch: return False
    return uid == ch['owner_id'] or uid in ch.get('moderators', [])

def is_channel_admin(uid, channel):
    try:
        result = requests.post(
            f'https://api.telegram.org/bot{SELLER_TOKEN}/getChatMember',
            json={'chat_id': channel, 'user_id': uid}
        ).json()
        if not result.get('ok'): return False
        status = result['result'].get('status', '')
        return status in ('creator', 'administrator')
    except:
        return False

_seller_bot_id_cache = {'id': None}
def get_seller_bot_id():
    """Sotuvchi botning Telegram user_id sini qaytaradi (bir martagina /getMe chaqiriladi)."""
    if _seller_bot_id_cache['id']:
        return _seller_bot_id_cache['id']
    if not SELLER_TOKEN:
        return None
    try:
        r = requests.get(f'https://api.telegram.org/bot{SELLER_TOKEN}/getMe', timeout=5).json()
        if r.get('ok'):
            _seller_bot_id_cache['id'] = r['result'].get('id')
            return _seller_bot_id_cache['id']
    except Exception as e:
        logging.error(f"getMe error: {e}")
    return None

def channel_exists(channel):
    """Kanal mavjudligini getChat orqali tekshiradi."""
    try:
        r = requests.post(
            f'https://api.telegram.org/bot{SELLER_TOKEN}/getChat',
            json={'chat_id': channel}, timeout=5
        ).json()
        return bool(r.get('ok'))
    except:
        return False

def is_bot_admin_in(channel):
    """Sotuvchi bot kanaldagi admin yoki creator ekanligini tekshiradi."""
    bot_id = get_seller_bot_id()
    if not bot_id:
        return False
    try:
        r = requests.post(
            f'https://api.telegram.org/bot{SELLER_TOKEN}/getChatMember',
            json={'chat_id': channel, 'user_id': bot_id}, timeout=5
        ).json()
        if not r.get('ok'):
            return False
        return r['result'].get('status', '') in ('creator', 'administrator')
    except:
        return False

MYPRODUCTS_PER_PAGE = 7

def format_price_short(amount):
    """1_400_000 → '1.4M', 2_000_000 → '2M', 550_000 → '550K', 999 → '999'."""
    try:
        n = int(amount or 0)
    except (TypeError, ValueError):
        return "0"
    if n >= 1_000_000:
        m = n / 1_000_000
        return f"{int(m)}M" if m == int(m) else f"{m:.1f}M"
    if n >= 1_000:
        return f"{n // 1_000}K"
    return str(n)

def _classify_product_status(p):
    """Mahsulot status'ini list va detail uchun yagona qilib qaytaradi.
    Returns {'emoji': str, 'label': str, 'archived': bool}.
    """
    status    = p.get('status', 'active')
    is_active = p.get('is_active', True)
    source    = p.get('source', 'manual')
    if status == 'closed':
        return {'emoji': '🔒', 'label': 'Yopilgan', 'archived': True}
    if source == 'billz' and not is_active:
        return {'emoji': '⏸', 'label': 'Yoqilmagan', 'archived': False}
    ddt = p.get('deadline_dt', '')
    if ddt:
        try:
            if datetime.strptime(ddt, '%Y-%m-%d %H:%M') < datetime.now():
                return {'emoji': '⏰', 'label': 'Muddati tugagan', 'archived': True}
        except (ValueError, TypeError):
            pass
    return {'emoji': '🔥', 'label': 'Aktiv', 'archived': False}

def render_customer_list(uid, cid, page=1, cur_filter='all'):
    """CRM mijozlar ro'yxatini chiqaradi (text handler va callback dispatcher uchun).
    Mavjud `menu_mycustomers` callback ichidagi rendering logikasini qayta ishlatadi.
    """
    if page > 1:
        data = f'crm_page_{page}'
    elif cur_filter != 'all':
        data = f'crm_filter_{cur_filter}'
    else:
        data = 'menu_mycustomers'
    # Soxta callback_query — seller_handle_cb ichida answer_cb('0', ...) silently fail bo'ladi
    seller_handle_cb({'id': '0', 'from': {'id': uid}, 'data': data})

def _truncate_name(name, limit=25):
    if not name:
        return '—'
    return name if len(name) <= limit else name[:limit-1].rstrip() + '…'

def render_myproducts(uid, cid, page=0, mode='active'):
    """Mahsulotlar ro'yxati 1-ustunli inline keyboard bilan.
    mode='active'  — aktiv + Billz draft'lar (default)
    mode='archived' — yopilgan + muddati tugagan (kelajakdagi Arxiv tugmasi uchun)
    """
    pids = []
    for pid in seller_products.get(uid, []):
        if pid not in products: continue
        cls = _classify_product_status(products[pid])
        if mode == 'active' and cls['archived']:
            continue
        if mode == 'archived' and not cls['archived']:
            continue
        pids.append(pid)
    if not pids:
        empty_msg = ("📦 Sizda mahsulot yo'q." if mode == 'active'
                     else "📦 Arxivda mahsulot yo'q.")
        send_seller(cid, empty_msg,
            {'inline_keyboard': [[{'text': "➕ Qo'shish", 'callback_data': 'menu_addproduct'}]]})
        return
    total       = len(pids)
    total_pages = (total + MYPRODUCTS_PER_PAGE - 1) // MYPRODUCTS_PER_PAGE
    page        = max(0, min(page, total_pages - 1))
    start       = page * MYPRODUCTS_PER_PAGE
    chunk       = pids[start:start + MYPRODUCTS_PER_PAGE]

    title = "Mening mahsulotlarim" if mode == 'active' else "Arxiv"
    txt = f"📦 <b>{title}</b> ({total} ta) · Sahifa {page+1}/{total_pages}"

    kb = []
    for pid in chunk:
        p     = products[pid]
        cls   = _classify_product_status(p)
        name  = _truncate_name(p.get('name', '—'), 25)
        count = len(groups.get(pid, []))
        min_g = p.get('min_group', 0)
        # MXIK belgisi — aktiv/draft mahsulot uchun yo'q bo'lsa ⚠️ prefix
        mxik_warn = '⚠️ ' if (cls['label'] in ('Aktiv', 'Yoqilmagan')
                              and not p.get('mxik_code')) else ''
        if cls['label'] == 'Aktiv':
            price = format_price_short(p.get('group_price', 0) or p.get('original_price', 0))
            label = f"{mxik_warn}{name} · {price} · 👥{count}/{min_g}"
        elif cls['label'] == 'Yoqilmagan':
            # Draft: group_price=0 — narx asl narxdan
            price = format_price_short(p.get('original_price', 0))
            label = f"{cls['emoji']} {mxik_warn}{name} · {price} · 👥{count}/{min_g}"
        else:
            # Yopilgan / Muddati tugagan
            label = f"{cls['emoji']} {name} · {cls['label']}"
        kb.append([{'text': label, 'callback_data': f'mp_view_{pid}'}])

    page_cb_prefix = 'mp_page_' if mode == 'active' else 'mp_arch_page_'
    nav = []
    if page > 0:
        nav.append({'text': "⬅️ Oldingi", 'callback_data': f'{page_cb_prefix}{page-1}'})
    if page < total_pages - 1:
        nav.append({'text': "➡️ Keyingi", 'callback_data': f'{page_cb_prefix}{page+1}'})
    nav.append({'text': "🔙 Menyu", 'callback_data': 'back_menu'})
    kb.append(nav)
    send_seller(cid, txt, {'inline_keyboard': kb})

def render_mxik_results(uid, cid, keyword, results, page=0):
    """MXIK qidiruv natijalarini ko'rsatadi (5 ta per sahifa)."""
    PER_PAGE = 5
    total    = len(results)
    start    = page * PER_PAGE
    chunk    = results[start:start + PER_PAGE]
    if not chunk:
        send_seller(cid,
            f"🔍 \"{keyword}\" — natija topilmadi.\n\nBoshqa kalit so'z, qo'lda kod yoki o'tkazib yuborish:",
            {'inline_keyboard': [
                [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                [{'text': "🔄 Qayta qidirish",        'callback_data': 'prod_mxik_again'}],
                [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
            ]})
        return
    txt = (
        f"🔍 \"{keyword}\" — {total} ta natija "
        f"(sahifa {page+1}/{(total + PER_PAGE - 1) // PER_PAGE})\n"
        f"━━━━━━━━━━━━━━━\n\n"
        "Quyidagilardan birini tanlang:"
    )
    kb = []
    for i, item in enumerate(chunk):
        idx = start + i
        kb.append([{
            'text': f"{i+1}. {item['name'][:40]}",
            'callback_data': f'prod_mxik_pick_{idx}',
        }])
    nav = []
    if page > 0:
        nav.append({'text': "◀️", 'callback_data': f'prod_mxik_page_{page-1}'})
    if start + PER_PAGE < total:
        nav.append({'text': "▶️ Yana", 'callback_data': f'prod_mxik_page_{page+1}'})
    if nav:
        kb.append(nav)
    kb.append([
        {'text': "🔢 Qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'},
        {'text': "🔄 Boshqa so'z",     'callback_data': 'prod_mxik_again'},
    ])
    kb.append([{'text': "⏭ O'tkazib yuborish", 'callback_data': 'prod_mxik_skip'}])
    send_seller(cid, txt, {'inline_keyboard': kb})

def render_mxik_confirm(uid, cid, mxik_code, mxik_name, classify=''):
    """Tanlangan MXIK kodni tasdiqlash ekrani."""
    classify_line = f"\n📋 Tasnif: {classify}" if classify else ''
    send_seller(cid,
        f"✅ <b>MXIK tanlandi:</b>\n\n"
        f"🏷 Kod: <code>{mxik_code}</code>\n"
        f"📦 Nomi: {mxik_name}"
        f"{classify_line}\n\n"
        f"Tasdiqlaysizmi?",
        {'inline_keyboard': [
            [{'text': "✅ Ha, tasdiqlash", 'callback_data': 'prod_mxik_confirm'}],
            [{'text': "🔄 Boshqasini qidirish", 'callback_data': 'prod_mxik_again'}],
        ]})

def seller_has_legal(uid):
    """Sotuvchi yuridik ma'lumotlarini to'liq kiritganmi tekshiradi."""
    prof = seller_profiles.get(uid) or {}
    return bool(prof.get('legal_completed_at'))

def _legal_status_label(status):
    return {'yatt': 'YaTT', 'mchj': 'MChJ'}.get(status, '—')

def _format_account(acc):
    """20 raqamli hisobni 4-4-4-4-4 formatida formatlash."""
    if not acc or not acc.isdigit():
        return acc or '—'
    return ' '.join(acc[i:i+4] for i in range(0, len(acc), 4))

def render_legal_summary(uid, cid):
    """Sotuvchining yuridik ma'lumotlarini ko'rsatadi (yoki to'ldirish taklif qiladi)."""
    if not seller_has_legal(uid):
        send_seller(cid,
            "📋 <b>Yuridik ma'lumotlar</b>\n\n"
            "Hali to'ldirilmagan. To'ldirsangiz — Payme split to'lov va fiskal chek "
            "ulansa, birinchilardan bo'lib pul olish imkoniyatiga ega bo'lasiz.",
            {'inline_keyboard': [
                [{'text': "▶️ To'ldirishni boshlash", 'callback_data': 'leg_start'}],
                [{'text': "🔙 Menyu",                  'callback_data': 'back_menu'}],
            ]})
        return
    prof = seller_profiles.get(uid, {})
    director_line = ""
    if prof.get('legal_status') == 'mchj' and prof.get('director_name'):
        director_line = f"👔 Direktor: {prof['director_name']}\n"
    txt = (
        "📋 <b>Yuridik ma'lumotlaringiz</b>\n\n"
        f"📌 Status: <b>{_legal_status_label(prof.get('legal_status'))}</b>\n"
        f"🏛 STIR: <code>{prof.get('stir','—')}</code>\n"
        f"🏦 Bank: <b>{prof.get('bank_name','—')}</b>\n"
        f"💳 Hisob: <code>{_format_account(prof.get('bank_account',''))}</code>\n"
        f"🔢 MFO: <code>{prof.get('bank_mfo','—')}</code>\n"
        f"{director_line}"
        f"\n📅 Tasdiqlangan: {prof.get('legal_completed_at','—')}"
    )
    send_seller(cid, txt, {'inline_keyboard': [
        [{'text': "✏️ O'zgartirish", 'callback_data': 'leg_edit_menu'}],
        [{'text': "🔙 Menyu",         'callback_data': 'back_menu'}],
    ]})

def start_legal_flow(uid, cid, after='menu'):
    """Yuridik ma'lumotlarni kiritish flow'ini boshlaydi.
    after='menu' — leg_confirm da save bo'lganda menyuga qaytaramiz.
    after='channel' — keyin ob_channel step'iga o'tkazamiz (onboarding paytida).
    """
    seller_state[uid] = {'step': 'leg_status', 'leg_after': after}
    send_seller(cid,
        "📋 <b>Yuridik ma'lumotlar — qadam 1/6</b>\n\n"
        "Yuridik statusingiz qanday?\n\n"
        "<i>Bu ma'lumotlar Payme split to'lov va fiskal chek uchun kerak. "
        "Bekor qilish: /cancel</i>",
        {'inline_keyboard': [
            [{'text': "👤 YaTT (yakka tartibdagi tadbirkor)", 'callback_data': 'leg_pick_yatt'}],
            [{'text': "🏢 MChJ (mas'uliyati cheklangan jamiyat)", 'callback_data': 'leg_pick_mchj'}],
        ]})

def render_legal_confirm(uid, cid):
    """Yuridik ma'lumotlarni kiritish so'nggi tasdiqlash ekrani."""
    prof = seller_profiles.get(uid, {})
    status = prof.get('legal_status', '—')
    director_line = ""
    if status == 'mchj' and prof.get('director_name'):
        director_line = f"👔 Direktor: {prof['director_name']}\n"
    txt = (
        "📋 <b>Tasdiqlash — Yuridik ma'lumotlar</b>\n\n"
        f"📌 Status: <b>{_legal_status_label(status)}</b>\n"
        f"🏛 STIR: <code>{prof.get('stir','—')}</code>\n"
        f"🏦 Bank: <b>{prof.get('bank_name','—')}</b>\n"
        f"💳 Hisob: <code>{_format_account(prof.get('bank_account',''))}</code>\n"
        f"🔢 MFO: <code>{prof.get('bank_mfo','—')}</code>\n"
        f"{director_line}\n"
        "Hammasi to'g'rimi?"
    )
    send_seller(cid, txt, {'inline_keyboard': [
        [{'text': "✅ Tasdiqlash",   'callback_data': 'leg_confirm'}],
        [{'text': "✏️ O'zgartirish", 'callback_data': 'leg_edit_menu'}],
    ]})

def finalize_shop_onboarding(uid, cid, s, channel):
    """Onboarding state'ni yakunlab do'konni saqlaydi va sotuvchiga xush kelibsiz xabarini yuboradi."""
    if channel not in verified_channels:
        verified_channels[channel] = {'owner_id': uid, 'moderators': []}
    if uid not in seller_shops:
        seller_shops[uid] = []
    shop = {
        'name':              s['ob_shop_name'],
        'phone':             s['ob_phone'],
        'phone2':            s.get('ob_phone2', ''),
        'address':           s.get('ob_address', ''),
        'social':            s.get('ob_social', {}),
        'delivery':          s.get('ob_delivery', 'pickup'),
        'channel':           channel,
        'verified':          True,
        'onboarding_status': 'active',
    }
    edit_idx = s.get('edit_shop_idx')
    if edit_idx is not None and edit_idx < len(seller_shops[uid]):
        seller_shops[uid][edit_idx] = shop
    else:
        seller_shops[uid].append(shop)
    save_data()
    seller_state.pop(uid, None)
    social_lines = ''
    if shop['social']:
        for k, v in shop['social'].items():
            social_lines += f"\n🔗 {k}: {v}"
    send_seller(cid,
        f"✅ <b>Do'kon profili saqlandi!</b>\n\n"
        f"🏪 {shop['name']}\n📞 {shop['phone']}"
        f"{chr(10)+'📱 '+shop['phone2'] if shop['phone2'] else ''}"
        f"{chr(10)+'📍 '+shop['address'] if shop['address'] else ''}"
        f"{social_lines}\n📢 {channel}\n\n"
        "Endi mahsulot qo'sha olasiz!",
        {'keyboard': [
            [{'text': '📦 Mahsulotlarim'},      {'text': '📋 Buyurtmalar'}],
            [{'text': "➕ Mahsulot qo'shish"},  {'text': '👥 Mijozlar'}],
            [{'text': '📊 Statistika'},         {'text': '🔌 Integratsiyalar'}],
        ], 'resize_keyboard': True, 'is_persistent': True})
    # 2-xabar — Mini App taklifi (inline web_app tugma)
    # Eslatma: web_app inline tugmasi ishlashi uchun BotFather'da bot domeni
    # https://seller.joynshop.uz ga o'rnatilgan bo'lishi kerak (/setdomain).
    send_seller(cid,
        "✅ <b>Onboarding tugadi!</b>\n\n"
        "Endi siz Joynshop sotuvchisi sifatida ishlay olasiz.\n\n"
        "💡 <b>Yangilik:</b> Mahsulotlaringizni kengaytirilgan ko'rinishda — "
        "Sotuvchi paneli orqali ham boshqarishingiz mumkin.",
        {'inline_keyboard': [[
            {'text': "🚀 Sotuvchi panelini ochish",
             'web_app': {'url': 'https://seller.joynshop.uz'}},
        ]]})

def gen_mod_code():
    return 'MOD-' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

# ─── INTEGRATIONS REGISTRY ──────────────────────────────────────────
# Kelajakda yangi integration qo'shish uchun shu ro'yxatga entry qo'shing.
# 'active' bo'lsa — handler='render_billz_menu' (yoki tegishli funksiya nomi)
# 'coming_soon' bo'lsa — bosilganda toast ko'rsatiladi.
INTEGRATIONS = [
    {'id': 'billz',    'name': 'Billz POS',  'icon': '🟢', 'status': 'active',      'handler': 'render_billz_menu'},
    {'id': 'iiko',     'name': 'iiko',       'icon': '🔒', 'status': 'coming_soon', 'handler': None},
    {'id': 'moysklad', 'name': 'MoySklad',   'icon': '🔒', 'status': 'coming_soon', 'handler': None},
    {'id': 'smartup',  'name': 'Smartup',    'icon': '🔒', 'status': 'coming_soon', 'handler': None},
]

# ─── BILLZ INTEGRATION ──────────────────────────────────────────────
BILLZ_BASE_URL          = 'https://api-admin.billz.ai'
BILLZ_ENCRYPTION_KEY    = os.environ.get('BILLZ_ENCRYPTION_KEY', '')
_billz_fernet_cache     = {'fernet': None, 'tried': False}
# In-memory access_token cache: {(seller_uid, shop_idx): {'token': '...', 'fetched_at': dt}}
_billz_access_tokens    = {}

def get_fernet():
    """Fernet shifrlovchini qaytaradi. Kalit yo'q yoki noto'g'ri bo'lsa None."""
    if _billz_fernet_cache['tried']:
        return _billz_fernet_cache['fernet']
    _billz_fernet_cache['tried'] = True
    if not BILLZ_ENCRYPTION_KEY:
        logging.warning("BILLZ_ENCRYPTION_KEY yo'q — Billz integratsiyasi o'chirilgan")
        return None
    try:
        from cryptography.fernet import Fernet
        _billz_fernet_cache['fernet'] = Fernet(BILLZ_ENCRYPTION_KEY.encode())
        return _billz_fernet_cache['fernet']
    except Exception as e:
        logging.error(f"BILLZ_ENCRYPTION_KEY noto'g'ri: {e}")
        return None

def encrypt_token(plain_token):
    """Plain string token ni Fernet bilan shifrlaydi va URL-safe base64 string qaytaradi."""
    f = get_fernet()
    if not f or not plain_token:
        return None
    return f.encrypt(plain_token.encode()).decode('ascii')

def decrypt_token(encrypted_str):
    """Shifrlangan stringni ochib plain token qaytaradi. Xato bo'lsa None."""
    f = get_fernet()
    if not f or not encrypted_str:
        return None
    try:
        return f.decrypt(encrypted_str.encode('ascii')).decode()
    except Exception as e:
        logging.error(f"decrypt_token error: {e}")
        return None

def billz_login(secret_token):
    """secret_token bilan Billz auth qiladi.
    Returns (access_token: str | None, error: str | None).
    """
    if not secret_token:
        return None, "Token bo'sh"
    try:
        r = requests.post(
            f'{BILLZ_BASE_URL}/v1/auth/login',
            json={'secret_token': secret_token},
            timeout=10
        )
        if r.status_code == 401:
            return None, "Secret token noto'g'ri"
        if r.status_code != 200:
            return None, f"Billz xatosi: HTTP {r.status_code}"
        data = r.json() or {}
        token = data.get('data', {}).get('access_token') or data.get('access_token')
        if not token:
            return None, "Billz javobida access_token topilmadi"
        return token, None
    except requests.Timeout:
        return None, "Billz ga ulanish vaqti o'tdi (timeout)"
    except Exception as e:
        logging.error(f"billz_login exception: {e}")
        return None, f"Tarmoq xatosi: {e}"

def _billz_get_access_token(uid, shop_idx, force_refresh=False):
    """Sotuvchi do'koni uchun amaldagi access_token'ni qaytaradi.
    Memory cache yoki secret_token'dan qaytadan auth.
    Returns (access_token: str | None, error: str | None).
    """
    cache_key = (uid, shop_idx)
    if not force_refresh and cache_key in _billz_access_tokens:
        cached = _billz_access_tokens[cache_key]
        # 23 soat — 24 soatlik token uchun xavfsizlik chegarasi
        if (datetime.now() - cached['fetched_at']).total_seconds() < 23 * 3600:
            return cached['token'], None
    shops = seller_shops.get(uid, [])
    if shop_idx >= len(shops):
        return None, "Do'kon topilmadi"
    encrypted = shops[shop_idx].get('billz_secret_token')
    if not encrypted:
        return None, "Billz ulanmagan"
    plain = decrypt_token(encrypted)
    if not plain:
        return None, "Token shifrini ochib bo'lmadi (BILLZ_ENCRYPTION_KEY o'zgarganmi?)"
    token, err = billz_login(plain)
    if not token:
        return None, err
    _billz_access_tokens[cache_key] = {'token': token, 'fetched_at': datetime.now()}
    return token, None

def billz_get(uid, shop_idx, path, params=None):
    """Billz GET so'rovi. 401 bo'lsa avtomatik qayta auth.
    Returns (data: dict | None, error: str | None).
    """
    for attempt in range(2):
        token, err = _billz_get_access_token(uid, shop_idx, force_refresh=(attempt > 0))
        if not token:
            return None, err
        try:
            r = requests.get(
                f'{BILLZ_BASE_URL}{path}',
                headers={'Authorization': f'Bearer {token}'},
                params=params or {},
                timeout=15,
            )
            if r.status_code == 401:
                # Token expired — qayta auth
                _billz_access_tokens.pop((uid, shop_idx), None)
                continue
            if r.status_code != 200:
                return None, f"Billz HTTP {r.status_code}: {r.text[:200]}"
            return r.json(), None
        except requests.Timeout:
            return None, "Billz timeout"
        except Exception as e:
            logging.error(f"billz_get exception: {e}")
            return None, f"Tarmoq xatosi: {e}"
    return None, "Avtorizatsiya muvaffaqiyatsiz"

def billz_extract_shops(products_response):
    """Billz mahsulot javobidan shop ro'yxatini chiqaradi.
    Returns [{'shop_id': str, 'shop_name': str}, ...] yoki [].
    """
    shops_seen = {}
    items = products_response.get('products') or products_response.get('data') or []
    if isinstance(items, dict):
        items = items.get('products', [])
    for prod in items:
        for smv in prod.get('shop_measurement_values', []) or []:
            sid = smv.get('shop_id')
            sname = smv.get('shop_name', sid)
            if sid and sid not in shops_seen:
                shops_seen[sid] = sname
    return [{'shop_id': k, 'shop_name': v} for k, v in shops_seen.items()]

def _billz_extract_price_for_shop(prod, billz_shop_id):
    """Billz mahsulot dict'idan tegishli shop uchun retail narxni topadi."""
    for sp in prod.get('shop_prices', []) or []:
        if str(sp.get('shop_id')) == str(billz_shop_id):
            return int(sp.get('retail_price') or sp.get('price') or 0)
    # Fallback — birinchi shop_price yoki retail_price
    sps = prod.get('shop_prices', []) or []
    if sps:
        return int(sps[0].get('retail_price') or sps[0].get('price') or 0)
    return int(prod.get('retail_price') or 0)

def _billz_extract_stock_for_shop(prod, billz_shop_id):
    """Billz mahsulot dict'idan tegishli shop uchun stock_value ni topadi."""
    for smv in prod.get('shop_measurement_values', []) or []:
        if str(smv.get('shop_id')) == str(billz_shop_id):
            try:
                return int(float(smv.get('active_measurement_value') or smv.get('value') or 0))
            except (TypeError, ValueError):
                return 0
    return 0

def _billz_make_product_dict(prod, uid, shop, channel):
    """Billz mahsulot JSON'idan Joynshop product dict yaratadi."""
    billz_shop_id = shop.get('billz_shop_id', '')
    price = _billz_extract_price_for_shop(prod, billz_shop_id)
    stock = _billz_extract_stock_for_shop(prod, billz_shop_id)
    photo = prod.get('main_image_url') or prod.get('photo_url') or ''
    cats  = prod.get('categories') or []
    cat_name = cats[0].get('name', '') if cats and isinstance(cats[0], dict) else ''
    return {
        # Identifikatsiya
        'billz_id':       prod.get('id') or prod.get('uuid') or '',
        'source':         'billz',
        # Kontent
        'name':           (prod.get('name') or '')[:200],
        'description':    (prod.get('description') or '')[:500],
        'photo_id':       None,
        'photo_ids':      [],
        'photo_url':      photo,
        'photo_urls':     [photo] if photo else [],
        'barcode':        prod.get('barcode', ''),
        'sku':            prod.get('sku', ''),
        'brand_name':     prod.get('brand_name', '') or (prod.get('brand', {}) or {}).get('name', ''),
        'category':       cat_name,
        'category_name':  cat_name,
        # Narx (faqat original — solo/group sotuvchi yoqishda kiritadi)
        'original_price': price,
        'group_price':    0,
        'solo_price':     0,
        'min_group':      0,
        'stock':          stock if stock > 0 else 9999,
        'stock_initial':  stock if stock > 0 else 9999,
        'stock_value':    stock,
        'stock_updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        # Sotuvchi va do'kon
        'seller_id':      uid,
        'seller_channel': channel,
        'shop_name':      shop.get('name', ''),
        'contact':        shop.get('phone', ''),
        'phone2':         shop.get('phone2', ''),
        'address':        shop.get('address', ''),
        'social':         shop.get('social', {}),
        'delivery_type':  shop.get('delivery', 'pickup'),
        'variants':       [],
        'sale_type':      'both',
        # Holat
        'status':         'draft',
        'is_active':      False,
        'solo_available': False,
        'channel_message_id': None,
        'channel_chat_id':    None,
        'deadline':       '',
        'deadline_dt':    '',
    }

def import_billz_products(uid, cid, shop_idx):
    """Background thread'da Billz mahsulotlarini import qiladi."""
    def worker():
        shops = seller_shops.get(uid, [])
        if shop_idx >= len(shops):
            send_seller(cid, "❌ Do'kon topilmadi"); return
        shop = shops[shop_idx]
        channel = shop.get('channel', '')
        send_seller(cid, "⏳ <b>Mahsulotlar yuklab olinmoqda...</b>")

        existing_billz_ids = {
            p.get('billz_id') for pid, p in products.items()
            if p.get('seller_id') == uid and p.get('source') == 'billz'
        }
        imported = 0
        skipped  = 0
        page     = 1
        last_progress = 0
        max_pages = 50  # 5000 mahsulot — havfsizlik chegarasi

        while page <= max_pages:
            data, err = billz_get(uid, shop_idx, '/v2/products',
                                  {'limit': 100, 'page': page})
            if err:
                send_seller(cid,
                    f"❌ <b>Import to'xtadi (sahifa {page}):</b>\n{err}\n\n"
                    f"Hozirgacha {imported} ta import qilindi.")
                return
            items = (data or {}).get('products') \
                    or (data or {}).get('data') \
                    or []
            if isinstance(items, dict):
                items = items.get('products', [])
            if not items:
                break  # Sahifalar tugadi

            for prod in items:
                billz_id = prod.get('id') or prod.get('uuid') or ''
                if not billz_id:
                    continue
                if billz_id in existing_billz_ids:
                    skipped += 1
                    continue
                pdict = _billz_make_product_dict(prod, uid, shop, channel)
                # Joynshop pid — Billz UUID ning birinchi 12 belgisi
                pid = 'bz' + ''.join(c for c in billz_id if c.isalnum())[:10].lower()
                # Collision bo'lsa qo'shimcha qator
                _i = 0
                while pid in products and _i < 5:
                    pid = 'bz' + ''.join(c for c in billz_id if c.isalnum())[:8].lower() + str(_i)
                    _i += 1
                products[pid] = pdict
                groups.setdefault(pid, [])
                seller_products.setdefault(uid, [])
                if pid not in seller_products[uid]:
                    seller_products[uid].append(pid)
                existing_billz_ids.add(billz_id)
                imported += 1

                if imported - last_progress >= 100:
                    send_seller(cid, f"📥 {imported} ta import qilindi...")
                    last_progress = imported

            if len(items) < 100:
                break  # Oxirgi sahifa
            page += 1

        save_data()
        send_seller(cid,
            f"✅ <b>Import tugadi!</b>\n\n"
            f"📦 Yangi: {imported} ta\n"
            f"⏭ O'tkazib yuborilgan (allaqachon bor): {skipped} ta\n\n"
            f"Endi /myproducts orqali har birini yoqing — narx va deadline kiritib kanalga e'lon qilasiz.",
            {'inline_keyboard': [
                [{'text': "📦 Mahsulotlarim", 'callback_data': 'menu_myproducts'}],
                [{'text': "🔌 Billz menyu",   'callback_data': 'billz_menu'}],
            ]})
    threading.Thread(target=worker, daemon=True).start()

def post_to_channel(uid, pid):
    """Mavjud product dict'ni kanalga e'lon qilib message_id'larni saqlaydi.
    publish_product'dan farqli — yangi product yaratmaydi, mavjud pid bilan ishlaydi.
    Returns (ok: bool, error: str | None).
    """
    p = products.get(pid)
    if not p:
        return False, "Mahsulot topilmadi"
    channel = p.get('seller_channel') or ''
    if not channel:
        # Sotuvchi shop'idan kanalni olish
        for shop in seller_shops.get(uid, []):
            if shop.get('channel'):
                channel = shop['channel']
                p['seller_channel'] = channel
                break
    if not channel:
        return False, "Kanal topilmadi"
    caption = post_caption(p, pid)
    count = len(groups.get(pid, []))
    kb = json.dumps(join_kb(pid, count, p.get('min_group', 0),
                            has_solo=bool(p.get('solo_price')),
                            sale_type=p.get('sale_type', 'both')))
    photo_ids  = p.get('photo_ids') or ([p.get('photo_id')] if p.get('photo_id') else [])
    photo_url  = p.get('photo_url') or ''
    try:
        if len(photo_ids) > 1:
            media = []
            for i, fid in enumerate(photo_ids):
                item = {'type': 'photo', 'media': fid}
                if i == 0:
                    item['caption'] = caption
                    item['parse_mode'] = 'HTML'
                media.append(item)
            r = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendMediaGroup',
                              json={'chat_id': channel, 'media': media}, timeout=20).json()
            if r.get('ok') and r.get('result'):
                p['channel_message_id'] = r['result'][0].get('message_id')
                p['channel_chat_id']    = channel
                requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendMessage', json={
                    'chat_id': channel, 'text': caption,
                    'parse_mode': 'HTML', 'reply_markup': kb,
                }, timeout=10)
                return True, None
            return False, f"sendMediaGroup xatosi: {r}"
        # Bitta rasm — file_id yoki URL
        photo_payload = photo_ids[0] if photo_ids else photo_url
        if not photo_payload:
            return False, "Rasm topilmadi"
        r = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendPhoto', json={
            'chat_id': channel, 'photo': photo_payload,
            'caption': caption, 'parse_mode': 'HTML', 'reply_markup': kb,
        }, timeout=20).json()
        if r.get('ok'):
            p['channel_message_id'] = r['result'].get('message_id')
            p['channel_chat_id']    = channel
            return True, None
        return False, f"sendPhoto xatosi: {r}"
    except Exception as e:
        return False, f"Tarmoq xatosi: {e}"

def seller_billz_connected_shops(uid):
    """Sotuvchining Billz ulangan do'konlari indekslarini qaytaradi."""
    return [i for i, sh in enumerate(seller_shops.get(uid, []))
            if sh.get('billz_secret_token')]

def integration_label(entry, uid):
    """Integration uchun dinamik tugma matni (sotuvchi holatiga qarab)."""
    if entry['status'] != 'active':
        return f"🔒 {entry['name']}"
    if entry['id'] == 'billz':
        return ("✅ Billz POS" if seller_billz_connected_shops(uid) else "➕ Billz POS")
    # Boshqa active integratsiyalar — kelajakda shu pattern'ga moslashtiriladi
    return f"➕ {entry['name']}"

def render_integrations_menu(uid, cid):
    """Integratsiyalar ro'yxatini sotuvchining ulanganlik holati bilan ko'rsatadi."""
    txt = "🔌 <b>Integratsiyalar</b>"
    kb = []
    for entry in INTEGRATIONS:
        kb.append([{
            'text':          integration_label(entry, uid),
            'callback_data': f"integ_{entry['id']}",
        }])
    kb.append([{'text': "🔙 Menyu", 'callback_data': 'back_menu'}])
    send_seller(cid, txt, {'inline_keyboard': kb})

def _open_billz_management(uid, cid, shop_idx):
    """Ulangan Billz do'koni uchun boshqaruv menyusi: import/discount/disconnect."""
    shops = seller_shops.get(uid, [])
    if shop_idx >= len(shops):
        send_seller(cid, "❌ Do'kon topilmadi"); return
    shop = shops[shop_idx]
    billz_count = sum(1 for p in products.values()
                      if p.get('seller_id') == uid and p.get('source') == 'billz')
    txt = (
        f"✅ <b>Billz POS — boshqaruv</b>\n\n"
        f"🏪 Joynshop do'koni: <b>{shop.get('name','—')}</b>\n"
        f"🏬 Billz do'koni: <b>{shop.get('billz_shop_name','—')}</b>\n"
        f"📅 Ulangan: {shop.get('billz_connected_at','—')}\n"
        f"📦 Import qilingan: {billz_count} ta\n"
    )
    kb = [
        [{'text': "📥 Mahsulotlarni yangilash",      'callback_data': f'billz_import_{shop_idx}'}],
        [{'text': "⚙️ Global chegirma sozlamalari", 'callback_data': f'billz_disc_{shop_idx}'}],
        [{'text': "🔌 Billz ni o'chirish",            'callback_data': f'billz_disconnect_{shop_idx}'}],
        [{'text': "🔙 Integratsiyalar",              'callback_data': 'menu_integrations'}],
    ]
    send_seller(cid, txt, {'inline_keyboard': kb})

def _start_billz_onboarding(uid, cid):
    """Hech qaysi do'kon ulanmagan — onboarding boshlash.
    1 do'kon bo'lsa to'g'ri token kutiladi, ko'p bo'lsa do'kon tanlash so'raladi.
    """
    if not get_fernet():
        send_seller(cid,
            "⚠️ Billz integratsiyasi server tomonidan o'chirilgan.\n"
            "Admin: BILLZ_ENCRYPTION_KEY env varni tekshiring.")
        return
    shops = seller_shops.get(uid, [])
    if not shops:
        send_seller(cid, "❌ Avval do'kon yarating: /start")
        return
    if len(shops) == 1:
        # Bitta do'kon — to'g'ri token so'raymiz
        seller_state[uid] = {'step': 'billz_secret_token', 'billz_shop_idx': 0}
        send_seller(cid,
            f"🔌 <b>Billz ulash — {shops[0].get('name','')}</b>\n\n"
            f"<b>Qadamlar:</b>\n"
            f"1. Billz UI → Sozlamalar → API → <b>Создать ключ</b>\n"
            f"2. Yaratilgan secret token'ni nusxalang\n"
            f"3. Token'ni shu chatga yuboring\n\n"
            f"⚠️ Token shifrlanib saqlanadi. Bekor qilish: /cancel")
        return
    # Ko'p do'kon — qaysi birini ulashni so'raymiz
    kb = []
    for idx, sh in enumerate(shops):
        connected = bool(sh.get('billz_secret_token'))
        if connected:
            label = f"✅ {sh.get('name','—')[:25]}"
        else:
            label = f"🟢 {sh.get('name','—')[:25]}"
        kb.append([{'text': label, 'callback_data': f'billz_view_{idx}'}])
    kb.append([{'text': "🔙 Integratsiyalar", 'callback_data': 'menu_integrations'}])
    send_seller(cid,
        "🟢 <b>Billz POS — ulash</b>\n\n"
        "Qaysi do'koningizni Billz bilan bog'lashni xohlaysiz?",
        {'inline_keyboard': kb})

def render_billz_menu(uid, cid):
    """Sotuvchining do'konlarini va Billz holatini ko'rsatadi."""
    if not get_fernet():
        send_seller(cid,
            "⚠️ Billz integratsiyasi server tomonidan o'chirilgan.\n"
            "Admin: BILLZ_ENCRYPTION_KEY env varni tekshiring.")
        return
    shops = seller_shops.get(uid, [])
    if not shops:
        send_seller(cid, "❌ Avval do'kon yarating: /start")
        return
    txt = "🔌 <b>Billz integratsiyasi</b>\n\n"
    txt += "Billz POS dan mahsulotlaringizni avtomatik olib kelish.\n\n"
    kb = []
    for idx, shop in enumerate(shops):
        connected = bool(shop.get('billz_secret_token'))
        billz_name = shop.get('billz_shop_name', '')
        if connected:
            label = f"✅ {shop.get('name','—')[:20]} → {billz_name[:15]}"
        else:
            label = f"⚪️ {shop.get('name','—')[:25]} (ulanmagan)"
        kb.append([{'text': label, 'callback_data': f'billz_view_{idx}'}])
    kb.append([{'text': "🔙 Menyu", 'callback_data': 'back_menu'}])
    send_seller(cid, txt, {'inline_keyboard': kb})


def _send_or_edit_prod(cid, s, text, kb):
    """Mahsulot tasdiqlash ekranida xabarni yangilaydi (yangi yubormaydi).
    state['prod_msg_id'] mavjud bo'lsa edit, aks holda yangi yuboradi va saqlaydi.
    Telegram 48s edit cheklovi yoki boshqa xato — fallback: yangi xabar.
    """
    mid = s.get('prod_msg_id') if isinstance(s, dict) else None
    if mid:
        try:
            r = edit_message(cid, mid, text, kb)
            if r and r.get('ok'):
                return r
        except Exception as e:
            logging.warning(f"prod_msg edit failed: {e}")
    # Fallback yoki birinchi xabar — yangi yuborish
    r = send_seller(cid, text, kb)
    if r and isinstance(r, dict) and isinstance(s, dict):
        result = r.get('result', {})
        if result.get('message_id'):
            s['prod_msg_id'] = result['message_id']
    return r

def show_prod_confirm(cid, s, shop):
    orig = s.get('original_price', 0); grp = s.get('group_price', 0)
    disc = round((orig-grp)/orig*100) if orig else 0
    photos = len(s.get('photo_ids', []))
    desc_line     = f"\n📝 {s['description']}" if s.get('description') else ''
    solo_line     = f"\n👤 Yakka: {s['solo_price']:,} so'm" if s.get('solo_price') else ''
    variants_line = f"\n🎨 {', '.join(s['variants'])}" if s.get('variants') else ''
    cat = s.get('category', '')
    cat_icon = next((icon for name, icon in CATEGORIES if name == cat), '📦')
    cat_line  = f"\n{cat_icon} {cat}" if cat else ''
    sale_labels = {'group': '👥 Guruhli', 'solo': '👤 Yakka', 'both': '👥+👤 Ikkalasi'}
    sale_line   = f"\n{sale_labels.get(s.get('sale_type','both'), '')}"
    min_group_line = f"\n👥 Min guruh: {s['min_group']} kishi" if s.get('sale_type') != 'solo' else ''
    # Deadline
    hours = int(s.get('deadline_hours', 48))
    deadline_labels = {24: '24 soat', 48: '2 kun', 72: '3 kun', 168: '1 hafta'}
    deadline_line = f"\n⏰ Muddat: {deadline_labels.get(hours, str(hours)+' soat')}"
    mxik_code = s.get('mxik_code')
    mxik_line = f"\n🏷 MXIK: <code>{mxik_code}</code>" if mxik_code else "\n🏷 MXIK: ⚠️ Yo'q"
    # Deadline tugmalari — hozirgi tanlangan belgilangan
    def dl_btn(h, label):
        mark = '✅ ' if hours == h else ''
        return {'text': f"{mark}{label}", 'callback_data': f'prod_deadline_{h}'}
    text = (
        f"📋 <b>Mahsulotni tekshiring:</b>\n\n"
        f"📦 <b>{s['name']}</b>\n🏪 {shop.get('name','')}"
        f"{cat_line}{sale_line}\n"
        f"📸 {photos} ta rasm\n💰 {orig:,} → {grp:,} so'm (-{disc}%)"
        f"{min_group_line}\n📢 {shop.get('channel','—')}"
        f"{deadline_line}{mxik_line}"
        f"{desc_line}{solo_line}{variants_line}"
    )
    kb = {'inline_keyboard': [
        [{'text': "🚀 E'lon qilish!", 'callback_data': 'prod_confirm_publish'}],
        [dl_btn(24,'24 soat'), dl_btn(48,'2 kun'), dl_btn(72,'3 kun'), dl_btn(168,'1 hafta')],
        [{'text': "📝 Tavsif",        'callback_data': 'prod_add_desc'},
         {'text': "💰 Yakka narx",    'callback_data': 'prod_add_solo'}],
        [{'text': "🎨 Variantlar",    'callback_data': 'prod_add_variants'}],
    ]}
    _send_or_edit_prod(cid, s, text, kb)

def show_confirm(cid, s):
    channel = s.get('seller_channel', '')
    send_seller(cid,
        f"📋 <b>Ma'lumotlarni tekshiring:</b>\n\n"
        f"📦 Mahsulot: <b>{s['name']}</b>\n"
        f"🏪 Do'kon: <b>{s['shop_name']}</b>\n"
        f"📝 Tavsif: {s['description']}\n\n"
        f"💰 Asl narx: <b>{fmt(s['original_price'])} so'm</b>\n"
        f"👤 Yakka narx: <b>{fmt(s.get('solo_price',0)) if s.get('solo_price') else 'Yoq'} so'm</b>\n"
        f"🏷 Guruh narxi: <b>{fmt(s['group_price'])} so'm</b>\n\n"
        f"👥 Minimal guruh: <b>{s['min_group']} kishi</b>\n"
        f"📞 Aloqa: <b>{s['contact']}</b>\n"
        f"📢 Kanal: <b>{channel}</b>",
        {'inline_keyboard': [
            [{'text': "✅ To'g'ri, e'lon qil!", 'callback_data': 'confirm_product'}],
            [{'text': "✏️ Nom",         'callback_data': 'edit_name'},
             {'text': "✏️ Do'kon",      'callback_data': 'edit_shop_name'}],
            [{'text': "✏️ Tavsif",      'callback_data': 'edit_description'},
             {'text': "✏️ Asl narx",    'callback_data': 'edit_original_price'}],
            [{'text': "✏️ Guruh narxi", 'callback_data': 'edit_group_price'},
             {'text': "✏️ Min guruh",   'callback_data': 'edit_min_group'}],
            [{'text': "✏️ Rasm",        'callback_data': 'edit_photo'},
             {'text': "✏️ Aloqa",       'callback_data': 'edit_contact'}],
            [{'text': "✏️ Kanal",       'callback_data': 'edit_seller_channel'}],
        ]}
    )

def publish_product(uid, cid, s):
    shop_idx = s.get('shop_idx')
    if shop_idx is not None:
        shops    = seller_shops.get(uid, [])
        shop     = shops[shop_idx] if shop_idx < len(shops) else {}
        channel  = shop.get('channel',  s.get('seller_channel',''))
        contact  = shop.get('phone',    s.get('contact',''))
        phone2   = shop.get('phone2',   '')
        address  = shop.get('address',  '')
        social   = shop.get('social',   {})
        delivery = shop.get('delivery', s.get('delivery_type','pickup'))
        shop_name= shop.get('name',     s.get('shop_name',''))
    else:
        channel  = s.get('seller_channel','')
        contact  = s.get('contact','')
        phone2   = ''
        address  = ''
        social   = {}
        delivery = s.get('delivery_type','pickup')
        shop_name= s.get('shop_name','')

    pid      = ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))
    deadline = datetime.now() + timedelta(hours=int(s.get('deadline_hours', 48)))
    photo_ids  = s.get('photo_ids') or ([s['photo_id']] if s.get('photo_id') else [])
    photo_urls = s.get('photo_urls', [])
    first_photo = photo_ids[0] if photo_ids else None
    first_url   = photo_urls[0] if photo_urls else None
    if not first_photo:
        send_seller(cid, "❌ Rasm topilmadi!"); return

    products[pid] = {
        'name':           s['name'],
        'shop_name':      shop_name,
        'description':    s.get('description', ''),
        'original_price': s['original_price'],
        'group_price':    s['group_price'],
        'solo_price':     s.get('solo_price', 0),
        'min_group':      s['min_group'],
        'stock':          s.get('stock', 9999),
        'stock_initial':  s.get('stock', 9999),
        'photo_id':       first_photo,
        'photo_ids':      photo_ids,
        'photo_url':      first_url,
        'photo_urls':     photo_urls,
        'contact':        contact,
        'phone2':         phone2,
        'address':        address,
        'social':         social,
        'delivery_type':  delivery,
        'variants':       s.get('variants', []),
        'category':       s.get('category', ''),
        'sale_type':      s.get('sale_type', 'both'),
        'seller_channel': channel,
        'seller_id':      uid,
        'deadline':       deadline.strftime('%d.%m.%Y %H:%M'),
        'deadline_dt':    deadline.strftime('%Y-%m-%d %H:%M'),
        'channel_message_id': None,
        'channel_chat_id':    None,
        'status':         'active',
        'is_active':      True,
        'solo_available': True,
        'mxik_code':      s.get('mxik_code'),
        'mxik_name':      s.get('mxik_name'),
    }
    groups[pid] = []
    if uid not in seller_products: seller_products[uid] = []
    seller_products[uid].append(pid)

    caption = post_caption(products[pid], pid)
    kb      = json.dumps(join_kb(pid, 0, s['min_group'], has_solo=bool(s.get('solo_price')), sale_type=s.get('sale_type','both')))

    if len(photo_ids) > 1:
        media = []
        for i, fid in enumerate(photo_ids):
            item = {'type': 'photo', 'media': fid}
            if i == 0:
                item['caption'] = caption
                item['parse_mode'] = 'HTML'
            media.append(item)
        result = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendMediaGroup', json={
            'chat_id': channel, 'media': media,
        }).json()
        if result.get('ok') and result.get('result'):
            first_msg = result['result'][0]
            products[pid]['channel_message_id'] = first_msg.get('message_id')
            products[pid]['channel_chat_id']    = channel
            # Caption + tugmalar alohida xabar (media group da reply_markup ishlamaydi)
            requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendMessage', json={
                'chat_id': channel, 'text': caption,
                'parse_mode': 'HTML', 'reply_markup': kb,
            })
    else:
        result = requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/sendPhoto', json={
            'chat_id': channel, 'photo': first_photo,
            'caption': caption, 'parse_mode': 'HTML', 'reply_markup': kb,
        }).json()
        if result.get('ok'):
            msg_result = result['result']
            products[pid]['channel_message_id'] = msg_result.get('message_id')
            products[pid]['channel_chat_id']    = channel

    del seller_state[uid]
    if result.get('ok'):
        save_data()
        send_seller(cid,
            f"✅ <b>E'lon qilindi!</b>\n\n"
            f"📦 {s['name']}\n📢 {channel}\n"
            f"🆔 <code>{pid}</code>\n⏰ {deadline.strftime('%d.%m.%Y %H:%M')}",
            {'inline_keyboard': [
                [{'text': '📊 Statistika',  'callback_data': 'menu_mystats'},
                 {'text': "📢 Qayta e'lon", 'callback_data': f'boost_{pid}'}],
                [{'text': '✏️ Tahrirlash',  'callback_data': f'edit_prod_{pid}'},
                 {'text': "🗑 O'chirish",   'callback_data': f'delete_prod_{pid}'}],
            ]}
        )
        # 5-mahsulotda yumshoq eslatma — yuridik to'ldirilmaganlar uchun
        prod_count = len(seller_products.get(uid, []))
        if prod_count and prod_count % 5 == 0 and not seller_has_legal(uid):
            send_seller(cid,
                "💡 <b>Eslatma:</b> Tez orada to'lov tizimi (Payme split, fiskal chek) ulanadi.\n\n"
                "Yuridik ma'lumotlaringizni to'ldirsangiz — birinchilardan bo'lib pul olish "
                "imkoniyatiga ega bo'lasiz.",
                {'inline_keyboard': [
                    [{'text': "▶️ Hozir to'ldirish", 'callback_data': 'leg_start'}],
                ]})
    else:
        del products[pid]
        seller_products[uid].remove(pid)
        s['step'] = 'confirm'
        seller_state[uid] = s
        send_seller(cid,
            f"❌ Kanalga post qo'yib bo'lmadi!\n\n"
            f"Tekshiring:\n"
            f"• Sotuvchi bot {channel} ga admin sifatida qo'shilganmi?\n"
            f"• Kanal username to'g'rimi?\n\n"
            f"Tahrirlash yoki qayta urinish:"
        )
        show_confirm(cid, s)

def seller_handle_msg(msg):
    cid  = msg['chat']['id']
    uid  = msg['from']['id']
    text = msg.get('text', '')

    # ─── GLOBAL ESCAPE COMMANDS ────────────────────────────────
    # /cancel — har qanday step'dan chiqib menyuga qaytadi
    # /menu   — back_menu ekranini chiqaradi (state'ga tegmasdan)
    # Bu /start dan ham, prod_* dan ham, ob_* dan ham ishlaydi.
    if text == '/cancel':
        seller_state.pop(uid, None)
        send_seller(cid,
            "❌ Bekor qilindi.\n\n🏪 <b>Bosh menyu:</b>",
            {'inline_keyboard': [
                [{'text': "🏠 Menyu ochish", 'callback_data': 'back_menu'}],
            ]})
        return
    if text == '/menu':
        send_seller(cid,
            "🏪 <b>Joynshop Sotuvchi Paneli</b>",
            {'inline_keyboard': [
                [{'text': "📦 Mahsulotlarim",      'callback_data': 'menu_myproducts'},
                 {'text': "📋 Buyurtmalar",        'callback_data': 'menu_myorders'}],
                [{'text': "➕ Mahsulot qo'shish",  'callback_data': 'menu_addproduct'},
                 {'text': "👥 Mijozlar",            'callback_data': 'menu_mycustomers'}],
                [{'text': "📊 Statistika",         'callback_data': 'menu_mystats'},
                 {'text': "🔌 Integratsiyalar",    'callback_data': 'menu_integrations'}],
            ]})
        return

    if is_prod_in_progress(uid) and text in PROD_BLOCKED_TEXTS:
        send_seller(cid, get_prod_progress_text(uid),
            {'inline_keyboard': [
                [{'text': "▶️ Davom etish", 'callback_data': 'prod_continue'}],
                [{'text': "🗑 Bekor qilish", 'callback_data': 'prod_restart'}],
            ]}
        )
        return

    if uid == ADMIN_ID and text == '/stats':
        conf   = sum(1 for o in orders.values() if o['status'] == 'confirmed')
        rev    = sum(o['amount'] for o in orders.values() if o['status'] == 'confirmed')
        active = sum(1 for p in products.values() if p.get('status') != 'closed')
        send_seller(cid,
            f"📊 <b>Umumiy statistika</b>\n\n"
            f"📦 Aktiv mahsulotlar: {active}\n"
            f"✅ Tasdiqlangan buyurtmalar: {conf}\n"
            f"💰 Jami aylanma: {fmt(rev)} so'm\n"
            f"📊 Komissiya (5%): {fmt(int(rev*COMMISSION_RATE))} so'm"
        )
        return

    if text == '/start':
        shops = seller_shops.get(uid) or seller_shops.get(str(uid), [])
        is_new = not shops and str(uid) not in [str(k) for k in seller_shops.keys()]
        if is_new:
            cur_state = seller_state.get(uid, {})
            if not cur_state.get('step','').startswith('ob_'):
                seller_state[uid] = {'step': 'ob_shop_name'}
                send_seller(cid,
                    "🏪 <b>Joynshop Sotuvchi Paneliga xush kelibsiz!</b>\n\n"
                    "Bir marta profilingizni to'ldiring.\n\n"
                    "<b>1/4</b> Do'kon nomini yozing:\n<i>Masalan: Nike Toshkent</i>"
                )
            else:
                send_seller(cid, "📝 Do'kon ma'lumotlarini kiritishni davom eting.")
        else:
            shop_names = ', '.join(s['name'] for s in shops) if shops else ''
            send_seller(cid,
                f"🏪 <b>Joynshop Sotuvchi Paneli</b>\n\n"
                f"{'🏬 ' + shop_names + chr(10) if shop_names else ''}"
                f"Guruh savdosi orqali ko'proq soting!",
                {'keyboard': [
                    [{'text': '📦 Mahsulotlarim'},      {'text': '📋 Buyurtmalar'}],
                    [{'text': '➕ Mahsulot qo\'shish'}, {'text': '👥 Mijozlar'}],
                    [{'text': '📊 Statistika'},         {'text': '🔌 Integratsiyalar'}],
                ], 'resize_keyboard': True, 'is_persistent': True}
            )
        return

    if text == '/myproducts' or text == '📦 Mahsulotlarim':
        render_myproducts(uid, cid, page=0)
        return

    if text == '/mystats' or text == '📊 Statistika':
        my = seller_products.get(uid, [])
        if not my:
            send_seller(cid, "📊 Statistika yo'q.\n\n/addproduct — mahsulot qo'shing!"); return
        revenue    = sum(o['amount'] for o in orders.values() if o.get('product_id') in my and o['status'] == 'confirmed')
        commission = int(revenue * COMMISSION_RATE)
        send_seller(cid,
            f"📊 <b>Sizning statistikangiz:</b>\n\n"
            f"📦 Jami mahsulot: {len(my)}\n"
            f"🔥 Aktiv: {sum(1 for pid in my if products.get(pid,{}).get('status')!='closed')}\n"
            f"✅ Muvaffaqiyatli guruh: {sum(1 for pid in my if len(groups.get(pid,[]))>=products.get(pid,{}).get('min_group',99))}\n"
            f"👥 Jami qo'shilgan: {sum(len(groups.get(pid,[])) for pid in my)}\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"💰 Jami sotuv: {fmt(revenue)} so'm\n"
            f"📊 Komissiya (5%): {fmt(commission)} so'm\n"
            f"✅ Sof daromad: {fmt(revenue-commission)} so'm",
            {'inline_keyboard': [[{'text': "📑 Excel eksport", 'callback_data': 'menu_export'}]]}
        )
        return

    if text == '/myorders' or text == '📋 Buyurtmalar':
        my_pids = seller_products.get(uid, [])
        pending = {k:v for k,v in orders.items() if v.get('product_id') in my_pids and v['status']=='confirming'}
        if not pending:
            send_seller(cid, "📋 Tasdiqlanmagan buyurtma yo'q."); return
        for code, o in list(pending.items())[-10:]:
            p = products.get(o['product_id'], {})
            send_seller(cid,
                f"🔔 <b>YANGI TO'LOV!</b>\n\n"
                f"📦 {p.get('name','')}\n👤 {o['user_name']}\n"
                f"💰 {fmt(o['amount'])} so'm\n"
                f"🛒 {'Yakka' if o.get('type')=='solo' else 'Guruh'}\n🆔 #{code}",
                {'inline_keyboard': [[
                    {'text': '✅ Tasdiqlash', 'callback_data': f'seller_ac_{code}'},
                    {'text': '❌ Rad',        'callback_data': f'seller_ar_{code}'}
                ]]}
            )
        return

    if text.startswith('/boost'):
        parts = text.split()
        if len(parts) < 2:
            send_seller(cid, "❌ Format: /boost [ID]"); return
        pid = parts[1]
        if pid not in products:
            send_seller(cid, '❌ Topilmadi!'); return
        p = products[pid]
        if p.get('seller_id') != uid and uid != ADMIN_ID:
            send_seller(cid, '❌ Bu sizning mahsulotingiz emas!'); return
        count = len(groups.get(pid, []))
        send_seller(cid,
            f"📢 <b>{p['name']}</b> qayta e'lon qilasizmi?\n\n"
            f"👥 Hozir: {count}/{p['min_group']} kishi\n💰 {fmt(p['group_price'])} so'm",
            {'inline_keyboard': [[
                {'text': "✅ E'lon qil", 'callback_data': f'boost_confirm_{pid}'},
                {'text': "❌ Yo'q",      'callback_data': 'noop'}
            ]]}
        )
        return

    if text.startswith('/delete'):
        parts = text.split()
        if len(parts) < 2:
            send_seller(cid, "❌ Format: /delete [ID]"); return
        pid = parts[1]
        if pid not in products:
            send_seller(cid, '❌ Topilmadi!'); return
        p = products[pid]
        if p.get('seller_id') != uid and uid != ADMIN_ID:
            send_seller(cid, "❌ Ruxsat yo'q!"); return
        p['status']    = 'closed'
        p['is_active'] = False
        ch_cid = p.get('channel_chat_id'); ch_mid = p.get('channel_message_id')
        if ch_cid and ch_mid:
            try:
                requests.post(
                    f'https://api.telegram.org/bot{SELLER_TOKEN}/deleteMessage',
                    json={'chat_id': ch_cid, 'message_id': ch_mid}, timeout=5
                )
            except Exception as e:
                logging.error(f"deleteMessage error: {e}")
        if uid in seller_products and pid in seller_products[uid]:
            seller_products[uid].remove(pid)
        save_data()
        send_seller(cid, f"✅ <b>{p['name']}</b> o'chirildi.")
        return

    if text == '/integrations' or text == '🔌 Integratsiyalar':
        render_integrations_menu(uid, cid)
        return

    if text == '/mycustomers' or text == '👥 Mijozlar':
        render_customer_list(uid, cid)
        return

    if text == '/legal':
        render_legal_summary(uid, cid)
        return

    if text == '/billz' or text == '🔌 Billz':
        # Backward compatibility — /billz to'g'ridan Billz menyuga olib boradi
        render_billz_menu(uid, cid)
        return

    if text == '/help':
        send_seller(cid,
            "📘 <b>Sotuvchi yordam</b>\n\n"
            "/start       — 🏠 Bosh sahifa\n"
            "/addproduct  — ➕ Mahsulot qo'shish\n"
            "/myproducts  — 📦 Mahsulotlarim\n"
            "/myorders    — 📋 Buyurtmalar\n"
            "/mystats     — 📊 Statistika\n"
            "/billz       — 🔌 Billz integratsiyasi\n"
            "/legal       — 📋 Yuridik ma'lumotlar\n"
            "/menu        — 📱 Bosh menyu\n"
            "/cancel      — ❌ Bekor qilish\n"
            "/help        — ℹ️ Yordam\n\n"
            "💬 Yordam: @joynshop_support"
        )
        return

    if text == '/golive' or text == "🔴 Live":
        my_pids = seller_products.get(uid, [])
        active_prods = [(pid, products[pid]) for pid in my_pids
                        if pid in products and products[pid].get('status') == 'active']
        if not active_prods:
            send_seller(cid,
                "❌ Avval faol mahsulot qo'shing.\n\n/addproduct yozing."
            )
            return
        # Tanlov klaviaturasi
        kb = []
        for pid, p in active_prods[:10]:
            kb.append([{'text': f"📦 {p.get('name','')[:30]}",
                        'callback_data': f'live_pick_{pid}'}])
        kb.append([{'text': "❌ Bekor", 'callback_data': 'live_cancel'}])
        send_seller(cid,
            "🔴 <b>LIVE boshlash</b>\n\n"
            "Qaysi mahsulot uchun Live qilasiz?",
            {'inline_keyboard': kb}
        )
        return

    if text == '/mylive':
        # Sotuvchining faol live'larini ko'rsatish
        my_lives = [(lid, lv) for lid, lv in lives.items()
                    if lv.get('seller_id') == uid and lv.get('status') == 'live']
        if not my_lives:
            send_seller(cid,
                "🔴 <b>Live Dashboard</b>\n\n"
                "Hozir faol live yo'q.\n"
                "Yangi boshlash uchun /golive yozing."
            )
            return
        text_msg = "🔴 <b>Faol Liveler</b>\n━━━━━━━━━━━━━━━\n\n"
        kb = []
        for lid, lv in my_lives:
            p = products.get(lv.get('product_id',''), {})
            text_msg += (
                f"📦 {p.get('name','—')}\n"
                f"👀 {lv.get('viewer_count',0)} ko'rdi\n"
                f"👥 {len(lv.get('joiners',[]))} qo'shildi\n\n"
            )
            kb.append([{'text': f"📊 {p.get('name','')[:25]}", 'callback_data': f'live_dash_{lid}'}])
        send_seller(cid, text_msg, {'inline_keyboard': kb})
        return

    if text == '/addproduct' or text == "➕ Mahsulot qo'shish":
        if is_prod_in_progress(uid):
            send_seller(cid, get_prod_progress_text(uid),
                {'inline_keyboard': [
                    [{'text': '▶️ Davom etish', 'callback_data': 'prod_continue'}],
                    [{'text': '🗑 Bekor qilish', 'callback_data': 'prod_restart'}],
                ]}
            )
            return
        shops = seller_shops.get(uid, [])
        if not shops:
            send_seller(cid, "❌ Avval do'kon profilingizni to'ldiring.\n\n/start yozing.")
            return
        if len(shops) == 1:
            seller_state[uid] = {
                'step': 'prod_name', 'shop_idx': 0,
                'shop_name':      shops[0]['name'],
                'contact':        shops[0]['phone'],
                'delivery_type':  shops[0].get('delivery','pickup'),
                'seller_channel': shops[0].get('channel',''),
            }
            send_seller(cid,
                f"📦 <b>{shops[0]['name']}</b> uchun yangi mahsulot\n\n"
                f"<b>1/4</b> Mahsulot nomini yozing:\n<i>Masalan: Nike Air Max 270</i>"
            )
        else:
            btns = [[{'text': f"🏪 {s['name']}", 'callback_data': f"sel_shop_{i}"}] for i,s in enumerate(shops)]
            send_seller(cid, "Qaysi do'kon uchun mahsulot qo'shmoqchisiz?", {'inline_keyboard': btns})
        return

    if text in ('/mychannels', '/shops', "📢 Do'konlarim", '📢 Kanallarim'):
        shops = seller_shops.get(uid, [])
        if shops:
            r = "🏪 <b>Do'konlaringiz:</b>\n\n"
            btns = []
            for i, sh in enumerate(shops):
                social_text = ' | '.join(f"{k}: {v}" for k,v in sh.get('social',{}).items())
                r += (f"━━━━━━━━━━━━━━\n"
                      f"🏪 <b>{sh['name']}</b>\n"
                      f"📞 {sh['phone']}"
                      + (f"\n📱 {sh['phone2']}" if sh.get('phone2') else '')
                      + (f"\n📍 {sh['address']}" if sh.get('address') else '')
                      + (f"\n🌐 {social_text}" if social_text else '')
                      + f"\n📢 {sh.get('channel','—')}\n")
                btns.append([{'text': f"✏️ {sh['name']} ni tahrirlash", 'callback_data': f'edit_shop_{i}'}])
            btns.append([{'text': "➕ Yangi do'kon qo'shish", 'callback_data': 'add_new_shop'}])
            send_seller(cid, r, {'inline_keyboard': btns})
            return

    if text == '/addmoderator':
        my_channels = [ch for ch, data in verified_channels.items() if data['owner_id'] == uid]
        if not my_channels:
            send_seller(cid, "❌ Siz hech bir kanalning egasi emassiz.\n\nAvval mahsulot qo'shib, kanal tasdiqlang.")
            return
        if len(my_channels) == 1:
            seller_state[uid] = {'step': 'add_mod_user', 'mod_channel': my_channels[0]}
            send_seller(cid,
                f"🛡 <b>{my_channels[0]}</b> uchun moderator qo'shish\n\n"
                f"Moderatorning Telegram @username ini yozing:\n"
                f"<i>Masalan: @username</i>\n\n"
                f"⚠️ U avval seller botni ishga tushirgan bo'lishi kerak!"
            )
        else:
            btns = [[{'text': ch, 'callback_data': f'addmod_ch_{ch}'}] for ch in my_channels]
            send_seller(cid, "Qaysi kanal uchun moderator qo'shmoqchisiz?", {'inline_keyboard': btns})
        return

    if text.startswith('MOD-'):
        code = text.strip()
        if code not in pending_moderator_codes:
            send_seller(cid, "❌ Kod topilmadi yoki muddati o'tgan."); return
        data    = pending_moderator_codes[code]
        channel = data['channel']
        if channel in verified_channels:
            if uid not in verified_channels[channel]['moderators']:
                verified_channels[channel]['moderators'].append(uid)
        del pending_moderator_codes[code]
        save_data()
        send_seller(cid,
            f"✅ <b>{channel}</b> kanalida moderator bo'ldingiz!\n\n"
            f"Endi /addproduct orqali mahsulot qo'sha olasiz."
        )
        owner_id = verified_channels.get(channel, {}).get('owner_id')
        if owner_id:
            send_seller(owner_id,
                f"🛡 Yangi moderator qo'shildi!\n\n"
                f"Kanal: {channel}\n"
                f"Moderator ID: <code>{uid}</code>"
            )
        return

    if uid in seller_state:
        s    = seller_state[uid]
        step = s.get('step')

        # ─── LIVE: video qabul qilish ───
        if step == 'live_video':
            video = msg.get('video') or msg.get('document')
            if not video:
                send_seller(cid, "❌ Iltimos, video yuboring (matn emas).")
                return
            duration = video.get('duration', 30)  # document da duration bo'lmasligi mumkin
            file_id  = video.get('file_id', '')
            if not file_id:
                send_seller(cid, "❌ Video fayl topilmadi, qayta yuboring.")
                return
            s['video_file_id']  = file_id
            s['video_duration'] = duration
            s['step'] = 'live_duration'
            send_seller(cid,
                f"✅ Video qabul qilindi ({duration} soniya)\n\n"
                f"⏰ <b>Live davomiyligi?</b>",
                {'inline_keyboard': [
                    [{'text': "1 soat",  'callback_data': 'live_dur_1'},
                     {'text': "3 soat",  'callback_data': 'live_dur_3'},
                     {'text': "24 soat", 'callback_data': 'live_dur_24'}],
                    [{'text': "❌ Bekor", 'callback_data': 'live_cancel'}],
                ]}
            )
            return

        # ─── CRM: Mijozga xabar yuborish ───
        if step == 'crm_send_msg':
            if text == '/cancel':
                del seller_state[uid]
                send_seller(cid, "❌ Bekor qilindi.")
                return
            target_uid = s.get('target_uid')
            target_name = s.get('target_name')
            del seller_state[uid]
            shop_name = (seller_shops.get(uid) or [{}])[0].get('name', '')
            try:
                send_buyer(target_uid,
                    f"💬 <b>{shop_name}</b> dan xabar:\n\n{text}\n\n"
                    f"<i>Joynshop orqali yuborildi</i>"
                )
                send_seller(cid, f"✅ Xabar <b>{target_name}</b> ga yuborildi.")
            except Exception as e:
                send_seller(cid, f"❌ Xabar yuborilmadi: {e}")
            return

        # ─── CRM: Mijozga izoh qo'shish ───
        if step == 'crm_add_note':
            if text == '/cancel':
                del seller_state[uid]
                send_seller(cid, "❌ Bekor qilindi.")
                return
            sid = str(uid)
            cuid = s.get('target_cuid')
            target_name = s.get('target_name')
            del seller_state[uid]
            if sid in customers and cuid in customers[sid]:
                customers[sid][cuid]['note'] = text[:300]
                save_data()
                send_seller(cid, f"✅ Izoh <b>{target_name}</b> uchun saqlandi.",
                    {'inline_keyboard': [[{'text': "👤 Mijozga qaytish", 'callback_data': 'crm_view_'+cuid}]]}
                )
            else:
                send_seller(cid, "❌ Mijoz topilmadi.")
            return

        # ─── CRM: Qidiruv ───
        if step == 'crm_search_query':
            del seller_state[uid]
            sid = str(uid)
            my_customers = customers.get(sid, {})
            q = text.lower().strip()
            if not q:
                send_seller(cid, "❌ Qidiruv so'rovi bo'sh.")
                return
            results = []
            for cuid, c in my_customers.items():
                if q in c.get('name','').lower() or q in str(c.get('user_id','')).lower():
                    results.append((cuid, c))
            if not results:
                send_seller(cid, f"🔍 \"{text}\" bo'yicha hech narsa topilmadi.",
                    {'inline_keyboard': [[{'text': "⬅️ CRM", 'callback_data': 'menu_mycustomers'}]]}
                )
                return
            kb = []
            txt = f"🔍 <b>Qidiruv natijasi: {len(results)} ta</b>\n\n"
            for cuid, c in results[:15]:
                badges = " ⭐" if 'vip' in c.get('tags', []) else ""
                txt += f"• <b>{c['name']}{badges}</b> — {fmt(c['total_spent'])} so'm\n"
                kb.append([{'text': "👤 " + c['name'], 'callback_data': 'crm_view_' + cuid}])
            kb.append([{'text': "⬅️ CRM", 'callback_data': 'menu_mycustomers'}])
            send_seller(cid, txt, {'inline_keyboard': kb})
            return

        if step == 'add_mod_user':
            username = text if text.startswith('@') else f'@{text}'
            channel  = s.get('mod_channel')
            code     = gen_mod_code()
            pending_moderator_codes[code] = {'channel': channel, 'added_by': uid}
            del seller_state[uid]
            send_seller(cid,
                f"✅ Moderator uchun kod yaratildi!\n\n"
                f"Quyidagi kodni <b>{username}</b> ga yuboring:\n\n"
                f"<code>{code}</code>\n\n"
                f"U seller botga shu kodni yuborishi kerak.\n"
                f"Kod 24 soat amal qiladi."
            )
            return

        if step == 'ob_shop_name':
            s['ob_shop_name'] = text; s['step'] = 'ob_phone'
            s.pop('ob_msg_id', None)
            send_or_edit_seller(cid,
                "📞 Asosiy telefon raqam:\n<i>+998XXXXXXXXX</i>",
                state=s)

        elif step == 'ob_phone':
            s['ob_phone'] = text.strip(); s['step'] = 'ob_phone2'
            s.pop('ob_msg_id', None)
            send_or_edit_seller(cid,
                "📱 Qo'shimcha telefon (ixtiyoriy):\n<i>+998XXXXXXXXX</i>",
                {'inline_keyboard': [[{'text': "⏭ O'tkazib yuborish", 'callback_data': 'ob_skip_phone2'}]]},
                state=s)

        elif step == 'ob_phone2':
            s['ob_phone2'] = text.strip() if text != '/skip' else ''
            s['step'] = 'ob_address'
            s.pop('ob_msg_id', None)
            send_or_edit_seller(cid,
                "📍 Do'kon manzili (ixtiyoriy):\n<i>Toshkent, Chilonzor, 3-mavze</i>",
                {'inline_keyboard': [[{'text': "⏭ O'tkazib yuborish", 'callback_data': 'ob_skip_address'}]]},
                state=s)

        elif step == 'ob_address':
            s['ob_address'] = text.strip() if text != '/skip' else ''
            s['step'] = 'ob_social'
            s.pop('ob_msg_id', None)
            send_or_edit_seller(cid,
                "🌐 Ijtimoiy tarmoqlar (ixtiyoriy):\n"
                "<code>instagram: @dokon_uz\ntelegram: @kanal\nwebsite: dokon.uz</code>",
                {'inline_keyboard': [[{'text': "⏭ O'tkazib yuborish", 'callback_data': 'ob_skip_social'}]]},
                state=s)

        elif step == 'ob_social':
            if text != '/skip':
                social = {}
                for line in text.strip().splitlines():
                    if ':' in line:
                        k, v = line.split(':', 1)
                        social[k.strip().lower()] = v.strip()
                s['ob_social'] = social
            else:
                s['ob_social'] = {}
            s['step'] = 'ob_delivery'
            s.pop('ob_msg_id', None)
            send_or_edit_seller(cid,
                "🚚 Yetkazib berish turini tanlang:",
                {'inline_keyboard': [
                    [{'text': "🚚 Yetkazib beraman",   'callback_data': 'ob_delivery_deliver'}],
                    [{'text': "🏪 Xaridor olib ketadi", 'callback_data': 'ob_delivery_pickup'}],
                    [{'text': "🚚🏪 Ikkalasi ham",       'callback_data': 'ob_delivery_both'}],
                ]},
                state=s)

        elif step == 'ob_channel':
            channel = text if text.startswith('@') else f'@{text}'
            send_seller(cid, f"🔍 <b>{channel}</b> tekshirilmoqda...")
            # 1) Kanal mavjudligini tekshirish
            if not channel_exists(channel):
                send_seller(cid,
                    f"❌ <b>{channel}</b> kanali topilmadi yoki shaxsiy.\n\n"
                    "Kanal username to'g'rimi? <code>@kanalim</code> formatida qayta kiriting:")
                return
            # 2) Foydalanuvchi admin yoki egasi ekanligini tekshirish
            user_admin = can_manage_channel(uid, channel) or is_channel_admin(uid, channel)
            if not user_admin:
                send_seller(cid,
                    f"❌ Siz <b>{channel}</b> kanalining admini emassiz!\n\n"
                    "Avval o'zingizni kanalga admin qiling, keyin qayta kiriting:")
                return
            # 3) Bot admin ekanligini tekshirish
            if is_bot_admin_in(channel):
                finalize_shop_onboarding(uid, cid, s, channel)
            else:
                # Bot admin emas — /confirm flow ga o'tamiz
                s['ob_pending_channel'] = channel
                s['step'] = 'ob_confirm_admin'
                send_seller(cid,
                    f"⚠️ <b>Bot {channel} kanalida admin emas!</b>\n\n"
                    f"Quyidagini bajaring:\n"
                    f"1️⃣ {channel} → Settings → Administrators\n"
                    f"2️⃣ Add Admin → seller botni qidirib qo'shing\n"
                    f"3️⃣ Post Messages, Edit, Delete ruxsatlarini bering\n\n"
                    f"Tayyor bo'lganingizda <code>/confirm</code> yozing.\n"
                    f"Boshqa kanal kiritmoqchimisiz — /cancel")

        elif step == 'ob_confirm_admin':
            if text.strip() == '/cancel':
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Bekor qilindi. Yangi do'kon uchun /start"); return
            if text.strip() != '/confirm':
                send_seller(cid,
                    "Bot admin qilingach <code>/confirm</code> yozing yoki bekor qilish uchun /cancel"); return
            channel = s.get('ob_pending_channel')
            if not channel:
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Holat yo'qoldi. /start"); return
            send_seller(cid, f"🔍 <b>{channel}</b> qayta tekshirilmoqda...")
            if is_bot_admin_in(channel):
                finalize_shop_onboarding(uid, cid, s, channel)
            else:
                send_seller(cid,
                    f"❌ Bot hali ham {channel} kanalida admin emas.\n\n"
                    f"Tekshiring:\n"
                    f"• Bot username to'g'ri qo'shilganmi?\n"
                    f"• Post Messages ruxsati berilganmi?\n\n"
                    f"Tayyor bo'lgach yana <code>/confirm</code> yozing yoki /cancel")

        elif step == 'edit_shop_name':
            s['ob_shop_name'] = text; s['step'] = 'ob_phone'
            send_seller(cid, f"✅ Do'kon nomi yangilandi.\n\n📞 Telefon raqam:\n<i>+998XXXXXXXXX yoki /skip</i>",
                {'inline_keyboard': [[{'text': "⏭ O'zgartirmaslik", 'callback_data': 'ob_keep_phone'}]]})

        elif step == 'edit_phone_direct':
            s['ob_phone'] = text.strip()
            idx = s.get('edit_shop_idx', 0)
            shops = seller_shops.get(uid, [])
            if idx < len(shops):
                shops[idx]['phone'] = s['ob_phone']
                save_data()
            del seller_state[uid]
            send_seller(cid, f"✅ Telefon yangilandi: {s['ob_phone']}")

        elif step == 'edit_address_direct':
            s['ob_address'] = text.strip()
            idx = s.get('edit_shop_idx', 0)
            shops = seller_shops.get(uid, [])
            if idx < len(shops):
                shops[idx]['address'] = s['ob_address']
                save_data()
            del seller_state[uid]
            send_seller(cid, f"✅ Manzil yangilandi: {s['ob_address']}")

        elif step == 'edit_social_direct':
            social = {}
            for line in text.strip().splitlines():
                if ':' in line:
                    k, v = line.split(':', 1)
                    social[k.strip().lower()] = v.strip()
            idx = s.get('edit_shop_idx', 0)
            shops = seller_shops.get(uid, [])
            if idx < len(shops):
                shops[idx]['social'] = social
                save_data()
            del seller_state[uid]
            lines = '\n'.join(f"🔗 {k}: {v}" for k, v in social.items())
            send_seller(cid, f"✅ Ijtimoiy tarmoqlar yangilandi:\n{lines}")

        elif step == 'prod_name':
            s['name'] = text; s['step'] = 'prod_category'
            kb = []
            row = []
            for i, (cat, icon) in enumerate(CATEGORIES):
                row.append({'text': f"{icon} {cat}", 'callback_data': f"cat_{cat}"})
                if len(row) == 2:
                    kb.append(row); row = []
            if row: kb.append(row)
            send_seller(cid,
                f"✅ <b>{text}</b>\n\n"
                "<b>2/7</b> Kategoriyani tanlang:",
                {'inline_keyboard': kb}
            )

        elif step == 'prod_photo':
            photo = msg.get('photo')
            video = msg.get('video')
            media_group_id = msg.get('media_group_id')
            media_id = None
            if photo:
                media_id = photo[-1]['file_id']
            elif video:
                media_id = video['file_id']
            if media_id:
                if media_id not in s['photo_ids'] and len(s['photo_ids']) < 5:
                    s['photo_ids'].append(media_id)
                    upload_photo_async(media_id, SELLER_TOKEN, s)
                if media_group_id:
                    s['last_media_group'] = media_group_id
                    if s.get('album_timer'):
                        s['album_timer'].cancel()
                    def send_album_count(cid=cid, s=s):
                        count = len(s.get('photo_ids', []))
                        send_seller(cid,
                            f"✅ {count}/5 media qabul qilindi. Yana yuboring yoki davom eting:",
                            {'inline_keyboard': [[{'text': f"➡️ Davom etish ({count} ta)", 'callback_data': 'prod_photo_done'}]]}
                        )
                    import threading as _t
                    timer = _t.Timer(0.8, send_album_count)
                    s['album_timer'] = timer
                    timer.start()
                else:
                    count = len(s['photo_ids'])
                    send_seller(cid,
                        f"✅ {count}/5 media. Yana yuboring yoki davom eting:",
                        {'inline_keyboard': [[{'text': f"➡️ Davom etish ({count} ta)", 'callback_data': 'prod_photo_done'}]]}
                    )
            else:
                send_seller(cid, "❌ Rasm yoki video yuboring!")

        elif step == 'prod_price':
            sale_type = s.get('sale_type', 'both')
            parts = [p.strip() for p in text.replace(' ','').replace(',','').split('/')]
            orig = parse_price(parts[0]) if parts else None
            if orig is None:
                send_seller(cid, "❌ Format: <code>850000 / 550000</code>"); return
            if sale_type == 'solo':
                ok, err = validate_prices(orig, 0, orig, sale_type='solo')
                if not ok:
                    send_seller(cid, err); return
                s['original_price'] = orig
                s['solo_price']     = orig
                s['group_price']    = orig
                s['min_group']      = 1
                s['step'] = 'prod_desc'; s['description'] = ''; s['variants'] = []
                send_seller(cid,
                    f"✅ Narx: {orig:,} so'm\n\n"
                    "<b>5/6</b> Mahsulot tavsifi (ixtiyoriy):\n"
                    "<i>Mahsulot haqida qo'shimcha ma'lumot...</i>",
                    {'inline_keyboard': [[{'text': "⏭ O'tkazib yuborish", 'callback_data': 'prod_skip_desc'}]]}
                )
            else:
                grp = parse_price(parts[1]) if len(parts) > 1 else None
                if grp is None:
                    send_seller(cid, "❌ Format: <code>850000 / 550000</code>"); return
                ok, err = validate_prices(orig, grp, grp if sale_type == 'both' else 0, sale_type=sale_type)
                if not ok:
                    send_seller(cid, err); return
                disc = round((orig-grp)/orig*100)
                s['original_price'] = orig
                s['group_price']    = grp
                s['solo_price']     = grp if sale_type == 'both' else 0
                s['step'] = 'prod_min_group'
                send_seller(cid, f"✅ {orig:,} → {grp:,} so'm (-{disc}%)\n\n<b>5/5</b> Minimal guruh soni (2-100):")

        elif step == 'prod_min_group':
            ok, mg, err = validate_min_group_text(text)
            if not ok:
                send_seller(cid, err); return
            s['min_group'] = mg
            s['step'] = 'prod_desc'
            s['description'] = ''
            s['variants'] = []
            send_seller(cid,
                f"✅ Minimal guruh: {mg} kishi\n\n"
                "<b>6/6</b> Mahsulot tavsifi (ixtiyoriy):\n"
                "<i>Mahsulot haqida qo'shimcha ma'lumot...</i>",
                {'inline_keyboard': [[{'text': "⏭ O'tkazib yuborish", 'callback_data': 'prod_skip_desc'}]]}
            )

        elif step == 'prod_desc':
            s['description'] = text[:300]
            s['step'] = 'prod_mxik_search'
            send_seller(cid,
                "✅ Tavsif saqlandi.\n\n"
                "🔍 <b>MXIK kodi (ixtiyoriy)</b>\n\n"
                "Mahsulot nomini yoki kalit so'z kiriting yoki o'tkazib yuboring:\n"
                "<i>Masalan: krem, ko'ylak paxta, telefon</i>\n\n"
                "Bekor qilish: /cancel",
                {'inline_keyboard': [
                    [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                    [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
                ]})

        elif step == 'prod_mxik_search':
            keyword = text.strip()
            if len(keyword) < 2:
                send_seller(cid, "❌ Kamida 2 ta belgi kiriting"); return
            results, err = mxik_search(keyword)
            if err:
                send_seller(cid,
                    f"⚠️ {err}\n\n"
                    f"Qaytadan urining, kodni qo'lda kiriting yoki o'tkazib yuboring:",
                    {'inline_keyboard': [
                        [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                        [{'text': "🔄 Qayta qidirish",        'callback_data': 'prod_mxik_again'}],
                        [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
                    ]})
                return
            if not results:
                send_seller(cid,
                    f"🔍 \"{keyword}\" — natija yo'q.\n\nBoshqa kalit so'z bilan urinib ko'ring yoki o'tkazib yuboring.",
                    {'inline_keyboard': [
                        [{'text': "🔢 Kodni qo'lda kiritish", 'callback_data': 'prod_mxik_manual_btn'}],
                        [{'text': "🔄 Boshqa so'z",          'callback_data': 'prod_mxik_again'}],
                        [{'text': "⏭ O'tkazib yuborish",     'callback_data': 'prod_mxik_skip'}],
                    ]})
                return
            s['mxik_results'] = results
            s['mxik_keyword'] = keyword
            render_mxik_results(uid, cid, keyword, results, page=0)

        elif step == 'prod_mxik_manual':
            ok, code, err = mxik_validate_code(text)
            if not ok:
                send_seller(cid, err); return
            # Qo'lda kiritilgan kodni saqlaymiz, "manual" deb belgi
            s['mxik_code'] = code
            s['mxik_name'] = "(qo'lda kiritilgan)"
            s['step'] = 'prod_mxik_confirm_state'
            render_mxik_confirm(uid, cid, code, s['mxik_name'])

        elif step == 'prod_edit_desc':
            s['description'] = text[:300]; s['step'] = 'prod_confirm'
            shop = seller_shops.get(uid,[{}])[s.get('shop_idx',0)]
            show_prod_confirm(cid, s, shop)

        elif step == 'prod_edit_solo':
            solo = parse_price(text)
            if solo is None or solo <= 0:
                send_seller(cid, "❌ To'g'ri raqam kiriting (masalan: 850000)"); return
            orig = s.get('original_price', 0)
            grp  = s.get('group_price', 0)
            sale_type = s.get('sale_type', 'both')
            ok, err = validate_prices(orig, grp, solo, sale_type=sale_type)
            if not ok:
                send_seller(cid, err); return
            s['solo_price'] = solo; s['step'] = 'prod_confirm'
            shop = seller_shops.get(uid,[{}])[s.get('shop_idx',0)]
            show_prod_confirm(cid, s, shop)

        elif step == 'prod_edit_variants':
            raw = [v.strip() for v in text.replace('،',',').split(',') if v.strip()]
            if not raw: send_seller(cid, "❌ Kamida 1 ta variant!"); return
            s['variants'] = raw; s['step'] = 'prod_confirm'
            shop = seller_shops.get(uid,[{}])[s.get('shop_idx',0)]
            show_prod_confirm(cid, s, shop)

        elif step in ('leg_stir', 'leg_account', 'leg_bank_name', 'leg_mfo', 'leg_director'):
            # Yuridik ma'lumot kiritish — har step alohida validatsiya
            if step == 'leg_stir':
                ok, val, err = validate_stir(text)
                field = 'stir'
                next_step  = 'leg_account'
                next_label = "📋 <b>Qadam 3/6 — Hisob raqami</b>\n\nBank hisob raqamingizni kiriting (20 raqam):"
            elif step == 'leg_account':
                ok, val, err = validate_bank_account(text)
                field = 'bank_account'
                next_step  = 'leg_bank_name'
                next_label = "📋 <b>Qadam 4/6 — Bank nomi</b>\n\nQaysi bankda? (masalan: Kapitalbank, Davr Bank, Agrobank):"
            elif step == 'leg_bank_name':
                ok, val, err = validate_bank_name(text)
                field = 'bank_name'
                next_step  = 'leg_mfo'
                next_label = "📋 <b>Qadam 5/6 — MFO</b>\n\nBank MFO kodini kiriting (5 raqam):"
            elif step == 'leg_mfo':
                ok, val, err = validate_mfo(text)
                field = 'bank_mfo'
                # MFO dan keyin: MChJ bo'lsa direktor, aks holda confirm
                prof_status = seller_profiles.get(uid, {}).get('legal_status', '')
                if prof_status == 'mchj':
                    next_step  = 'leg_director'
                    next_label = "📋 <b>Qadam 6/6 — Direktor</b>\n\nDirektor F.I.O. ni to'liq kiriting (familiya, ism, sharif):"
                else:
                    next_step  = 'leg_confirm'
                    next_label = None  # confirm ekranini render_legal_confirm chiqaradi
            else:  # leg_director
                ok, val, err = validate_director_name(text)
                field = 'director_name'
                next_step  = 'leg_confirm'
                next_label = None
            if not ok:
                send_seller(cid, err)
                return
            prof = seller_profiles.setdefault(uid, {})
            prof[field] = val
            # Edit rejimida — to'g'ridan leg_confirm'ga qaytamiz
            if s.get('leg_editing'):
                s.pop('leg_editing', None)
                s['step'] = 'leg_confirm'
                render_legal_confirm(uid, cid)
                return
            s['step'] = next_step
            if next_step == 'leg_confirm':
                render_legal_confirm(uid, cid)
            else:
                send_seller(cid, next_label + "\n\n<i>Bekor qilish: /cancel</i>")
            return

        elif step == 'bz_set_disc':
            if text.strip() == '/cancel':
                idx = s.get('bz_disc_idx', 0)
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Bekor qilindi.",
                    {'inline_keyboard': [[{'text': "⬅️ Orqaga", 'callback_data': f'billz_disc_{idx}'}]]})
                return
            try:
                pct = int(text.strip().rstrip('%'))
            except (ValueError, TypeError):
                send_seller(cid, "❌ Butun son kiriting (0-90)"); return
            if pct < 0 or pct > 90:
                send_seller(cid, "❌ 0 dan 90 gacha bo'lishi kerak"); return
            idx  = s.get('bz_disc_idx', 0)
            kind = s.get('bz_disc_kind')
            shops = seller_shops.get(uid, [])
            if idx >= len(shops):
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Do'kon topilmadi"); return
            field = 'billz_global_solo_discount' if kind == 'solo' else 'billz_global_group_discount'
            shops[idx][field] = pct
            save_data()
            seller_state.pop(uid, None)
            label = "Solo" if kind == 'solo' else "Guruh"
            send_seller(cid,
                f"✅ {label} chegirma: <b>{pct}%</b>\n\n"
                f"Bu qiymat keyingi Billz mahsulotni yoqishda tavsiya narx hisoblashda ishlatiladi.",
                {'inline_keyboard': [[{'text': "⬅️ Sozlamalar", 'callback_data': f'billz_disc_{idx}'}]]})
            return

        elif step in ('bz_act_solo', 'bz_act_grp', 'bz_act_min'):
            if text.strip() == '/cancel':
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Yoqish bekor qilindi.",
                    {'inline_keyboard': [[{'text': "📦 Mahsulotlarim", 'callback_data': 'menu_myproducts'}]]})
                return
            pid = s.get('bz_pid')
            p = products.get(pid)
            if not p or p.get('seller_id') != uid:
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Mahsulot topilmadi"); return
            orig = int(p.get('original_price', 0) or 0)

            if step == 'bz_act_solo':
                solo = parse_price(text)
                if solo is None or solo <= 0:
                    send_seller(cid, "❌ To'g'ri raqam kiriting"); return
                if solo >= orig:
                    send_seller(cid, "❌ Yakka narx asl narxdan past bo'lishi kerak. Qayta kiriting."); return
                s['bz_solo'] = solo
                s['step'] = 'bz_act_grp'
                send_seller(cid,
                    f"✅ Yakka: {fmt(solo)} so'm\n\n"
                    f"<b>2/4</b> Guruh narxini yozing (so'm).\n"
                    f"💡 Tavsiya: <b>{fmt(s.get('bz_suggested_group', 0))}</b> so'm")
                return

            if step == 'bz_act_grp':
                grp = parse_price(text)
                if grp is None or grp <= 0:
                    send_seller(cid, "❌ To'g'ri raqam kiriting"); return
                if grp >= s.get('bz_solo', 0):
                    send_seller(cid, "❌ Guruh narxi yakka narxdan past bo'lishi kerak. Qayta kiriting."); return
                s['bz_group'] = grp
                s['step'] = 'bz_act_min'
                send_seller(cid,
                    f"✅ Guruh: {fmt(grp)} so'm\n\n"
                    f"<b>3/4</b> Minimal guruh sonini yozing (2-100):")
                return

            if step == 'bz_act_min':
                ok, mg, err = validate_min_group_text(text)
                if not ok:
                    send_seller(cid, err); return
                s['bz_min'] = mg
                s['step'] = 'bz_act_deadline'
                send_seller(cid,
                    f"✅ Min guruh: {mg} kishi\n\n"
                    f"<b>4/4</b> Muddatni tanlang:",
                    {'inline_keyboard': [
                        [{'text': "24 soat",  'callback_data': 'bz_deadline_24'},
                         {'text': "2 kun",   'callback_data': 'bz_deadline_48'}],
                        [{'text': "3 kun",   'callback_data': 'bz_deadline_72'},
                         {'text': "1 hafta",'callback_data': 'bz_deadline_168'}],
                    ]})
                return

        elif step == 'billz_secret_token':
            if text.strip() == '/cancel':
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Bekor qilindi.",
                    {'inline_keyboard': [[{'text': "🔌 Billz menyu", 'callback_data': 'billz_menu'}]]})
                return
            secret = text.strip()
            if len(secret) < 10:
                send_seller(cid, "❌ Token juda qisqa. Qayta yuboring yoki /cancel"); return
            shop_idx = s.get('billz_shop_idx', 0)
            send_seller(cid, "🔍 Billz bilan ulanish tekshirilmoqda...")
            access_token, err = billz_login(secret)
            if not access_token:
                send_seller(cid,
                    f"❌ Ulanish muvaffaqiyatsiz: {err}\n\n"
                    f"Tokenni qayta tekshirib yuboring yoki /cancel")
                return
            # Mahsulot olib ko'rib do'konlarni aniqlash
            try:
                r = requests.get(
                    f'{BILLZ_BASE_URL}/v2/products',
                    headers={'Authorization': f'Bearer {access_token}'},
                    params={'limit': 1, 'page': 1}, timeout=15
                )
                if r.status_code != 200:
                    send_seller(cid,
                        f"❌ Billz dan mahsulot olishda xato: HTTP {r.status_code}\n\n"
                        f"Token to'g'ri lekin ruxsat cheklanganmi? Qayta urining yoki /cancel")
                    return
                resp = r.json() or {}
            except Exception as e:
                send_seller(cid, f"❌ Tarmoq xatosi: {e}\n\nQayta urining yoki /cancel")
                return
            candidates = billz_extract_shops(resp)
            if not candidates:
                send_seller(cid,
                    "⚠️ Billz hisobingizda mahsulot topilmadi yoki shop_measurement_values bo'sh.\n\n"
                    "Avval Billz UI'da kamida bitta mahsulot va do'kon yarating, keyin qayta urining.")
                seller_state.pop(uid, None)
                return
            if len(candidates) == 1:
                # Avtomatik tanlash
                only = candidates[0]
                encrypted = encrypt_token(secret)
                if not encrypted:
                    send_seller(cid, "❌ Shifrlash xatosi"); seller_state.pop(uid, None); return
                shops = seller_shops.get(uid, [])
                if shop_idx >= len(shops):
                    send_seller(cid, "❌ Do'kon topilmadi"); seller_state.pop(uid, None); return
                shops[shop_idx]['billz_secret_token']  = encrypted
                shops[shop_idx]['billz_shop_id']       = only['shop_id']
                shops[shop_idx]['billz_shop_name']     = only['shop_name']
                shops[shop_idx]['billz_connected_at']  = datetime.now().strftime('%Y-%m-%d %H:%M')
                save_data()
                seller_state.pop(uid, None)
                send_seller(cid,
                    f"✅ <b>Billz ulandi!</b>\n\n"
                    f"🏬 Billz do'koni: <b>{only['shop_name']}</b>\n\n"
                    f"Mahsulotlarni import qilish — Faza 2 (keyingi deploy).",
                    {'inline_keyboard': [[{'text': "🔌 Billz menyu", 'callback_data': 'billz_menu'}]]})
                return
            # Bir nechta do'kon — tanlash so'raymiz
            s['step']               = 'billz_shop_select'
            s['billz_pending_token']= secret
            s['billz_candidates']   = candidates
            kb = [[{'text': c['shop_name'][:40],
                    'callback_data': f"billz_pickshop_{shop_idx}_{c['shop_id']}"}]
                  for c in candidates[:20]]
            kb.append([{'text': "❌ Bekor", 'callback_data': 'billz_menu'}])
            send_seller(cid,
                f"✅ Token to'g'ri.\n\n"
                f"Billz hisobingizda <b>{len(candidates)}</b> ta do'kon topildi. "
                f"Joynshop'dagi <b>{seller_shops.get(uid,[{}])[shop_idx].get('name','')}</b> bilan qaysisini bog'lashni xohlaysiz?",
                {'inline_keyboard': kb})
            return

        elif step and step.startswith('pp_edit_'):
            if text.strip() == '/cancel':
                pid_back = s.get('pp_pid')
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Bekor qilindi.",
                    {'inline_keyboard': [[{'text': "⬅️ Mahsulotga qaytish",
                                           'callback_data': f'mp_view_{pid_back}'}]]} if pid_back else None)
                return
            pid = s.get('pp_pid')
            p   = products.get(pid) if pid else None
            if not p or p.get('seller_id') != uid:
                seller_state.pop(uid, None)
                send_seller(cid, "❌ Mahsulot topilmadi"); return

            # ─── Photo edit (faqat photo qabul qiladi) ───
            if step == 'pp_edit_photo':
                photo = msg.get('photo')
                if not photo:
                    send_seller(cid, "❌ Rasm yuboring (faylni emas)"); return
                file_id = photo[-1]['file_id']
                p['photo_id']   = file_id
                p['photo_ids']  = [file_id]
                p['photo_url']  = ''
                p['photo_urls'] = []
                # S3 yuklash (async, eski kabi)
                upload_photo_async(file_id, SELLER_TOKEN, p)
                save_data()
                seller_state.pop(uid, None)
                send_seller(cid,
                    f"✅ Rasm yangilandi.\n\n"
                    f"⚠️ Telegram media group'larini tahrirlab bo'lmaydi — "
                    f"kanal post'ini yangilash uchun /boost qiling yoki tugma orqali qayta e'lon qiling.",
                    {'inline_keyboard': [
                        [{'text': "📢 Qayta e'lon", 'callback_data': f'boost_{pid}'}],
                        [{'text': "⬅️ Orqaga",     'callback_data': f'mp_view_{pid}'}],
                    ]})
                return

            # ─── Text-based field edits ───
            if step == 'pp_edit_name':
                new_name = text.strip()
                if not new_name:
                    send_seller(cid, "❌ Nom bo'sh bo'lmasligi kerak"); return
                p['name'] = new_name[:100]
            elif step == 'pp_edit_orig':
                val = parse_price(text)
                if val is None or val <= 0:
                    send_seller(cid, "❌ To'g'ri raqam kiriting"); return
                ok, err = validate_prices(val, p.get('group_price', 0), p.get('solo_price', 0),
                                          sale_type=p.get('sale_type', 'both'))
                if not ok:
                    send_seller(cid, err); return
                p['original_price'] = val
            elif step == 'pp_edit_grp':
                val = parse_price(text)
                if val is None or val <= 0:
                    send_seller(cid, "❌ To'g'ri raqam kiriting"); return
                ok, err = validate_prices(p.get('original_price', 0), val, p.get('solo_price', 0),
                                          sale_type=p.get('sale_type', 'both'))
                if not ok:
                    send_seller(cid, err); return
                p['group_price'] = val
            elif step == 'pp_edit_solo':
                val = parse_price(text)
                if val is None or val <= 0:
                    send_seller(cid, "❌ To'g'ri raqam kiriting"); return
                ok, err = validate_prices(p.get('original_price', 0), p.get('group_price', 0), val,
                                          sale_type=p.get('sale_type', 'both'))
                if not ok:
                    send_seller(cid, err); return
                p['solo_price'] = val
            elif step == 'pp_edit_min':
                ok, mg, err = validate_min_group_text(text)
                if not ok:
                    send_seller(cid, err); return
                p['min_group'] = mg
            elif step == 'pp_edit_deadline':
                try:
                    hours = int(text.strip())
                except (ValueError, TypeError):
                    send_seller(cid, "❌ Soatlar sonini kiriting (masalan: 48)"); return
                if hours < 1 or hours > 720:
                    send_seller(cid, "❌ 1 dan 720 soatgacha kiriting"); return
                deadline_dt = datetime.now() + timedelta(hours=hours)
                p['deadline']    = deadline_dt.strftime('%d.%m.%Y %H:%M')
                p['deadline_dt'] = deadline_dt.strftime('%Y-%m-%d %H:%M')
            elif step == 'pp_edit_desc':
                p['description'] = text[:300]
            elif step == 'pp_edit_variants':
                raw = [v.strip() for v in text.replace('،', ',').split(',') if v.strip()]
                if not raw:
                    send_seller(cid, "❌ Kamida 1 ta variant kiriting"); return
                p['variants'] = raw
            else:
                send_seller(cid, "❌ Noma'lum step"); seller_state.pop(uid, None); return

            # Saqlash + kanal post caption'ini yangilash
            save_data()
            try:
                ch_cid = p.get('channel_chat_id')
                ch_mid = p.get('channel_message_id')
                if ch_cid and ch_mid:
                    count = len(groups.get(pid, []))
                    edit_caption(ch_cid, ch_mid, post_caption(p, pid),
                                 join_kb(pid, count, p.get('min_group', 0),
                                         has_solo=bool(p.get('solo_price')),
                                         sale_type=p.get('sale_type', 'both')))
            except Exception as e:
                logging.error(f"pp_edit channel update error: {e}")
            seller_state.pop(uid, None)
            send_seller(cid,
                f"✅ Yangilandi.",
                {'inline_keyboard': [
                    [{'text': "✏️ Yana tahrirlash", 'callback_data': f'mp_edit_{pid}'}],
                    [{'text': "⬅️ Mahsulotga",     'callback_data': f'mp_view_{pid}'}],
                ]})
            return

        elif step == 'name':
            s['name'] = text; s['step'] = 'shop_name'
            send_seller(cid, "2️⃣ Do'kon nomingiz:\n<i>Masalan: Nike Toshkent</i>")

        elif step == 'shop_name':
            s['shop_name'] = text; s['step'] = 'description'
            send_seller(cid, "3️⃣ Mahsulot tavsifi:")

        elif step == 'description':
            s['description'] = text; s['step'] = 'original_price'
            send_seller(cid, "4️⃣ Asl narx (so'm):\n<i>Masalan: 850000</i>")

        elif step == 'original_price':
            try:
                s['original_price'] = int(text.replace(' ','').replace(',',''))
                s['step'] = 'group_price'
                send_seller(cid, "5️⃣ Guruh narxi (so'm):\n<i>Masalan: 550000</i>")
            except:
                send_seller(cid, "❌ Faqat raqam kiriting!")

        elif step == 'group_price':
            try:
                s['group_price'] = int(text.replace(' ','').replace(',',''))
                s['step'] = 'solo_price'
                send_seller(cid,
                    "5️⃣b Yakka sotish narxi (so'm):\n"
                    "<i>Masalan: 720000</i>\n\n"
                    "Yakka sotish bo'lmasa /skip yozing"
                )
            except:
                send_seller(cid, "❌ Faqat raqam kiriting!")

        elif step == 'solo_price':
            if text.strip() == '/skip':
                s['solo_price'] = 0
            else:
                try:
                    s['solo_price'] = int(text.replace(' ','').replace(',',''))
                except:
                    send_seller(cid, "❌ Faqat raqam kiriting yoki /skip yozing!"); return
            s['step'] = 'has_variants'
            send_seller(cid,
                "5️⃣c Variantlar bormi? (o'lcham, rang va h.k.)",
                {'inline_keyboard': [
                    [{'text': "✅ Ha, variantlar bor", 'callback_data': 'variants_yes'}],
                    [{'text': "❌ Yo'q, oddiy mahsulot", 'callback_data': 'variants_no'}],
                ]}
            )

        elif step == 'variants_input':
            raw = [v.strip() for v in text.replace('،',',').split(',') if v.strip()]
            if not raw:
                send_seller(cid, "❌ Kamida 1 ta variant kiriting!"); return
            s['variants'] = raw
            s['step'] = 'min_group'
            send_seller(cid,
                f"✅ Variantlar: {', '.join(raw)}\n\n"
                f"6️⃣ Minimal guruh soni (2-100):"
            )

        elif step == 'min_group':
            ok, mg, err = validate_min_group_text(text)
            if not ok:
                send_seller(cid, err); return
            s['min_group'] = mg; s['step'] = 'photo'
            send_seller(cid, "7️⃣ Mahsulot rasmini yuboring 📸")

        elif step == 'photo':
            photo = msg.get('photo')
            if photo:
                s['photo_id'] = photo[-1]['file_id']
                s['step']     = 'contact'
                send_seller(cid, "8️⃣ Aloqa ma'lumotingiz:\n<i>@username yoki +998XXXXXXXXX</i>")
            else:
                send_seller(cid, "❌ Rasm yuboring!")

        elif step == 'contact':
            s['contact'] = text
            s['step']    = 'delivery_type'
            send_seller(cid,
                "8️⃣b Yetkazib berish turi:",
                {'inline_keyboard': [
                    [{'text': "🚚 Sotuvchi yetkazadi", 'callback_data': 'delivery_deliver'}],
                    [{'text': "🏪 Xaridor olib ketadi", 'callback_data': 'delivery_pickup'}],
                ]}
            )

        elif step == 'seller_channel':
            channel = text if text.startswith('@') else f'@{text}'
            s['seller_channel'] = channel
            if can_manage_channel(uid, channel):
                s['step'] = 'confirm'
                show_confirm(cid, s)
            else:
                send_seller(cid, f"🔍 <b>{channel}</b> adminligi tekshirilmoqda...")
                if is_channel_admin(uid, channel):
                    verified_channels[channel] = {'owner_id': uid, 'moderators': []}
                    save_data()
                    send_seller(cid, f"✅ <b>{channel}</b> tasdiqlandi!")
                    s['step'] = 'confirm'
                    show_confirm(cid, s)
                else:
                    send_seller(cid,
                        f"❌ <b>{channel}</b> kanalining admini emassiz!\n\n"
                        f"Tekshiring:\n"
                        f"• Seller bot kanalga admin sifatida qo'shilganmi?\n"
                        f"• Kanal username to'g'rimi?\n\n"
                        f"Qayta urinish: /addproduct"
                    )

        elif step == 'editing':
            field = s.get('edit_field')
            if field in ('original_price', 'group_price'):
                val = parse_price(text)
                if val is None or val <= 0:
                    send_seller(cid, "❌ To'g'ri raqam kiriting (masalan: 850000)"); return
                # Tahrirlangach narxlar mantig'ini tekshiramiz
                trial = dict(s); trial[field] = val
                ok, err = validate_prices(
                    trial.get('original_price', 0),
                    trial.get('group_price', 0),
                    trial.get('solo_price', 0),
                    sale_type=trial.get('sale_type', 'both'),
                )
                if not ok:
                    send_seller(cid, err); return
                s[field] = val
            elif field == 'min_group':
                ok, mg, err = validate_min_group_text(text)
                if not ok:
                    send_seller(cid, err); return
                s[field] = mg
            elif field == 'photo':
                photo = msg.get('photo')
                if photo:
                    s['photo_id'] = photo[-1]['file_id']
                else:
                    send_seller(cid, "❌ Rasm yuboring!"); return
            elif field == 'seller_channel':
                s[field] = text if text.startswith('@') else f'@{text}'
            else:
                s[field] = text
            s['step'] = 'confirm'
            send_seller(cid, "✅ Yangilandi!")
            show_confirm(cid, s)

# ══════════════════════════════════════════════════════════════════════
#  BUYER WEBHOOK
# ══════════════════════════════════════════════════════════════════════
@app.route('/buyer/webhook', methods=['POST'])
def buyer_webhook():
    data = request.json
    if 'callback_query' in data:
        buyer_handle_cb(data['callback_query'])
    elif 'pre_checkout_query' in data:
        handle_pre_checkout(data['pre_checkout_query'])
    elif 'message' in data:
        msg = data['message']
        if 'successful_payment' in msg:
            handle_successful_payment(msg)
        elif msg.get('chat', {}).get('type', '') in ['group', 'supergroup']:
            moderate_chat(msg)
        else:
            buyer_handle_msg(msg)
    return 'ok'

def handle_pre_checkout(query):
    """Telegram to'lovni tasdiqlashdan oldin chaqiradi — biz OK deymiz"""
    payload = query.get('invoice_payload', '')
    uid     = query['from']['id']
    # Channel invoice: channel_PID formatida
    if payload.startswith('channel_'):
        pid = payload[8:]
        if pid in products and products[pid].get('status') == 'active':
            answer_pre_checkout(query['id'], ok=True)
        else:
            answer_pre_checkout(query['id'], ok=False, error="Mahsulot topilmadi yoki yopilgan")
    elif payload in orders and orders[payload]['status'] == 'pending':
        answer_pre_checkout(query['id'], ok=True)
    else:
        answer_pre_checkout(query['id'], ok=False, error="Buyurtma topilmadi yoki muddati o'tgan")

def handle_successful_payment(msg):
    """To'lov muvaffaqiyatli — buyurtmani tasdiqlash"""
    uid     = msg['from']['id']
    uname   = msg['from'].get('first_name', 'Foydalanuvchi')
    payment = msg['successful_payment']
    payload = payment['invoice_payload']
    amount  = payment['total_amount'] // 100

    # Kanal invoice: channel_PID
    if payload.startswith('channel_'):
        pid = payload[8:]
        p   = products.get(pid, {})
        if not p:
            send_buyer(uid, "❌ Mahsulot topilmadi.")
            return
        # Yangi buyurtma yaratish
        code = gen_code()
        orders[code] = {
            'product_id': pid, 'user_id': uid,
            'user_name':  uname, 'amount': amount,
            'type': 'group', 'status': 'confirming',
            'variant': '', 'payment_method': 'click',
            'telegram_payment_charge_id': payment.get('telegram_payment_charge_id',''),
            'created': datetime.now().strftime('%d.%m.%Y %H:%M')
        }
        # Guruhga qo'shish
        if pid not in groups: groups[pid] = []
        if uid not in groups[pid]: groups[pid].append(uid)
        sid = p.get('seller_id')
        update_customer(sid, uid, uname, amount, p.get('name',''),
                        username=msg['from'].get('username',''))
        save_data()
        # Sotuvchiga xabar
        if sid:
            send_seller(sid,
                f"💳 <b>KANAL TO'LOV!</b>\n\n"
                f"📦 {p.get('name','')}\n"
                f"👤 {uname} (ID: {uid})\n"
                f"💰 {fmt(amount)} so'm\n"
                f"👥 Guruh: {len(groups[pid])}/{p.get('min_group',3)}\n"
                f"💳 Click ✅ • 🆔 #{code}",
                {'inline_keyboard': [[
                    {'text': '✅ Tasdiqlash', 'callback_data': f'seller_ac_{code}'},
                    {'text': '❌ Rad',        'callback_data': f'seller_rj_{code}'},
                ]]}
            )
        send_buyer(uid,
            f"✅ <b>To'lov qabul qilindi!</b>\n\n"
            f"📦 {p.get('name','')}\n"
            f"💰 {fmt(amount)} so'm\n"
            f"👥 Guruh: {len(groups[pid])}/{p.get('min_group',3)}\n"
            f"🆔 #{code}\n\n"
            f"⏳ Sotuvchi tasdiqlashi kutilmoqda..."
        )
        return

    # Bot orqali buyurtma
    code = payload
    if code not in orders:
        send_buyer(uid, "❌ Buyurtma topilmadi.")
        return
    orders[code]['status']       = 'confirming'
    orders[code]['payment_method'] = 'click'
    orders[code]['telegram_payment_charge_id'] = payment.get('telegram_payment_charge_id','')
    save_data()
    p   = products.get(orders[code]['product_id'], {})
    sid = p.get('seller_id')
    if sid:
        o = orders[code]
        variant_text = f"\n🎨 Variant: {o.get('variant','')}" if o.get('variant') else ''
        send_seller(sid,
            f"💳 <b>TO'LOV TASDIQLANDI!</b>\n\n"
            f"📦 {p.get('name','')}{variant_text}\n"
            f"👤 {o['user_name']} (ID: {uid})\n"
            f"💰 {fmt(amount)} so'm\n"
            f"💳 Click ✅ • 🆔 #{code}",
            {'inline_keyboard': [[
                {'text': '✅ Tasdiqlash', 'callback_data': f'seller_ac_{code}'},
                {'text': '❌ Rad',        'callback_data': f'seller_rj_{code}'},
            ]]}
        )
    auto_check(code, orders[code], p)
    send_buyer(uid,
        f"✅ <b>To'lov qabul qilindi!</b>\n\n"
        f"📦 {p.get('name','')}\n"
        f"💰 {fmt(amount)} so'm\n"
        f"🆔 #{code}\n\n"
        f"⏳ Sotuvchi tasdiqlashi kutilmoqda..."
    )

def delivery_notice(p):
    dtype = p.get('delivery_type', 'pickup')
    if dtype == 'deliver':
        return "🚚 <b>Yetkazib berish:</b> Sotuvchi yetkazadi\nTo'lovdan so'ng manzil so'raladi."
    return "🏪 <b>Olish:</b> Sotuvchi manzilidan olib ketasiz\n📞 " + p.get('contact', '')

def auto_check(code, order, p):
    dtype = p.get('delivery_type', 'pickup')
    delivery_text = "🚚 Sotuvchi yetkazadi" if dtype == 'deliver' else "🏪 Xaridor olib ketadi"
    send_buyer(order['user_id'],
        f"📋 <b>BUYURTMA QABUL QILINDI</b>\n"
        f"━━━━━━━━━━━━━━━\n"
        f"🏪 {p.get('shop_name','')}\n"
        f"📦 {p.get('name','')}\n"
        f"🛒 {'Yakka' if order.get('type')=='solo' else 'Guruh'} buyurtma\n"
        f"💰 {fmt(order['amount'])} so'm\n"
        f"📅 {order.get('created','')}\n"
        f"🆔 #{code}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"{delivery_text}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"⏳ Sotuvchi 15 daqiqa ichida tasdiqlaydi\n"
        f"🔒 Joynshop kafolati ostida"
    )

def buyer_handle_cb(cb):
    cbid = cb['id']
    uid  = cb['from']['id']
    d    = cb['data']

    if d == 'noop':
        answer_cb(cbid); return

    # ── OPEN SHOP ──
    if d == 'settings_toggle_notif':
        prof = get_profile(uid)
        prof['notifications'] = not prof.get('notifications', True)
        save_data()
        notif_on = prof['notifications']
        answer_cb(cbid, ("🔔 Yoqildi" if notif_on else "🔕 O'chirildi"))
        send_buyer(uid,
            "⚙️ <b>Sozlamalar</b>\n\n"
            f"🔔 Bildirishnomalar: {'Yoqilgan ✅' if notif_on else 'O‘chirilgan ❌'}",
            {'inline_keyboard': [[
                {'text': ("🔕 O‘chirish" if notif_on else "🔔 Yoqish"),
                 'callback_data': 'settings_toggle_notif'},
            ]]}
        )
        return

    if d == 'open_shop':
        answer_cb(cbid)
        if APP_URL:
            send_buyer(uid,
                "🛍 <b>Joynshop do'koni</b>\n\nMahsulotlarni ko'rish uchun:",
                {'inline_keyboard': [
                    [{'text': "🌐 Saytga o'tish",  'url': APP_URL}],
                    [{'text': "📱 Miniapp ochish", 'web_app': {'url': f"{(BACKEND_URL or APP_URL or '').rstrip('/')}/miniapp"}}],
                ]}
            )
        else:
            send_buyer(uid,
                "🛍 Tez kunda to'liq sayt ishga tushadi!\n"
                "Hozircha botdan foydalaning 🤝"
            )
        return

    # ── BUYER MENU ──
    def buyer_main_menu():
        send_buyer(uid,
            "👋 <b>Joynshop</b>\n\nNimani qilmoqchisiz?",
            {'inline_keyboard': [
                [{'text': "🛍 Do'konga o'tish",  'callback_data': 'open_shop'}],
                [{'text': "📋 Buyurtmalarim",    'callback_data': 'buyer_mystatus'}],
                [
                    {'text': "👤 Profilim",      'callback_data': 'buyer_myprofile'},
                    {'text': "🤍 Wishlist",      'callback_data': 'buyer_mywishlist'},
                ],
                [
                    {'text': "↩️ Qaytarish",    'callback_data': 'buyer_refund'},
                    {'text': "❓ Yordam",         'callback_data': 'buyer_help'},
                ],
            ]}
        )

    if d == 'buyer_mystatus':
        answer_cb(cbid)
        my = {k:v for k,v in orders.items() if v['user_id']==uid}
        if not my:
            send_buyer(uid, "📋 Buyurtma yo'q.",
                {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}]]}
            ); return
        r  = "📋 <b>Buyurtmalaringiz:</b>\n\n"
        em = {'pending':'⏳','confirming':'🔄','confirmed':'✅','rejected':'❌','cancelled':'🚫'}
        st = {'pending':"To'lov kutilmoqda",'confirming':'Tekshirilmoqda',
              'confirmed':'Tasdiqlandi','rejected':'Rad','cancelled':'Bekor'}
        for k, o in list(my.items())[-5:]:
            p  = products.get(o['product_id'],{})
            r += f"{em.get(o['status'],'?')} <b>#{k}</b>\n{p.get('name','?')} — {fmt(o['amount'])} so'm\n{st.get(o['status'],'')}\n\n"
        send_buyer(uid, r, {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}]]})
        return

    if d == 'buyer_myprofile':
        answer_cb(cbid)
        p      = get_profile(uid)
        ref_d  = referrals.get(str(uid), {'count': 0, 'cashback': 0})
        ref_link = f"https://t.me/{BUYER_BOT_USERNAME}?start=ref_{uid}"
        send_buyer(uid,
            f"👤 <b>Profilingiz</b>\n\n"
            f"🛒 Jami xaridlar: {p['total_orders']}\n"
            f"💰 Tejagan: {fmt(p['total_saved'])} so'm\n"
            f"👥 Guruhlar: {p['groups_joined']}\n"
            f"🎁 Cashback: {fmt(p['cashback'])} so'm\n"
            f"👫 Taklif qilganlar: {ref_d['count']} kishi\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"🔗 Referral linkingiz:\n"
            f"<code>{ref_link}</code>\n\n"
            f"Har taklif uchun +10,000 so'm cashback!",
            {'inline_keyboard': [
                [{'text': "🔗 Referral linkni ulashish", 'url': f"https://t.me/share/url?url={ref_link}&text=🛍%20Do'stlarim%20bilan%20birgalikda%20xarid%20qilib%2040%25%20gacha%20tejayapman!%20Sen%20ham%20ulab%20ko'r%20👇"}],
                [{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}]
            ]}
        )
        return

    if d == 'buyer_mywishlist':
        answer_cb(cbid)
        wl = wishlists.get(uid, [])
        if not wl:
            send_buyer(uid, "🤍 Wishlist bo'sh.",
                {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}]]}
            ); return
        r = "🤍 <b>Wishlistingiz:</b>\n\n"
        for pid in wl:
            p = products.get(pid, {})
            if not p: continue
            count = len(groups.get(pid, []))
            r    += f"📦 {p.get('name','')}\n💰 {fmt(p.get('group_price',0))} so'm\n👥 {count}/{p.get('min_group',3)}\n\n"
        send_buyer(uid, r, {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}]]})
        return

    if d == 'buyer_refund':
        answer_cb(cbid)
        my = {k:v for k,v in orders.items() if v['user_id']==uid and v['status']=='confirmed'}
        if not my:
            send_buyer(uid, "Qaytarish uchun tasdiqlangan buyurtma yo'q.",
                {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}]]}
            ); return
        btns = []
        for k, o in list(my.items())[-5:]:
            p = products.get(o['product_id'],{})
            btns.append([{'text': f"#{k} — {p.get('name','')}", 'callback_data': f'refund_{k}'}])
        btns.append([{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}])
        send_buyer(uid, "Qaysi buyurtmani qaytarmoqchisiz?", {'inline_keyboard': btns})
        return

    if d == 'buyer_help':
        answer_cb(cbid)
        send_buyer(uid,
            "ℹ️ <b>Yordam</b>\n\n"
            "/mystatus   — Buyurtmalarim\n"
            "/myprofile  — Profilim\n"
            "/mywishlist — Saqlangan mahsulotlar\n"
            "/refund     — Qaytarish so'rovi\n\n"
            "🆘 Yordam: @joynshop_support",
            {'inline_keyboard': [[{'text': "🔙 Menyu", 'callback_data': 'buyer_back'}]]}
        )
        return

    if d == 'buyer_back':
        answer_cb(cbid)
        buyer_main_menu()
        return

    if d.startswith('choose_'):
        pid = d[7:]
        if pid not in products:
            answer_cb(cbid, '❌ Mahsulot topilmadi!'); return
        p = products[pid]
        if p.get('status') == 'closed':
            answer_cb(cbid, '⛔️ Yopilgan!'); return
        answer_cb(cbid)
        count    = len(groups.get(pid, []))
        min_g    = p['min_group']
        has_solo = p.get('solo_price')
        kb = []
        if count < min_g:
            kb.append([{'text': f"👥 Guruh narxi — {fmt(p['group_price'])} so'm ({count}/{min_g})", 'callback_data': f'join_{pid}'}])
        if has_solo:
            kb.append([{'text': f"👤 Yakka narxi — {fmt(p['solo_price'])} so'm", 'callback_data': f'solo_{pid}'}])
        kb.append([{'text': "❌ Bekor", 'callback_data': 'noop'}])
        send_buyer(uid,
            f"📦 <b>{p['name']}</b>\n\nQanday sotib olmoqchisiz?",
            {'inline_keyboard': kb}
        )
        return

    if d.startswith('join_'):
        pid = d[5:]
        if pid not in products:
            answer_cb(cbid, '❌ Mahsulot topilmadi!'); return
        p = products[pid]
        if p.get('status') == 'closed':
            answer_cb(cbid, '⛔️ Guruh yopilgan!'); return
        if pid not in groups: groups[pid] = []
        if uid in groups[pid]:
            answer_cb(cbid, '✅ Allaqachon guruhdasiz!'); return

        variants = p.get('variants', [])
        if variants:
            answer_cb(cbid)
            btns = [[{'text': v, 'callback_data': f'variant_{pid}_{v}'}] for v in variants]
            send_buyer(uid,
                f"📦 <b>{p['name']}</b>\n\nVariantni tanlang:",
                {'inline_keyboard': btns}
            )
            return

        code = gen_code()
        orders[code] = {
            'product_id': pid, 'user_id': uid,
            'user_name':  cb['from'].get('first_name', 'Foydalanuvchi'),
            'amount':     p['group_price'], 'type': 'group',
            'status':     'pending', 'variant': '',
            'created':    datetime.now().strftime('%d.%m.%Y %H:%M')
        }
        save_data()
        answer_cb(cbid, "To'lov sahifasi ochilmoqda...")
        # Click invoice yuborish
        if CLICK_TOKEN:
            photo = p.get('photos', [None])[0]
            photo_url = None
            if photo and CDN_BASE_URL:
                photo_url = f"{CDN_BASE_URL}/products/{photo}.jpg"
            send_invoice(
                uid,
                title=p['name'],
                description=f"Guruh xarid • {p.get('shop_name','Sotuvchi')} • #{code}",
                payload=code,
                amount=p['group_price'],
                photo_url=photo_url
            )
        else:
            # Fallback: qo'lda Payme
            send_buyer(uid,
                f"🛒 <b>{p.get('shop_name','Sotuvchi')} — Guruh buyurtma</b>\n"
                f"━━━━━━━━━━━━━━━\n"
                f"📦 {p['name']}\n💰 {fmt(p['group_price'])} so'm\n\n"
                f"💳 <b>Payme:</b>\n"
                f"📱 <code>{PAYME_NUMBER}</code>\n"
                f"💵 <code>{fmt(p['group_price'])}</code>\n"
                f"📝 Izoh: <code>{code}</code>\n\n"
                f"⚠️ Izohga <b>{code}</b> yozing!\n"
                f"━━━━━━━━━━━━━━━\n🔒 Joynshop kafolati ostida",
                {'inline_keyboard': [
                    [{'text': "✅ To'lovni tasdiqlayman", 'callback_data': f'paid_{code}'}],
                    [{'text': "❌ Bekor",                 'callback_data': f'cancel_{code}'}]
                ]}
            )
        return

    if d.startswith('paid_'):
        code = d[5:]
        if code not in orders:
            answer_cb(cbid, '❌ Buyurtma topilmadi!'); return
        o = orders[code]
        if o['status'] != 'pending':
            answer_cb(cbid, '⚠️ Allaqachon yuborilgan!'); return
        orders[code]['status'] = 'confirming'
        save_data()
        answer_cb(cbid, '⏳ Sotuvchi tasdiqlamoqda...')
        p_auto = products.get(orders[code]['product_id'], {})
        auto_check(code, orders[code], p_auto)
        p   = products.get(o['product_id'], {})
        sid = p.get('seller_id')
        if sid:
            variant_text = f"\n🎨 Variant: {o.get('variant','')}" if o.get('variant') else ''
            send_seller(sid,
                f"🔔 <b>YANGI TO'LOV!</b>\n\n"
                f"📦 {p.get('name','')}{variant_text}\n👤 {o['user_name']} (ID: {uid})\n"
                f"💰 {fmt(o['amount'])} so'm\n"
                f"🛒 {'Yakka' if o.get('type')=='solo' else 'Guruh'}\n🆔 #{code}",
                {'inline_keyboard': [[
                    {'text': '✅ Tasdiqlash', 'callback_data': f'seller_ac_{code}'},
                    {'text': '❌ Rad',        'callback_data': f'seller_ar_{code}'}
                ]]}
            )
        return

    if d.startswith('cancel_'):
        code = d[7:]
        if code in orders:
            orders[code]['status'] = 'cancelled'
            save_data()
        answer_cb(cbid, '❌ Bekor qilindi')
        send_buyer(uid, f"❌ #{code} bekor qilindi.")
        return

    if d.startswith('rate_start_'):
        pid = d[11:]
        answer_cb(cbid)
        send_buyer(uid, "⭐ Sotuvchiga baho bering:",
            {'inline_keyboard': [[
                {'text': '⭐',         'callback_data': f'rate_{pid}_1'},
                {'text': '⭐⭐',       'callback_data': f'rate_{pid}_2'},
                {'text': '⭐⭐⭐',     'callback_data': f'rate_{pid}_3'},
                {'text': '⭐⭐⭐⭐',   'callback_data': f'rate_{pid}_4'},
                {'text': '⭐⭐⭐⭐⭐', 'callback_data': f'rate_{pid}_5'},
            ]]}
        )
        return

    if d.startswith('rate_') and not d.startswith('rate_start_'):
        parts  = d.split('_')
        pid    = parts[1]
        rating = int(parts[2])
        answer_cb(cbid, f"{'⭐'*rating} Baho berildi!")
        p   = products.get(pid, {})
        uname = cb['from'].get('first_name', 'Xaridor')
        sid = p.get('seller_id')
        if sid:
            send_seller(sid, f"⭐ <b>Yangi baho!</b>\n\n📦 {p.get('name','')}\n{'⭐'*rating} — {uname}")
        send_buyer(uid, "✅ Rahmat! Sharhingiz uchun bonus: +5,000 so'm cashback")
        get_profile(uid)['cashback'] = get_profile(uid).get('cashback', 0) + 5000
        return

    if d.startswith('save_'):
        pid = d[5:]
        if uid not in wishlists: wishlists[uid] = []
        if pid not in wishlists[uid]:
            wishlists[uid].append(pid)
            answer_cb(cbid, '✅ Wishlistga saqlandi!')
        else:
            answer_cb(cbid, '✅ Allaqachon saqlangan!')
        return

    if d.startswith('variant_'):
        parts   = d.split('_', 2)
        pid     = parts[1]
        variant = parts[2] if len(parts) > 2 else ''
        if pid not in products:
            answer_cb(cbid, '❌ Topilmadi!'); return
        p = products[pid]
        if p.get('status') == 'closed':
            answer_cb(cbid, '⛔️ Yopilgan!'); return
        if pid not in groups: groups[pid] = []
        if uid in groups[pid]:
            answer_cb(cbid, '✅ Allaqachon guruhdasiz!'); return
        answer_cb(cbid, f"✅ {variant} tanlandi!")
        code = gen_code()
        orders[code] = {
            'product_id': pid, 'user_id': uid,
            'user_name':  cb['from'].get('first_name', 'Foydalanuvchi'),
            'amount':     p['group_price'], 'type': 'group',
            'status':     'pending', 'variant': variant,
            'created':    datetime.now().strftime('%d.%m.%Y %H:%M')
        }
        save_data()
        send_buyer(uid,
            f"🛒 <b>{p.get('shop_name','Sotuvchi')} — Guruh buyurtma</b>\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📦 {p['name']}\n"
            f"🎨 Variant: <b>{variant}</b>\n"
            f"💰 {fmt(p['group_price'])} so'm\n\n"
            f"💳 <b>Payme:</b>\n"
            f"📱 <code>{PAYME_NUMBER}</code>\n"
            f"💵 <code>{fmt(p['group_price'])}</code>\n"
            f"📝 Izoh: <code>{code}</code>\n\n"
            f"⚠️ Izohga <b>{code}</b> yozing!\n"
            f"━━━━━━━━━━━━━━━\n🔒 Joynshop kafolati ostida",
            {'inline_keyboard': [
                [{'text': "✅ To'lovni tasdiqlayman", 'callback_data': f'paid_{code}'}],
                [{'text': "❌ Bekor",                 'callback_data': f'cancel_{code}'}]
            ]}
        )
        return

    if d.startswith('refund_') and not d.startswith('refund_reason_'):
        code = d[7:]
        answer_cb(cbid)
        send_buyer(uid, "↩️ <b>Qaytarish sababi:</b>",
            {'inline_keyboard': [
                [{'text': '📦 Mahsulot kelmadi', 'callback_data': f'refund_reason_notarrived_{code}'}],
                [{'text': '😕 Sifat yomon',       'callback_data': f'refund_reason_quality_{code}'}],
                [{'text': '❌ Boshqa sabab',       'callback_data': f'refund_reason_other_{code}'}],
            ]}
        )
        return

    if d.startswith('refund_reason_'):
        parts      = d.replace('refund_reason_', '').split('_')
        reason_map = {'notarrived': 'Mahsulot kelmadi', 'quality': 'Sifat yomon', 'other': 'Boshqa sabab'}
        reason_key = parts[0]
        code       = '_'.join(parts[1:])
        reason     = reason_map.get(reason_key, 'Boshqa')
        refund_requests[code] = {'user_id': uid, 'reason': reason, 'status': 'pending'}
        save_data()
        answer_cb(cbid, "✅ Qaytarish so'rovi yuborildi!")
        send_buyer(uid,
            f"✅ <b>Qaytarish so'rovi yuborildi</b>\n\n#{code}\nSabab: {reason}\n\nSotuvchi 24 soat ichida ko'rib chiqadi."
        )
        o   = orders.get(code, {})
        p   = products.get(o.get('product_id',''), {})
        sid = p.get('seller_id')
        if sid:
            send_seller(sid,
                f"↩️ <b>QAYTARISH SO'ROVI!</b>\n\n#{code}\n📦 {p.get('name','')}\n"
                f"💰 {fmt(o.get('amount',0))} so'm\nSabab: {reason}",
                {'inline_keyboard': [[
                    {'text': '✅ Qaytarish', 'callback_data': f'seller_approve_refund_{code}'},
                    {'text': '❌ Rad',       'callback_data': f'seller_deny_refund_{code}'}
                ]]}
            )
        return

def buyer_handle_msg(msg):
    cid  = msg['chat']['id']
    uid  = msg['from']['id']
    text = msg.get('text', '')

    # Deep link: /start join_xxx  /start solo_xxx  /start ref_uid
    if text.startswith('/start ') or text.startswith('/start\n'):
        param = text.split(None, 1)[1].strip() if len(text.split(None, 1)) > 1 else ''

        # buy_PID_TYPE — kanaldan kelgan, to'g'ri Mini App ga yo'naltirish
        if param.startswith('buy_'):
            parts = param.split('_')
            if len(parts) >= 2:
                pid = parts[1]
                buy_type = parts[2] if len(parts) > 2 else 'group'
                p = products.get(pid)
                if not p:
                    send_buyer(cid, "❌ Mahsulot topilmadi yoki yopilgan.")
                    return
                _base = (BACKEND_URL or APP_URL or '').rstrip('/')
                miniapp_url = f'{_base}/miniapp?pid={pid}&action=buy&type={buy_type}' if _base else None
                if miniapp_url:
                    # Faqat bitta xabar — Mini App ochish tugmasi bilan
                    send_buyer(cid,
                        f"🛒 <b>{p['name']}</b>\n\n"
                        f"Xarid qilish uchun pastdagi tugmani bosing:",
                        {'inline_keyboard': [[
                            {'text': "🛍 Xarid qilish", 'web_app': {'url': miniapp_url}}
                        ]]}
                    )
                else:
                    send_buyer(cid, "⚠️ Sayt vaqtinchalik ishlamayapti, keyinroq urinib ko'ring.")
                return

        if param.startswith('ref_'):
            try:
                referrer_uid = int(param[4:])
                if referrer_uid != uid and uid not in referral_map:
                    referral_map[uid] = referrer_uid
                    if str(referrer_uid) not in referrals:
                        referrals[str(referrer_uid)] = {'count': 0, 'cashback': 0}
                    referrals[str(referrer_uid)]['count']   += 1
                    referrals[str(referrer_uid)]['cashback'] += 10000
                    prof = get_profile(referrer_uid)
                    prof['cashback']  = prof.get('cashback', 0) + 10000
                    prof['referrals'] = prof.get('referrals', 0) + 1
                    save_data()
                    send_buyer(referrer_uid,
                        f"🎉 <b>Yangi taklif!</b>\n\n"
                        f"Do'stingiz Joynshop ga qo'shildi!\n"
                        f"💰 +10,000 so'm cashback oldiniz!\n\n"
                        f"Jami cashback: {fmt(prof['cashback'])} so'm"
                    )
            except: pass
            # /start with ref — welcome message
            _base = (BACKEND_URL or APP_URL or '').rstrip('/')
            miniapp_url = f'{_base}/miniapp' if _base else None
            site_url    = APP_URL if APP_URL else None

            inline_row = []
            if site_url:
                inline_row.append({'text': "🌐 Saytga o'tish", 'url': site_url})
            if miniapp_url:
                inline_row.append({'text': "📱 Miniapp ochish", 'web_app': {'url': miniapp_url}})
            if not inline_row:
                inline_row.append({'text': "🛍 Do'konga o'tish", 'callback_data': 'open_shop'})

            send_buyer(cid,
                "👋 <b>Joynshop ga xush kelibsiz!</b>\n\n"
                "🛍 Do'stlaringiz bilan xarid qiling — <b>40% gacha tejang!</b>\n\n"
                "💰 Do'stingiz sizni taklif qildi — birinchi xariddan chegirma olasiz!",
                {
                    'keyboard': [
                        [{'text': "📋 Buyurtmalarim"}, {'text': "👤 Profilim"}],
                        [{'text': "🤍 Wishlist"},       {'text': "❓ Yordam"}],
                    ],
                    'resize_keyboard': True,
                    'is_persistent':   True,
                }
            )
            send_buyer(cid,
                "🏪 Xarid boshlash uchun:",
                {'inline_keyboard': [inline_row]}
            )
            return

        if param.startswith('join_') or param.startswith('solo_'):
            action, pid = param.split('_', 1)
            if pid not in products:
                send_buyer(cid, "❌ Mahsulot topilmadi yoki muddati o'tgan.")
                return
            p = products[pid]
            if p.get('status') == 'closed':
                send_buyer(cid, "⛔️ Bu guruh/sotuv yopilgan.")
                return

            if action == 'solo':
                if not p.get('solo_price'):
                    send_buyer(cid, "❌ Bu mahsulotda yakka sotish yo'q.")
                    return
                code = gen_code()
                orders[code] = {
                    'product_id': pid, 'user_id': uid,
                    'user_name':  msg['from'].get('first_name', 'Foydalanuvchi'),
                    'amount':     p['solo_price'], 'type': 'solo',
                    'status':     'pending',
                    'created':    datetime.now().strftime('%d.%m.%Y %H:%M')
                }
                save_data()
                if CLICK_TOKEN:
                    photo = p.get('photos', [None])[0]
                    photo_url = f"{CDN_BASE_URL}/products/{photo}.jpg" if photo and CDN_BASE_URL else None
                    send_invoice(cid, title=p['name'],
                        description=f"{p.get('shop_name','Sotuvchi')} • Yakka • #{code}",
                        payload=code, amount=p['solo_price'], photo_url=photo_url)
                else:
                    send_buyer(cid,
                        f"🛒 <b>{p.get('shop_name','Sotuvchi')} — Yakka buyurtma</b>\n"
                        f"━━━━━━━━━━━━━━━\n"
                        f"📦 {p['name']}\n💰 {fmt(p['solo_price'])} so'm\n\n"
                        f"💳 <b>Payme:</b>\n"
                        f"📱 <code>{PAYME_NUMBER}</code>\n"
                        f"💵 <code>{fmt(p['solo_price'])}</code>\n"
                        f"📝 Izoh: <code>{code}</code>\n\n"
                        f"⚠️ Izohga <b>{code}</b> yozing!\n"
                        f"━━━━━━━━━━━━━━━\n🔒 Joynshop kafolati ostida",
                        {'inline_keyboard': [
                            [{'text': "✅ To'lovni tasdiqlayman", 'callback_data': f'paid_{code}'}],
                            [{'text': "❌ Bekor",                 'callback_data': f'cancel_{code}'}]
                        ]}
                    )
                return

            # action == 'join'
            if pid not in groups: groups[pid] = []
            if uid in groups[pid]:
                send_buyer(cid, '✅ Siz allaqachon guruhdasiz!'); return

            variants = p.get('variants', [])
            if variants:
                btns = [[{'text': v, 'callback_data': f'variant_{pid}_{v}'}] for v in variants]
                send_buyer(cid,
                    f"📦 <b>{p['name']}</b>\n\nVariantni tanlang:",
                    {'inline_keyboard': btns}
                )
                return

            code = gen_code()
            orders[code] = {
                'product_id': pid, 'user_id': uid,
                'user_name':  msg['from'].get('first_name', 'Foydalanuvchi'),
                'amount':     p['group_price'], 'type': 'group',
                'status':     'pending', 'variant': '',
                'created':    datetime.now().strftime('%d.%m.%Y %H:%M')
            }
            save_data()
            if CLICK_TOKEN:
                photo = p.get('photos', [None])[0]
                photo_url = f"{CDN_BASE_URL}/products/{photo}.jpg" if photo and CDN_BASE_URL else None
                send_invoice(cid, title=p['name'],
                    description=f"{p.get('shop_name','Sotuvchi')} • Guruh • #{code}",
                    payload=code, amount=p['group_price'], photo_url=photo_url)
            else:
                send_buyer(cid,
                    f"🛒 <b>{p.get('shop_name','Sotuvchi')} — Guruh buyurtma</b>\n"
                    f"━━━━━━━━━━━━━━━\n"
                    f"📦 {p['name']}\n💰 {fmt(p['group_price'])} so'm\n\n"
                    f"💳 <b>Payme:</b>\n"
                    f"📱 <code>{PAYME_NUMBER}</code>\n"
                    f"💵 <code>{fmt(p['group_price'])}</code>\n"
                    f"📝 Izoh: <code>{code}</code>\n\n"
                    f"⚠️ Izohga <b>{code}</b> yozing!\n"
                    f"━━━━━━━━━━━━━━━\n🔒 Joynshop kafolati ostida",
                    {'inline_keyboard': [
                        [{'text': "✅ To'lovni tasdiqlayman", 'callback_data': f'paid_{code}'}],
                        [{'text': "❌ Bekor",                 'callback_data': f'cancel_{code}'}]
                    ]}
                )
            return

    # Manzil kutilayotgan bo'lsa
    prof = get_profile(uid)
    if prof.get('awaiting_address') and not text.startswith('/'):
        code = prof['awaiting_address']
        o    = orders.get(code, {})
        p    = products.get(o.get('product_id',''), {})
        sid  = p.get('seller_id')
        orders[code]['address'] = text
        del prof['awaiting_address']
        save_data()
        send_buyer(cid,
            f"✅ <b>Manzil saqlandi!</b>\n\n"
            f"📍 {text}\n\n"
            f"Sotuvchi yetkazib berish vaqtini belgilaydi."
        )
        if sid:
            send_seller(sid,
                f"📍 <b>Yangi manzil!</b>\n\n"
                f"📦 {p.get('name','')}\n"
                f"👤 {o.get('user_name','')}\n"
                f"🆔 #{code}\n\n"
                f"📍 Manzil: <b>{text}</b>"
            )
        return

    # ── /start (parametrsiz) ──
    if text == '/start':
        _base = (BACKEND_URL or APP_URL or '').rstrip('/')
        miniapp_url = f'{_base}/miniapp' if _base else None
        site_url    = APP_URL if APP_URL else None

        # 1 ta xabar: salom matni + inline [Saytga o'tish] [Miniapp] buttonlar
        inline_row = []
        if site_url:
            inline_row.append({'text': "🌐 Saytga o'tish", 'url': site_url})
        if miniapp_url:
            inline_row.append({'text': "📱 Miniapp ochish", 'web_app': {'url': miniapp_url}})
        if not inline_row:
            inline_row.append({'text': "🛍 Do'konga o'tish", 'callback_data': 'open_shop'})

        # 1: reply keyboard (pastki menyu) — bo'sh matn bilan
        send_buyer(cid,
            "👋 <b>Joynshop ga xush kelibsiz!</b>\n\n"
            "🛍 Do'stlaringiz bilan xarid qiling — <b>40% gacha tejang!</b>",
            {'inline_keyboard': [inline_row]}
        )
        # 2: reply keyboard ni o'rnatish
        send_buyer(cid, "📌 Asosiy menyu:",
            {
                'keyboard': [
                    [{'text': "📋 Buyurtmalarim"}, {'text': "👤 Profilim"}],
                    [{'text': "🤍 Wishlist"},       {'text': "❓ Yordam"}],
                ],
                'resize_keyboard': True,
                'is_persistent':   True,
            }
        )
        return

    if text == '/myprofile' or text == '👤 Profilim':
        p      = get_profile(uid)
        ref_d  = referrals.get(str(uid), {'count': 0, 'cashback': 0})
        ref_link = f"https://t.me/{BUYER_BOT_USERNAME}?start=ref_{uid}"
        send_buyer(cid,
            f"👤 <b>Profilingiz</b>\n\n"
            f"🛒 Jami xaridlar: {p['total_orders']}\n"
            f"💰 Tejagan: {fmt(p['total_saved'])} so'm\n"
            f"👥 Guruhlar: {p['groups_joined']}\n"
            f"🎁 Cashback: {fmt(p['cashback'])} so'm\n"
            f"👫 Taklif qilganlar: {ref_d['count']} kishi\n\n"
            f"━━━━━━━━━━━━━━━\n"
            f"🔗 Referral linkingiz:\n"
            f"<code>{ref_link}</code>\n\n"
            f"Har taklif uchun +10,000 so'm cashback!",
            {'inline_keyboard': [
                [{'text': "🔗 Referral linkni ulashish", 'url': f"https://t.me/share/url?url={ref_link}&text=🛍%20Do'stlarim%20bilan%20birgalikda%20xarid%20qilib%2040%25%20gacha%20tejayapman!%20Sen%20ham%20ulab%20ko'r%20👇"}],
            ]}
        )
        return

    if text == '/mystatus' or text == '📋 Buyurtmalarim':
        my = {k:v for k,v in orders.items() if v['user_id']==uid}
        if not my:
            send_buyer(cid, "📋 Buyurtma yo'q."); return
        r  = "📋 <b>Buyurtmalaringiz:</b>\n\n"
        em = {'pending':'⏳','confirming':'🔄','confirmed':'✅','rejected':'❌','cancelled':'🚫'}
        st = {'pending':"To'lov kutilmoqda",'confirming':'Tekshirilmoqda',
              'confirmed':'Tasdiqlandi','rejected':'Rad','cancelled':'Bekor'}
        for k, o in list(my.items())[-5:]:
            p  = products.get(o['product_id'],{})
            r += f"{em.get(o['status'],'?')} <b>#{k}</b>\n{p.get('name','?')} — {fmt(o['amount'])} so'm\n{st.get(o['status'],'')}\n\n"
        send_buyer(cid, r)
        return

    if text == '/mywishlist' or text == '🤍 Wishlist':
        wl = wishlists.get(uid, [])
        if not wl:
            send_buyer(cid, "🤍 Wishlist bo'sh."); return
        r = "🤍 <b>Wishlistingiz:</b>\n\n"
        for pid in wl:
            p = products.get(pid, {})
            if not p: continue
            count = len(groups.get(pid, []))
            r    += f"📦 {p.get('name','')}\n💰 {fmt(p.get('group_price',0))} so'm\n👥 {count}/{p.get('min_group',3)}\n\n"
        send_buyer(cid, r)
        return

    if text == '/refund':
        my = {k:v for k,v in orders.items() if v['user_id']==uid and v['status']=='confirmed'}
        if not my:
            send_buyer(cid, "Qaytarish uchun tasdiqlangan buyurtma yo'q."); return
        btns = []
        for k, o in list(my.items())[-5:]:
            p = products.get(o['product_id'],{})
            btns.append([{'text': f"#{k} — {p.get('name','')}", 'callback_data': f'refund_{k}'}])
        send_buyer(cid, "Qaysi buyurtmani qaytarmoqchisiz?", {'inline_keyboard': btns})
        return

    if text == '/help' or text == '❓ Yordam':
        send_buyer(cid,
            "ℹ️ <b>Yordam</b>\n\n"
            "/mystatus   — Buyurtmalarim\n"
            "/myprofile  — Profilim\n"
            "/mywishlist — Saqlangan mahsulotlar\n"
            "/refund     — Qaytarish so'rovi\n"
            "/feedback   — Fikr bildirish\n"
            "/settings   — Sozlamalar\n\n"
            "🆘 Yordam: @joynshop_support"
        )
        return

    if text == '/feedback':
        prof = get_profile(uid)
        prof['awaiting_feedback'] = True
        save_data()
        send_buyer(cid,
            "✍️ <b>Fikr bildirish</b>\n\n"
            "Joynshop haqida fikr-mulohazalaringizni shu chatga yozib yuboring — "
            "biz albatta ko'rib chiqamiz va javob beramiz.\n\n"
            "Bekor qilish: /cancel\n"
            "Yoki to'g'ridan-to'g'ri: @joynshop_support"
        )
        return

    if text == '/settings':
        prof = get_profile(uid)
        notif_on = prof.get('notifications', True)
        send_buyer(cid,
            "⚙️ <b>Sozlamalar</b>\n\n"
            f"🔔 Bildirishnomalar: {'Yoqilgan ✅' if notif_on else 'O‘chirilgan ❌'}\n\n"
            "Profil ma'lumotlari uchun /myprofile",
            {'inline_keyboard': [[
                {'text': ("🔕 O‘chirish" if notif_on else "🔔 Yoqish"),
                 'callback_data': 'settings_toggle_notif'},
            ]]}
        )
        return

    if text == '/cancel':
        prof = get_profile(uid)
        cleared = []
        if prof.pop('awaiting_feedback', False):
            cleared.append('feedback')
        if prof.pop('awaiting_address', False):
            cleared.append('address')
        save_data()
        send_buyer(cid, "✅ Bekor qilindi." if cleared else "Hech qanday faol jarayon yo'q.")
        return

    # Awaiting feedback (oddiy text yuborilsa, fikr sifatida qabul qilamiz)
    prof = get_profile(uid)
    if prof.get('awaiting_feedback') and text and not text.startswith('/'):
        del prof['awaiting_feedback']
        save_data()
        if ADMIN_ID:
            uname = msg.get('from', {}).get('username', '')
            uname_str = f"@{uname}" if uname else f"ID {uid}"
            send_seller(ADMIN_ID,
                f"✍️ <b>Yangi fikr</b>\n\n"
                f"👤 {uname_str}\n\n"
                f"{text[:1000]}"
            )
        send_buyer(cid, "✅ Fikringiz yuborildi. Rahmat!")
        return

# ══════════════════════════════════════════════════════════════════════
#  FLASK ROUTES
# ══════════════════════════════════════════════════════════════════════
@app.route('/', methods=['GET'])
def index():
    return 'Joynshop Bot ishlayapti! 🏪🛍'

@app.route('/api/stats', methods=['GET'])
def api_stats():
    from flask import jsonify
    pwd = request.args.get('pwd', '')
    if pwd != DASHBOARD_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401

    now         = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start  = today_start - timedelta(days=7)

    all_confirmed = [o for o in orders.values() if o['status'] == 'confirmed']
    today_orders  = [o for o in all_confirmed if datetime.strptime(o['created'], '%d.%m.%Y %H:%M') >= today_start]
    week_orders   = [o for o in all_confirmed if datetime.strptime(o['created'], '%d.%m.%Y %H:%M') >= week_start]

    gmv_total = sum(o['amount'] for o in all_confirmed)
    gmv_today = sum(o['amount'] for o in today_orders)
    gmv_week  = sum(o['amount'] for o in week_orders)

    all_products  = list(products.values())
    active_prods  = [p for p in all_products if p.get('status') != 'closed']
    filled_groups = [p for p in all_products if len(groups.get(list(products.keys())[list(products.values()).index(p)], [])) >= p.get('min_group', 99)]

    unique_sellers = len(set(p.get('seller_id') for p in all_products if p.get('seller_id')))
    unique_buyers  = len(set(o['user_id'] for o in all_confirmed))
    today_buyers   = len(set(o['user_id'] for o in today_orders))
    total_attempts = len([o for o in orders.values()])
    conv_rate      = round(len(all_confirmed) / total_attempts * 100) if total_attempts else 0

    daily_data = []
    for i in range(13, -1, -1):
        day     = today_start - timedelta(days=i)
        day_end = day + timedelta(days=1)
        day_orders = [o for o in all_confirmed
                      if datetime.strptime(o['created'], '%d.%m.%Y %H:%M') >= day
                      and datetime.strptime(o['created'], '%d.%m.%Y %H:%M') < day_end]
        daily_data.append({
            'date':   day.strftime('%d.%m'),
            'gmv':    sum(o['amount'] for o in day_orders),
            'orders': len(day_orders)
        })

    return jsonify({
        'sellers': {
            'total':    unique_sellers,
            'channels': len(verified_channels),
            'channel_list': [
                {
                    'username': ch,
                    'owner':    data.get('owner_id'),
                    'mods':     len(data.get('moderators', [])),
                    'products': sum(1 for p in products.values() if p.get('seller_channel') == ch and p.get('status') != 'closed')
                }
                for ch, data in verified_channels.items()
            ],
        },
        'referrals': {
            'total_referrers': len(referrals),
            'total_referred':  len(referral_map),
            'total_cashback':  sum(v.get('cashback', 0) for v in referrals.values()),
        },
        'buyers': {
            'total': unique_buyers,
            'today': today_buyers,
        },
        'products': {
            'total':  len(all_products),
            'active': len(active_prods),
            'filled': len(filled_groups),
        },
        'orders': {
            'total':      len(all_confirmed),
            'today':      len(today_orders),
            'week':       len(week_orders),
            'conversion': conv_rate,
        },
        'finance': {
            'gmv_total':        gmv_total,
            'gmv_today':        gmv_today,
            'gmv_week':         gmv_week,
            'commission_total': int(gmv_total * COMMISSION_RATE),
            'commission_week':  int(gmv_week  * COMMISSION_RATE),
            'avg_order':        int(gmv_total / len(all_confirmed)) if all_confirmed else 0,
        },
        'chart': daily_data,
    })

@app.route('/api/user/<int:uid>/orders', methods=['GET'])
def api_user_orders(uid):
    from flask import jsonify
    my = {k: v for k, v in orders.items() if v.get('user_id') == uid}
    result = []
    em = {'pending':'⏳','confirming':'🔄','confirmed':'✅','rejected':'❌','cancelled':'🚫'}
    st = {'pending':"To'lov kutilmoqda",'confirming':'Tekshirilmoqda',
          'confirmed':'Tasdiqlandi','rejected':'Rad etildi','cancelled':'Bekor qilindi'}
    for code, o in sorted(my.items(), key=lambda x: x[1].get('created',''), reverse=True)[:20]:
        p = products.get(o.get('product_id',''), {})
        result.append({
            'code':        code,
            'name':        p.get('name',''),
            'shop_name':   p.get('shop_name',''),
            'amount':      o.get('amount', 0),
            'type':        o.get('type','group'),
            'status':      o.get('status',''),
            'status_text': st.get(o.get('status',''), ''),
            'status_icon': em.get(o.get('status',''), '?'),
            'created':     o.get('created',''),
            'address':     o.get('address',''),
            'photo_id':    p.get('photo_id',''),
            'delivery':    p.get('delivery_type','pickup'),
        })
    return jsonify(result)

@app.route('/api/user/<int:uid>/profile', methods=['GET'])
def api_user_profile(uid):
    from flask import jsonify
    prof    = get_profile(uid)
    ref_d   = referrals.get(str(uid), {'count': 0, 'cashback': 0})
    my_ords = {k:v for k,v in orders.items() if v.get('user_id')==uid and v.get('status')=='confirmed'}
    return jsonify({
        'total_orders':    prof.get('total_orders', 0),
        'total_saved':     prof.get('total_saved', 0),
        'groups_joined':   prof.get('groups_joined', 0),
        'cashback':        prof.get('cashback', 0),
        'referral_count':  ref_d.get('count', 0),
        'confirmed_orders': len(my_ords),
    })

@app.route('/api/photo/<file_id>', methods=['GET'])
def api_photo(file_id):
    from flask import redirect
    for p in products.values():
        if p.get('photo_id') == file_id and p.get('photo_url'):
            return redirect(p['photo_url'])
        if file_id in p.get('photo_ids', []):
            urls = p.get('photo_urls', [])
            idx  = p['photo_ids'].index(file_id) if file_id in p.get('photo_ids',[]) else -1
            if idx >= 0 and idx < len(urls) and urls[idx]:
                return redirect(urls[idx])
    cached = _photo_url_cache.get(file_id)
    if cached:
        return redirect(cached)
    s3_url = upload_photo_to_s3(file_id, SELLER_TOKEN)
    if s3_url:
        _photo_url_cache[file_id] = s3_url
        return redirect(s3_url)
    try:
        result = requests.get(
            f'https://api.telegram.org/bot{SELLER_TOKEN}/getFile',
            params={'file_id': file_id}
        ).json()
        if result.get('ok'):
            path = result['result']['file_path']
            url  = f'https://api.telegram.org/file/bot{SELLER_TOKEN}/{path}'
            _photo_url_cache[file_id] = url
            return redirect(url)
    except: pass
    return '', 404

@app.route('/fonts/<path:filename>', methods=['GET'])
def serve_font(filename):
    from flask import send_from_directory, Response
    font_path = os.path.join(os.getcwd(), 'fonts')
    try:
        return send_from_directory(font_path, filename, mimetype='font/truetype')
    except:
        return Response('', 404)

@app.route('/miniapp', methods=['GET'])
def miniapp():
    from flask import Response
    html = open('miniapp.html').read()
    return Response(html, mimetype='text/html')

@app.route('/pay/<pid>', methods=['GET'])
def pay_page(pid):
    """To'lov sahifasi — kanal postidan kelgan xaridorlar uchun"""
    from flask import Response
    try:
        html = open('pay.html').read()
        return Response(html, mimetype='text/html')
    except FileNotFoundError:
        return "<h1>Pay page</h1><p>pay.html fayli topilmadi</p>", 404

@app.route('/live/<live_id>', methods=['GET'])
def live_page(live_id):
    """Live tomosha qilish sahifasi"""
    from flask import Response
    try:
        html = open('live.html').read()
        return Response(html, mimetype='text/html')
    except FileNotFoundError:
        return "<h1>Live page</h1><p>live.html fayli topilmadi</p>", 404

@app.route('/api/lives', methods=['GET'])
def api_lives():
    """Faol live'lar ro'yxati"""
    from datetime import datetime as _dt
    now = _dt.now()
    result = []
    for lid, lv in lives.items():
        if lv.get('status') != 'live':
            continue
        # Vaqt tugaganmi?
        try:
            ends = _dt.strptime(lv['ends_at'], '%Y-%m-%d %H:%M')
            if ends < now:
                lv['status'] = 'ended'
                continue
            seconds_left = int((ends - now).total_seconds())
        except:
            seconds_left = 0
        p = products.get(lv.get('product_id',''), {})
        if not p:
            continue
        result.append({
            'id':            lid,
            'product_id':    lv['product_id'],
            'product_name':  p.get('name',''),
            'shop_name':     p.get('shop_name',''),
            'video_file_id': lv.get('video_file_id',''),
            'group_price':   p.get('group_price', 0),
            'live_price':    lv.get('live_price', 0),
            'discount_pct':  lv.get('discount_pct', 0),
            'viewer_count':  lv.get('viewer_count', 0),
            'joiner_count':  len(lv.get('joiners', [])),
            'min_group':     p.get('min_group', 0),
            'count':         len(groups.get(lv.get('product_id',''), [])),
            'seconds_left':  seconds_left,
            'photo_id':      p.get('photo_id', ''),
            'photo_url':     p.get('photo_url', ''),
        })
    save_data()
    result.sort(key=lambda x: x['viewer_count'], reverse=True)
    return jsonify(result)

@app.route('/api/live/<live_id>', methods=['GET'])
def api_live(live_id):
    """Bitta live ma'lumoti — live.html sahifasi uchun"""
    from datetime import datetime as _dt
    lv = lives.get(live_id)
    if not lv:
        return jsonify({'ok': False, 'error': 'Live topilmadi'}), 404
    p = products.get(lv.get('product_id',''), {})
    try:
        ends = _dt.strptime(lv['ends_at'], '%Y-%m-%d %H:%M')
        seconds_left = max(0, int((ends - _dt.now()).total_seconds()))
        if seconds_left == 0 and lv.get('status') == 'live':
            lv['status'] = 'ended'
            save_data()
    except:
        seconds_left = 0
    # Video URL — Telegram file_id orqali
    video_url = None
    file_id   = lv.get('video_file_id', '')
    if file_id:
        try:
            r = requests.get(
                f'https://api.telegram.org/bot{SELLER_TOKEN}/getFile',
                params={'file_id': file_id}, timeout=5
            ).json()
            if r.get('ok'):
                file_path = r['result']['file_path']
                video_url = f'https://api.telegram.org/file/bot{SELLER_TOKEN}/{file_path}'
                logging.info(f"getFile ok for live={live_id} file_id={file_id} -> {file_path}")
            else:
                logging.error(f"getFile failed for live={live_id} file_id={file_id}: {r}")
        except Exception as e:
            logging.error(f"Video URL fetch exception for live={live_id} file_id={file_id}: {e}", exc_info=True)
    else:
        logging.warning(f"Live {live_id} has no video_file_id (lv keys={list(lv.keys())})")
    return jsonify({
        'ok':           True,
        'id':           live_id,
        'status':       lv.get('status', 'live'),
        'product_id':   lv.get('product_id', ''),
        'product_name': p.get('name', ''),
        'shop_name':    p.get('shop_name', ''),
        'description':  p.get('description', ''),
        'video_url':    video_url,
        'group_price':  p.get('group_price', 0),
        'live_price':   lv.get('live_price', 0),
        'original_price': p.get('original_price', 0),
        'discount_pct': lv.get('discount_pct', 0),
        'viewer_count': lv.get('viewer_count', 0),
        'joiner_count': len(lv.get('joiners', [])),
        'min_group':    p.get('min_group', 0),
        'count':        len(groups.get(lv.get('product_id',''), [])),
        'seconds_left': seconds_left,
        'photo_id':     p.get('photo_id', ''),
        'photo_url':    p.get('photo_url', ''),
        'contact':      p.get('contact', ''),
        'questions':    lv.get('questions', [])[-20:],
    })

@app.route('/api/live/<live_id>/view', methods=['POST'])
def api_live_view(live_id):
    """Live ko'rganni qayd qilish"""
    lv = lives.get(live_id)
    if not lv:
        return jsonify({'ok': False}), 404
    data = request.json or {}
    user_id = data.get('user_id', 0)
    viewers = lv.setdefault('viewers', [])
    if user_id and user_id not in viewers:
        viewers.append(user_id)
        lv['viewer_count'] = len(viewers)
        save_data()
        # Sotuvchiga real-time update (har 5-chi ko'ruvchida)
        if len(viewers) % 5 == 0:
            seller_id = lv.get('seller_id')
            if seller_id:
                p = products.get(lv['product_id'], {})
                send_seller(seller_id,
                    f"🔥 <b>{p.get('name','')}</b> — Live\n"
                    f"👀 {len(viewers)} kishi tomosha qilyapti!"
                )
    return jsonify({'ok': True, 'viewer_count': lv['viewer_count']})

@app.route('/api/live/<live_id>/question', methods=['POST'])
def api_live_question(live_id):
    """Live ga savol yuborish"""
    lv = lives.get(live_id)
    if not lv:
        return jsonify({'ok': False}), 404
    data = request.json or {}
    q = {
        'user_name': data.get('user_name', 'Mehmon')[:30],
        'text':      data.get('text', '')[:200],
        'time':      datetime.now().strftime('%H:%M'),
    }
    if not q['text']:
        return jsonify({'ok': False}), 400
    lv.setdefault('questions', []).append(q)
    save_data()
    # Sotuvchiga xabar
    seller_id = lv.get('seller_id')
    if seller_id:
        p = products.get(lv['product_id'], {})
        send_seller(seller_id,
            f"💬 <b>Live savol</b> — {p.get('name','')}\n\n"
            f"<b>{q['user_name']}:</b> {q['text']}"
        )
    return jsonify({'ok': True})

@app.route('/api/product/<pid>', methods=['GET'])
def api_product(pid):
    """Bitta mahsulot ma'lumoti — pay sahifasi uchun"""
    p = products.get(pid)
    if not p or not p.get('is_active', True):
        return jsonify({'ok': False, 'error': 'Mahsulot topilmadi'}), 404
    if p.get('status') == 'closed':
        return jsonify({'ok': False, 'error': 'Mahsulot yopilgan'}), 400
    count = len(groups.get(pid, []))
    return jsonify({
        'ok':            True,
        'id':            pid,
        'name':          p['name'],
        'description':   p.get('description', ''),
        'photos':        p.get('photo_ids', []) or ([p.get('photo_id')] if p.get('photo_id') else []),
        'photo_url':     p.get('photo_url', ''),
        'photo_urls':    p.get('photo_urls', []),
        'category':      p.get('category', ''),
        'original_price': p['original_price'],
        'solo_price':    p.get('solo_price', 0),
        'group_price':   p['group_price'],
        'min_group':     p['min_group'],
        'count':         count,
        'stock':         p.get('stock', 9999),
        'sale_type':     p.get('sale_type', 'both'),
        'variants':      p.get('variants', []),
        'shop_name':     p.get('shop_name', ''),
        'contact':       p.get('contact', ''),
        'address':       p.get('address', ''),
        'deadline':      p.get('deadline', ''),
    })

@app.route('/manifest.json', methods=['GET'])
def manifest():
    from flask import Response
    import json as _json
    data = {
        "name": "Joynshop",
        "short_name": "Joynshop",
        "description": "Guruh xarid platformasi — 40% gacha tejang!",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#FA7319",
        "orientation": "portrait-primary",
        "icons": [
            {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"}
        ],
        "categories": ["shopping"],
        "lang": "uz"
    }
    return Response(_json.dumps(data), mimetype='application/manifest+json')

@app.route('/sw.js', methods=['GET'])
def service_worker():
    from flask import Response
    try:
        sw = open('sw.js').read()
    except:
        sw = "self.addEventListener('fetch',()=>{});"
    return Response(sw, mimetype='application/javascript')

@app.route('/api/buyer_stats', methods=['GET'])
def api_buyer_stats():
    from flask import jsonify
    uid = request.args.get('uid', type=int)
    if not uid: return jsonify({}), 400
    p   = buyer_profiles.get(uid, {})
    ref = referrals.get(str(uid), {})
    return jsonify({
        'total_orders':  p.get('total_orders', 0),
        'total_saved':   p.get('total_saved', 0),
        'groups_joined': p.get('groups_joined', 0),
        'cashback':      p.get('cashback', 0),
        'referrals':     ref.get('count', 0),
    })

@app.route('/api/buyer_orders', methods=['GET'])
def api_buyer_orders():
    from flask import jsonify
    uid = request.args.get('uid', type=int)
    if not uid: return jsonify([]), 400
    result = []
    for oid, o in orders.items():
        if o.get('buyer_id') != uid: continue
        pid = o.get('product_id','')
        p   = products.get(pid, {})
        result.append({
            'id':        oid,
            'code':      o.get('code', oid),
            'name':      p.get('name', o.get('product_name','')),
            'shop_name': p.get('shop_name', ''),
            'photo_url': p.get('photo_url') or (f"{request.host_url}api/photo/{p.get('photo_id')}" if p.get('photo_id') else ''),
            'amount':    o.get('amount', p.get('group_price' if o.get('type')=='group' else 'solo_price', 0)),
            'type':      o.get('type', 'group'),
            'delivery':  o.get('delivery_type', 'pickup'),
            'address':   o.get('address', ''),
            'status':    o.get('status', 'pending'),
            'created':   o.get('created_at', ''),
        })
    result.sort(key=lambda x: x['created'], reverse=True)
    return jsonify(result)

@app.route('/api/products', methods=['GET'])
def api_products():
    from flask import jsonify
    result = []
    for pid, p in products.items():
        if p.get('status') == 'closed': continue
        if not p.get('is_active', True): continue
        count = len(groups.get(pid, []))
        min_g = p.get('min_group', 3)
        orig  = p.get('original_price', 0)
        solo  = p.get('solo_price', 0)
        grp   = p.get('group_price', 0)
        result.append({
            'id':             pid,
            'name':           p.get('name',''),
            'shop_name':      p.get('shop_name',''),
            'description':    p.get('description',''),
            'category':       p.get('category',''),
            'sale_type':      p.get('sale_type','both'),
            'original_price': orig,
            'solo_price':     solo,
            'group_price':    grp,
            'min_group':      min_g,
            'count':          count,
            'deadline':       p.get('deadline',''),
            'photo_id':       p.get('photo_id',''),
            'contact':        p.get('contact',''),
            'solo_disc':      round((orig-solo)/orig*100) if solo and orig else 0,
            'grp_disc':       round((orig-grp)/orig*100)  if grp  and orig else 0,
            'join_url':       f"https://t.me/{BUYER_BOT_USERNAME}?start=join_{pid}",
            'solo_url':       f"https://t.me/{BUYER_BOT_USERNAME}?start=solo_{pid}" if solo else None,
            'photo_ids':      p.get('photo_ids', [p.get('photo_id')] if p.get('photo_id') else []),
            'photo_url':      p.get('photo_url', ''),
            'photo_urls':     p.get('photo_urls', []),
            'seller_channel': p.get('seller_channel',''),
            'solo_available': p.get('solo_available', True),
            'variants':       p.get('variants', []),
            'stock':          p.get('stock', 9999),
        })
    result.sort(key=lambda x: x['count'], reverse=True)
    return jsonify(result)

# ─── ADMIN API ───────────────────────────────────────────────────────
def admin_auth(req):
    return req.args.get('pwd','') == DASHBOARD_PASSWORD

@app.route('/api/admin/orders', methods=['GET'])
def api_admin_orders():
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    page=int(request.args.get('page',1)); status=request.args.get('status',''); per=20
    all_orders=list(orders.values())
    if status: all_orders=[o for o in all_orders if o.get('status')==status]
    all_orders.sort(key=lambda o:o.get('created',''),reverse=True)
    total=len(all_orders); pages=max(1,(total+per-1)//per)
    chunk=all_orders[(page-1)*per:page*per]
    result=[]
    for o in chunk:
        result.append({'code':o.get('code',''),'product':products.get(o.get('product_id',''),{}).get('name','—'),
            'shop':products.get(o.get('product_id',''),{}).get('shop_name',''),
            'buyer_id':o.get('user_id',''),'buyer_name':o.get('user_name','—'),
            'amount':o.get('amount',0),'type':o.get('type','group'),
            'status':o.get('status',''),'status_icon':o.get('status_icon',''),
            'status_text':o.get('status_text',o.get('status','')),
            'created':o.get('created',''),'delivery':o.get('delivery',''),'address':o.get('address','')})
    return jsonify({'orders':result,'total':total,'pages':pages,'page':page})

@app.route('/api/admin/order/<code>/confirm', methods=['POST'])
def api_admin_confirm(code):
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    o=orders.get(code)
    if not o: return jsonify({'error':'Not found'}),404
    o['status']='confirmed';o['status_text']='Tasdiqlandi';o['status_icon']='✅'
    save_data(); return jsonify({'ok':True})

@app.route('/api/admin/order/<code>/reject', methods=['POST'])
def api_admin_reject(code):
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    o=orders.get(code)
    if not o: return jsonify({'error':'Not found'}),404
    o['status']='rejected';o['status_text']='Rad etildi';o['status_icon']='❌'
    save_data(); return jsonify({'ok':True})

@app.route('/api/admin/products', methods=['GET'])
def api_admin_products():
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    result=[]
    for pid,p in products.items():
        count=len(groups.get(pid,[])); orig=p.get('original_price',0); grp=p.get('group_price',0)
        disc=round((orig-grp)/orig*100) if orig else 0
        result.append({'id':pid,'name':p.get('name',''),'shop_name':p.get('shop_name',''),
            'group_price':grp,'original_price':orig,'grp_disc':disc,
            'min_group':p.get('min_group',0),'count':count,'status':p.get('status','active'),
            'deadline':p.get('deadline',''),'seller_id':p.get('seller_id','')})
    result.sort(key=lambda x:x['status']!='active'); return jsonify(result)

@app.route('/api/admin/product/<pid>/close', methods=['POST'])
def api_admin_close(pid):
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    p=products.get(pid)
    if not p: return jsonify({'error':'Not found'}),404
    p['status']='closed'; save_data(); return jsonify({'ok':True})

@app.route('/api/admin/product/<pid>/extend', methods=['POST'])
def api_admin_extend(pid):
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    p=products.get(pid)
    if not p: return jsonify({'error':'Not found'}),404
    try: dt=datetime.strptime(p['deadline_dt'],'%Y-%m-%d %H:%M')+timedelta(hours=24)
    except: dt=datetime.now()+timedelta(hours=24)
    p['deadline']=dt.strftime('%d.%m.%Y %H:%M'); p['deadline_dt']=dt.strftime('%Y-%m-%d %H:%M')
    save_data(); return jsonify({'ok':True,'deadline':p['deadline']})

@app.route('/api/admin/buyers', methods=['GET'])
def api_admin_buyers():
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    page=int(request.args.get('page',1)); per=20
    stats={}
    for o in orders.values():
        uid=str(o.get('user_id',''))
        if not uid: continue
        if uid not in stats:
            stats[uid]={'uid':uid,'total_spent':0,'confirmed':0,'cashback':0,'referrals':0,'last_order':''}
        if o.get('status')=='confirmed':
            stats[uid]['total_spent']+=o.get('amount',0)
            stats[uid]['confirmed']+=1
            stats[uid]['cashback']+=int(o.get('amount',0)*0.02)
        if o.get('created','')>stats[uid]['last_order']:
            stats[uid]['last_order']=o.get('created','')
    for ref_uid,refs in referrals.items():
        uid=str(ref_uid)
        if uid in stats: stats[uid]['referrals']=len(refs)
    buyers_list=sorted(stats.values(),key=lambda x:x['total_spent'],reverse=True)
    total=len(buyers_list); pages=max(1,(total+per-1)//per)
    chunk=buyers_list[(page-1)*per:page*per]
    return jsonify({'buyers':chunk,'total':total,'pages':pages,'page':page})

@app.route('/api/admin/sellers', methods=['GET'])
def api_admin_sellers():
    from flask import jsonify
    if not admin_auth(request): return jsonify({'error':'Unauthorized'}),401
    result=[]
    for uid,pids in seller_products.items():
        revenue=0; order_cnt=0; active_cnt=0; channels=[]
        for pid in pids:
            p=products.get(pid,{})
            if not p: continue
            if p.get('status')=='active': active_cnt+=1
            ch=p.get('seller_channel','')
            if ch and ch not in channels: channels.append(ch)
            for o in orders.values():
                if o.get('product_id')==pid and o.get('status')=='confirmed':
                    revenue+=o.get('amount',0); order_cnt+=1
        commission=int(revenue*COMMISSION_RATE); payout=revenue-commission
        result.append({'uid':str(uid),'channels':channels,'products':len(pids),
            'active':active_cnt,'revenue':revenue,'commission':commission,
            'payout':payout,'orders':order_cnt})
    result.sort(key=lambda x:x['revenue'],reverse=True); return jsonify(result)

@app.route('/setup-menu', methods=['GET'])
def setup_menu_route():
    from flask import jsonify
    if request.args.get('key','')!=DASHBOARD_PASSWORD:
        return jsonify({'ok':False,'error':'unauthorized'}),403
    miniapp_url = f"{(BACKEND_URL or APP_URL or '').rstrip('/')}/miniapp" if (BACKEND_URL or APP_URL) else None
    results={}
    if BUYER_TOKEN and miniapp_url:
        r=requests.post(f'https://api.telegram.org/bot{BUYER_TOKEN}/setChatMenuButton',
            json={'menu_button':{'type':'web_app','text':'🛍 Joynshop','web_app':{'url':miniapp_url}}}).json()
        results['buyer_menu']=r
    elif BUYER_TOKEN:
        results['buyer_menu']='APP_URL not set'
    if SELLER_TOKEN:
        r=requests.post(f'https://api.telegram.org/bot{SELLER_TOKEN}/setChatMenuButton',
            json={'menu_button':{'type':'commands'}}).json()
        results['seller_menu']=r
    results['miniapp_url']=miniapp_url; results['APP_URL']=APP_URL; results['BACKEND_URL']=BACKEND_URL
    return jsonify({'ok':True,'results':results})

@app.route('/dashboard', methods=['GET'])
def dashboard():
    from flask import Response
    html = open('dashboard.html').read()
    return Response(html, mimetype='text/html')


@app.route('/api/wishlist', methods=['GET'])
def api_wishlist_get():
    from flask import jsonify
    uid = request.args.get('uid', type=int)
    if not uid: return jsonify([]), 400
    wl  = wishlists.get(uid, [])
    result = []
    for pid in wl:
        p = products.get(pid)
        if not p or p.get('status') == 'closed': continue
        count = len(groups.get(pid, []))
        orig  = p.get('original_price', 0)
        grp   = p.get('group_price', 0)
        result.append({
            'id':             pid,
            'name':           p.get('name', ''),
            'shop_name':      p.get('shop_name', ''),
            'group_price':    grp,
            'original_price': orig,
            'grp_disc':       round((orig - grp) / orig * 100) if orig else 0,
            'min_group':      p.get('min_group', 3),
            'count':          count,
            'photo_id':       p.get('photo_id', ''),
            'photo_url':      p.get('photo_url', ''),
            'deadline':       p.get('deadline', ''),
            'join_url':       f"https://t.me/{BUYER_BOT_USERNAME}?start=join_{pid}",
        })
    return jsonify(result)


@app.route('/api/wishlist/add', methods=['POST'])
def api_wishlist_add():
    from flask import jsonify
    data = request.json or {}
    uid  = data.get('uid')
    pid  = data.get('pid', '')
    if not uid or not pid:
        return jsonify({'ok': False, 'error': 'uid va pid kerak'}), 400
    uid = int(uid)
    if pid not in products:
        return jsonify({'ok': False, 'error': 'Mahsulot topilmadi'}), 404
    if uid not in wishlists: wishlists[uid] = []
    if pid not in wishlists[uid]:
        wishlists[uid].append(pid)
        save_data()
    return jsonify({'ok': True, 'count': len(wishlists[uid])})


@app.route('/api/wishlist/remove', methods=['POST'])
def api_wishlist_remove():
    from flask import jsonify
    data = request.json or {}
    uid  = data.get('uid')
    pid  = data.get('pid', '')
    if not uid or not pid:
        return jsonify({'ok': False, 'error': 'uid va pid kerak'}), 400
    uid = int(uid)
    if uid in wishlists and pid in wishlists[uid]:
        wishlists[uid].remove(pid)
        save_data()
    return jsonify({'ok': True, 'count': len(wishlists.get(uid, []))})

@app.route('/api/categories', methods=['GET'])
def api_categories():
    from flask import jsonify
    # Mahsulotlarda mavjud kategoriyalarni va sonini qaytaradi
    cat_counts = {}
    for p in products.values():
        if p.get('status') == 'closed': continue
        if not p.get('is_active', True): continue
        cat = p.get('category', '')
        if cat:
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
    # CATEGORIES tartibida qaytaramiz
    result = []
    for name, icon in CATEGORIES:
        if name in cat_counts:
            result.append({'name': name, 'icon': icon, 'count': cat_counts[name]})
    return jsonify(result)


@app.route('/api/web_checkout', methods=['POST'])
def api_web_checkout():
    """Sayt to'lov sahifasidan kelgan buyurtma — kanal postidan kelgan xaridorlar"""
    data = request.json or {}
    pid       = data.get('product_id', '')
    otype     = data.get('type', 'group')
    user_name = data.get('user_name', '').strip()
    user_phone = data.get('user_phone', '').strip()
    variant   = data.get('variant', '')
    delivery  = data.get('delivery', 'pickup')
    address   = data.get('address', '')

    # Validatsiya
    if not pid:
        return jsonify({'ok': False, 'error': 'product_id kerak'}), 400
    if not user_name or not user_phone:
        return jsonify({'ok': False, 'error': "Ism va telefon kiritilishi shart"}), 400

    p = products.get(pid)
    if not p or not p.get('is_active', True):
        return jsonify({'ok': False, 'error': 'Mahsulot topilmadi'}), 404
    if p.get('status') == 'closed':
        return jsonify({'ok': False, 'error': 'Mahsulot yopilgan'}), 400

    # Narx
    if otype == 'solo':
        if not p.get('solo_price'):
            return jsonify({'ok': False, 'error': 'Yakka sotish mavjud emas'}), 400
        amount = p['solo_price']
    else:
        amount = p['group_price']

    # Buyurtma yaratish (web_user_id — Telegram ID yo'q, telefon orqali)
    code = gen_code()
    orders[code] = {
        'product_id':  pid,
        'user_id':     0,  # web user, Telegram ID yo'q
        'user_name':   user_name,
        'user_phone':  user_phone,
        'amount':      amount,
        'type':        otype,
        'variant':     variant,
        'delivery':    delivery,
        'address':     address,
        'status':      'pending',
        'created':     datetime.now().strftime('%d.%m.%Y %H:%M'),
        'source':      'web',
    }
    save_data()

    # Sotuvchiga xabar
    sid = p.get('seller_id')
    if sid:
        variant_line  = f"\n🎨 Variant: <b>{variant}</b>" if variant else ''
        delivery_text = "🚚 Yetkazib berish" if delivery == 'deliver' else "🏪 Olib ketish"
        address_line  = f"\n📍 Manzil: {address}" if address else ''
        send_seller(sid,
            f"🔔 <b>YANGI BUYURTMA (Sayt)</b>\n\n"
            f"📦 {p.get('name','')}{variant_line}\n"
            f"👤 {user_name}\n"
            f"📞 <code>{user_phone}</code>\n"
            f"💰 {fmt(amount)} so'm\n"
            f"🛒 {'Yakka' if otype=='solo' else 'Guruh'}\n"
            f"🚚 {delivery_text}{address_line}\n"
            f"🆔 #{code}",
            {'inline_keyboard': [[
                {'text': '✅ Tasdiqlash', 'callback_data': f'seller_ac_{code}'},
                {'text': '❌ Rad',        'callback_data': f'seller_ar_{code}'},
            ]]}
        )

    # Hozircha checkout_url yo'q (Click rasmiy ulanmagan)
    # Kelajakda: checkout_url = create_click_invoice(amount, code, ...)
    return jsonify({
        'ok':           True,
        'code':         code,
        'amount':       amount,
        'checkout_url': None  # kelajakda Click checkout URL
    })


@app.route('/api/checkout', methods=['POST'])
def api_checkout():
    from flask import jsonify
    data = request.json or {}

    pid       = data.get('product_id', '')
    uid       = data.get('user_id')
    user_name = data.get('user_name', 'Foydalanuvchi')
    otype     = data.get('type', 'group')       # 'group' | 'solo'
    variant   = data.get('variant', '')
    delivery  = data.get('delivery', 'pickup')  # 'pickup' | 'deliver'
    address   = data.get('address', '')

    # Validatsiya
    if not pid or not uid:
        return jsonify({'ok': False, 'error': 'product_id va user_id kerak'}), 400

    p = products.get(pid)
    if not p or not p.get('is_active', True):
        return jsonify({'ok': False, 'error': 'Mahsulot topilmadi'}), 404
    if p.get('status') == 'closed':
        return jsonify({'ok': False, 'error': 'Mahsulot yopilgan'}), 400

    uid = int(uid)

    # Solo sotuv tekshiruvi
    if otype == 'solo':
        if not p.get('solo_price'):
            return jsonify({'ok': False, 'error': 'Yakka sotish mavjud emas'}), 400
        amount = p['solo_price']
    else:
        # Guruh — allaqachon qo'shilganmi?
        if uid in groups.get(pid, []):
            return jsonify({'ok': False, 'error': 'Allaqachon guruhdasiz'}), 400
        amount = p['group_price']

    # Buyurtma yaratish
    code = gen_code()
    orders[code] = {
        'product_id': pid,
        'user_id':    uid,
        'user_name':  user_name,
        'amount':     amount,
        'type':       otype,
        'variant':    variant,
        'delivery':   delivery,
        'address':    address,
        'status':     'pending',
        'created':    datetime.now().strftime('%d.%m.%Y %H:%M'),
        'source':     'miniapp',
    }
    save_data()

    variant_line  = f"\n🎨 Variant: <b>{variant}</b>" if variant else ''
    delivery_text = "🚚 Yetkazib berish" if delivery == 'deliver' else "🏪 Olib ketish"
    address_line  = f"\n📍 Manzil: {address}" if address else ''
    sale_type     = p.get('sale_type', 'both')

    # Click invoice link yaratish (Mini App ichida ochish uchun)
    invoice_link = None
    if CLICK_TOKEN:
        price_lbl = "Yakka narx" if otype == 'solo' else "Guruh narxi"
        # description: yakka bo'lsa minimal, guruh bo'lsa guruh ma'lumoti
        if sale_type == 'solo' or otype == 'solo':
            desc = (p.get('description') or p['name'])[:255]
        else:
            count   = len(groups.get(pid, []))
            min_g   = p['min_group']
            desc    = f"👥 Guruh: {count}/{min_g} • Kerak: {max(0, min_g-count)} kishi"
            if p.get('description'):
                desc += f"\n{p['description'][:100]}"
        desc = (desc + f"\n🏪 {p.get('shop_name','')} | {p.get('contact','')}")[:255]

        photo_url = p.get('photo_url') or None
        inv_data  = {
            'title':             strip_html(p['name'])[:32],
            'description':       desc,
            'payload':           code,
            'provider_token':    CLICK_TOKEN,
            'currency':          'UZS',
            'prices':            json.dumps([{'label': price_lbl, 'amount': amount * 100}]),
            'need_name':         True,
            'need_phone_number': True,
            'need_shipping_address': False,
            'is_flexible':       False,
        }
        if photo_url and photo_url.startswith('http'):
            inv_data['photo_url']  = photo_url
            inv_data['photo_size'] = 800
        # createInvoiceLink — URL qaytaradi (Mini App tg.openInvoice uchun)
        try:
            r = requests.post(f'https://api.telegram.org/bot{BUYER_TOKEN}/createInvoiceLink', json=inv_data, timeout=10).json()
            if r.get('ok'):
                invoice_link = r['result']
            else:
                logging.warning(f"createInvoiceLink failed: {r}")
                # Fallback: oddiy sendInvoice
                inv_data['chat_id'] = uid
                requests.post(f'https://api.telegram.org/bot{BUYER_TOKEN}/sendInvoice', json=inv_data, timeout=10)
        except Exception as e:
            logging.error(f"createInvoiceLink error: {e}")
    else:
        # Fallback: qo'lda Payme
        send_buyer(uid,
            f"🛒 <b>{p.get('shop_name','Sotuvchi')} — {'Yakka' if otype=='solo' else 'Guruh'} buyurtma</b>\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📦 {p['name']}{variant_line}\n"
            f"💰 {fmt(amount)} so'm\n"
            f"🚚 {delivery_text}{address_line}\n\n"
            f"💳 <b>Payme orqali to'lang:</b>\n"
            f"📱 <code>{PAYME_NUMBER}</code>\n"
            f"💵 <code>{fmt(amount)}</code>\n"
            f"📝 Izoh: <code>{code}</code>\n\n"
            f"⚠️ Izohga <b>{code}</b> yozing!\n"
            f"━━━━━━━━━━━━━━━\n"
            f"🔒 Joynshop kafolati ostida",
            {'inline_keyboard': [
                [{'text': "✅ To'lovni tasdiqlayman", 'callback_data': f'paid_{code}'}],
                [{'text': "❌ Bekor",                 'callback_data': f'cancel_{code}'}],
            ]}
        )

    # Sotuvchiga xabar
    sid = p.get('seller_id')
    if sid:
        send_seller(sid,
            f"🔔 <b>YANGI BUYURTMA (Miniapp)</b>\n\n"
            f"📦 {p.get('name','')}{variant_line}\n"
            f"👤 {user_name} (ID: <code>{uid}</code>)\n"
            f"💰 {fmt(amount)} so'm\n"
            f"🛒 {'Yakka' if otype=='solo' else 'Guruh'}\n"
            f"🚚 {delivery_text}{address_line}\n"
            f"🆔 #{code}",
            {'inline_keyboard': [[
                {'text': '✅ Tasdiqlash', 'callback_data': f'seller_ac_{code}'},
                {'text': '❌ Rad',        'callback_data': f'seller_ar_{code}'},
            ]]}
        )

    return jsonify({'ok': True, 'code': code, 'amount': amount, 'invoice_link': invoice_link})


# ══════════════════════════════════════════════════════════════════════
#  SELLER MINI APP API (v1) — seller.joynshop.uz frontend
# ══════════════════════════════════════════════════════════════════════
# Authorization: tma <Telegram WebApp initData>
# Validatsiya: HMAC-SHA256 with secret_key=HMAC-SHA256(b'WebAppData', SELLER_TOKEN)
# auth_date freshness: 24 soat

def verify_telegram_init_data(init_data, bot_token, max_age=86400):
    """Telegram WebApp initData ni HMAC bilan tekshiradi.
    Returns user dict (id, first_name, ...) yoki None.
    """
    if not init_data or not bot_token:
        return None
    try:
        parsed = dict(urllib.parse.parse_qsl(init_data, keep_blank_values=True))
    except Exception:
        return None
    received_hash = parsed.pop('hash', None)
    if not received_hash:
        return None
    # auth_date freshness
    try:
        auth_date = int(parsed.get('auth_date', '0'))
    except (ValueError, TypeError):
        return None
    if auth_date <= 0 or (datetime.now().timestamp() - auth_date) > max_age:
        return None
    # data_check_string — sorted alphabetically, joined with \n
    data_check_string = '\n'.join(f'{k}={v}' for k, v in sorted(parsed.items()))
    secret_key   = hmac.new(b'WebAppData', bot_token.encode(), hashlib.sha256).digest()
    computed     = hmac.new(secret_key, data_check_string.encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(computed, received_hash):
        return None
    # Valid — parse user
    user_json = parsed.get('user')
    if not user_json:
        return None
    try:
        return json.loads(user_json)
    except json.JSONDecodeError:
        return None

def require_seller(fn):
    """Decorator: /api/v1/seller/* endpoint'larini Telegram initData orqali himoyalaydi.
    Sets g.seller_uid (int) va g.seller_user (dict)."""
    @functools.wraps(fn)
    def wrapped(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('tma '):
            return jsonify({'error': 'unauthorized', 'reason': 'missing_init_data'}), 401
        user = verify_telegram_init_data(auth[4:], SELLER_TOKEN)
        if not user:
            return jsonify({'error': 'unauthorized', 'reason': 'invalid_init_data'}), 401
        uid = user.get('id')
        if not uid:
            return jsonify({'error': 'unauthorized', 'reason': 'no_user_id'}), 401
        # Sotuvchi ekanligini tekshirish — eng kamida bitta shop bo'lishi kerak
        has_shop = uid in seller_shops or str(uid) in seller_shops
        if not has_shop:
            return jsonify({'error': 'forbidden', 'reason': 'not_a_seller'}), 403
        g.seller_uid  = int(uid)
        g.seller_user = user
        return fn(*args, **kwargs)
    return wrapped

def _seller_get_shops(uid):
    """seller_shops dict'idan har xil key turlarini hisobga olib shop list qaytaradi."""
    return seller_shops.get(uid) or seller_shops.get(str(uid)) or []

def _seller_get_pids(uid):
    return seller_products.get(uid) or seller_products.get(str(uid)) or []

@app.route('/api/v1/seller/me', methods=['GET'])
@require_seller
def api_seller_me():
    """Sotuvchi profili va umumiy ma'lumot."""
    uid  = g.seller_uid
    user = g.seller_user
    shops = _seller_get_shops(uid)
    shops_info = [{
        'name':              shop.get('name', ''),
        'channel':           shop.get('channel', ''),
        'billz_connected':   bool(shop.get('billz_secret_token')),
        'billz_shop_name':   shop.get('billz_shop_name', ''),
        'onboarding_status': shop.get('onboarding_status', 'active'),
    } for shop in shops]
    pids = _seller_get_pids(uid)
    products_count  = sum(1 for pid in pids
                          if pid in products and products[pid].get('status') != 'closed')
    orders_pending  = sum(1 for o in orders.values()
                          if o.get('product_id') in pids and o.get('status') == 'confirming')
    # Stats summary
    today_str  = datetime.now().strftime('%d.%m.%Y')
    week_start = datetime.now() - timedelta(days=7)
    confirmed  = [o for o in orders.values()
                  if o.get('product_id') in pids and o.get('status') == 'confirmed']
    gmv_today = sum(o.get('amount', 0) for o in confirmed
                    if o.get('created', '').startswith(today_str))
    gmv_week  = 0
    for o in confirmed:
        try:
            if datetime.strptime(o.get('created','01.01.2000 00:00'), '%d.%m.%Y %H:%M') >= week_start:
                gmv_week += o.get('amount', 0)
        except (ValueError, TypeError):
            pass
    return jsonify({
        'uid':             uid,
        'first_name':      user.get('first_name', ''),
        'last_name':       user.get('last_name', ''),
        'username':        user.get('username', ''),
        'photo_url':       user.get('photo_url', ''),
        'shops':           shops_info,
        'legal_completed': seller_has_legal(uid),
        'billz_connected': bool(seller_billz_connected_shops(uid)),
        'products_count':  products_count,
        'orders_pending':  orders_pending,
        'stats_summary': {
            'gmv_today': gmv_today,
            'gmv_week':  gmv_week,
        },
    })

@app.route('/api/v1/seller/products', methods=['GET'])
@require_seller
def api_seller_products():
    """Sotuvchi mahsulotlari ro'yxati. Pagination + filter + search."""
    uid = g.seller_uid
    try:
        page = max(0, int(request.args.get('page', 0)))
    except ValueError:
        page = 0
    try:
        limit = min(50, max(1, int(request.args.get('limit', 20))))
    except ValueError:
        limit = 20
    filt   = (request.args.get('filter', 'active') or 'active').lower()
    search = (request.args.get('search', '') or '').lower().strip()

    pids_all = _seller_get_pids(uid)
    items_filtered = []
    for pid in pids_all:
        if pid not in products:
            continue
        p   = products[pid]
        cls = _classify_product_status(p)
        # Filter
        if filt == 'active':
            if cls['archived']:
                continue
        elif filt == 'archived':
            if not cls['archived']:
                continue
        # filt == 'all' — hammasi
        # Search (name substring)
        if search and search not in p.get('name', '').lower():
            continue
        items_filtered.append((pid, p, cls))

    # Sort: Billz draft'lar yuqorida, qolgani deadline_dt asc bo'yicha
    def sort_key(item):
        pid, p, cls = item
        is_draft = (p.get('source') == 'billz' and not p.get('is_active', True))
        return (0 if is_draft else 1, p.get('deadline_dt', ''))
    items_filtered.sort(key=sort_key)

    total = len(items_filtered)
    pages = max(1, (total + limit - 1) // limit)
    page  = min(page, pages - 1) if total > 0 else 0
    chunk = items_filtered[page*limit : (page+1)*limit]

    items_out = []
    for pid, p, cls in chunk:
        photo_url = p.get('photo_url') or ''
        if not photo_url and p.get('photo_id'):
            photo_url = f'/api/photo/{p["photo_id"]}'
        is_billz_draft = (p.get('source') == 'billz' and not p.get('is_active', True))
        price_amt = p.get('group_price', 0) or p.get('original_price', 0)
        items_out.append({
            'id':              pid,
            'name':            p.get('name', ''),
            'price':           price_amt,
            'price_short':     format_price_short(price_amt),
            'original_price':  p.get('original_price', 0),
            'min_group':       p.get('min_group', 0),
            'count':           len(groups.get(pid, [])),
            'status':          p.get('status', 'active'),
            'status_label':    cls['label'],
            'status_emoji':    cls['emoji'],
            'source':          p.get('source', 'manual'),
            'is_billz_draft':  is_billz_draft,
            'mxik_missing':    not p.get('mxik_code') and cls['label'] in ('Aktiv', 'Yoqilmagan'),
            'deadline':        p.get('deadline', ''),
            'deadline_dt':     p.get('deadline_dt', ''),
            'photo_url':       photo_url,
            'shop_name':       p.get('shop_name', ''),
            'channel':         p.get('seller_channel', ''),
        })

    return jsonify({
        'items':    items_out,
        'total':    total,
        'page':     page,
        'pages':    pages,
        'has_next': (page + 1) < pages,
    })

# ─── SPRINT 1.1: Read-only endpoints ────────────────────────────────

def _parse_order_dt(o):
    """Order created sanasini datetime ga o'giradi (parse fail bo'lsa None)."""
    try:
        return datetime.strptime(o.get('created', '01.01.2000 00:00'), '%d.%m.%Y %H:%M')
    except (ValueError, TypeError):
        return None

@app.route('/api/v1/seller/stats', methods=['GET'])
@require_seller
def api_seller_stats():
    """Sotuvchi statistikasi range filter bilan.
    range: today | week | month | all (default: week)"""
    uid = g.seller_uid
    range_key = (request.args.get('range', 'week') or 'week').lower()
    if range_key not in ('today', 'week', 'month', 'all'):
        return jsonify({'error': 'invalid_range', 'reason': "range: today|week|month|all"}), 400

    pids = set(_seller_get_pids(uid))
    if not pids:
        return jsonify({
            'range': range_key,
            'gmv': 0, 'commission': 0, 'net_income': 0,
            'orders_total': 0, 'orders_confirmed': 0, 'orders_pending': 0,
            'conversion_rate': 0,
            'products_total': 0, 'products_active': 0, 'products_archived': 0,
            'groups_filled': 0, 'buyers_unique': 0, 'avg_check': 0,
            'top_products': [], 'top_customers': [],
        })

    # Range filter
    now = datetime.now()
    if range_key == 'today':
        cutoff = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif range_key == 'week':
        cutoff = now - timedelta(days=7)
    elif range_key == 'month':
        cutoff = now - timedelta(days=30)
    else:
        cutoff = None

    def in_range(o):
        if cutoff is None:
            return True
        dt = _parse_order_dt(o)
        return dt is not None and dt >= cutoff

    my_orders = [o for o in orders.values() if o.get('product_id') in pids]
    in_range_orders = [o for o in my_orders if in_range(o)]
    confirmed = [o for o in in_range_orders if o.get('status') == 'confirmed']
    pending   = [o for o in in_range_orders if o.get('status') in ('pending', 'confirming')]

    gmv         = sum(o.get('amount', 0) for o in confirmed)
    commission  = int(gmv * COMMISSION_RATE)
    net_income  = gmv - commission
    orders_total     = len(in_range_orders)
    orders_confirmed = len(confirmed)
    orders_pending   = len(pending)
    conversion = round((orders_confirmed / orders_total) * 100, 1) if orders_total else 0
    avg_check  = (gmv // orders_confirmed) if orders_confirmed else 0
    buyers_unique = len({o.get('user_id') for o in confirmed if o.get('user_id')})

    products_total = sum(1 for pid in pids if pid in products)
    products_active = 0
    products_archived = 0
    groups_filled = 0
    for pid in pids:
        p = products.get(pid)
        if not p:
            continue
        cls = _classify_product_status(p)
        if cls['archived']:
            products_archived += 1
        else:
            products_active += 1
        if len(groups.get(pid, [])) >= p.get('min_group', 9999):
            groups_filled += 1

    # Top products by revenue (in range)
    rev_by_pid = {}
    sold_by_pid = {}
    for o in confirmed:
        pid = o.get('product_id')
        rev_by_pid[pid]  = rev_by_pid.get(pid, 0) + o.get('amount', 0)
        sold_by_pid[pid] = sold_by_pid.get(pid, 0) + 1
    top_products = []
    for pid, revenue in sorted(rev_by_pid.items(), key=lambda x: x[1], reverse=True)[:5]:
        p = products.get(pid, {})
        top_products.append({
            'id':      pid,
            'name':    p.get('name', '—')[:40],
            'sold':    sold_by_pid.get(pid, 0),
            'revenue': revenue,
        })

    # Top customers (from CRM, in range)
    sid = str(uid)
    my_custs = customers.get(sid, {})
    cust_revenue = {}
    for o in confirmed:
        cuid = str(o.get('user_id', ''))
        cust_revenue[cuid] = cust_revenue.get(cuid, 0) + o.get('amount', 0)
    top_customers = []
    for cuid, spent in sorted(cust_revenue.items(), key=lambda x: x[1], reverse=True)[:5]:
        c = my_custs.get(cuid, {})
        top_customers.append({
            'cuid':   cuid,
            'name':   c.get('name', '—')[:30],
            'spent':  spent,
            'orders': sum(1 for o in confirmed if str(o.get('user_id', '')) == cuid),
        })

    return jsonify({
        'range':            range_key,
        'gmv':              gmv,
        'commission':       commission,
        'net_income':       net_income,
        'orders_total':     orders_total,
        'orders_confirmed': orders_confirmed,
        'orders_pending':   orders_pending,
        'conversion_rate':  conversion,
        'products_total':   products_total,
        'products_active':  products_active,
        'products_archived':products_archived,
        'groups_filled':    groups_filled,
        'buyers_unique':    buyers_unique,
        'avg_check':        avg_check,
        'top_products':     top_products,
        'top_customers':    top_customers,
    })

@app.route('/api/v1/seller/products/<pid>', methods=['GET'])
@require_seller
def api_seller_product_detail(pid):
    """Bitta mahsulot to'liq detail."""
    uid = g.seller_uid
    p = products.get(pid)
    if not p or p.get('seller_id') != uid:
        return jsonify({'error': 'not_found'}), 404

    cls = _classify_product_status(p)
    cat_name = p.get('category', '')
    cat_icon = next((icon for name, icon in CATEGORIES if name == cat_name), '📦')

    # Photos array — S3 first, file_id proxy fallback
    photos = []
    photo_urls = p.get('photo_urls') or ([p['photo_url']] if p.get('photo_url') else [])
    photo_ids  = p.get('photo_ids')  or ([p['photo_id']]  if p.get('photo_id')  else [])
    primary_set = False
    for url in photo_urls:
        if url:
            photos.append({'url': url, 'is_primary': not primary_set})
            primary_set = True
    if not photos:
        for fid in photo_ids:
            if fid:
                photos.append({'url': f'/api/photo/{fid}', 'is_primary': not primary_set})
                primary_set = True

    # Channel post URL
    ch_chat = p.get('channel_chat_id')
    ch_msg  = p.get('channel_message_id')
    channel_post_url = None
    if ch_chat and ch_msg:
        if isinstance(ch_chat, str) and ch_chat.startswith('@'):
            channel_post_url = f"https://t.me/{ch_chat[1:]}/{ch_msg}"
        elif isinstance(ch_chat, (int, str)):
            chat_id_str = str(ch_chat).replace('-100', '')
            channel_post_url = f"https://t.me/c/{chat_id_str}/{ch_msg}"

    # Deadline seconds left
    deadline_seconds = 0
    ddt = p.get('deadline_dt')
    if ddt:
        try:
            deadline_seconds = max(0, int((datetime.strptime(ddt, '%Y-%m-%d %H:%M') - datetime.now()).total_seconds()))
        except (ValueError, TypeError):
            pass

    # Stats — barcha vaqtlar bo'yicha
    p_orders   = [o for o in orders.values() if o.get('product_id') == pid]
    confirmed  = [o for o in p_orders if o.get('status') == 'confirmed']
    revenue    = sum(o.get('amount', 0) for o in confirmed)
    wl_count   = sum(1 for wl in wishlists.values() if pid in wl)
    first_dt = min((_parse_order_dt(o) for o in confirmed if _parse_order_dt(o)), default=None)
    last_dt  = max((_parse_order_dt(o) for o in confirmed if _parse_order_dt(o)), default=None)

    return jsonify({
        'id':                    pid,
        'name':                  p.get('name', ''),
        'description':           p.get('description', ''),
        'category':              cat_name,
        'category_icon':         cat_icon,
        'sale_type':             p.get('sale_type', 'both'),
        'original_price':        p.get('original_price', 0),
        'group_price':           p.get('group_price', 0),
        'solo_price':            p.get('solo_price', 0),
        'min_group':             p.get('min_group', 0),
        'count':                 len(groups.get(pid, [])),
        'status':                p.get('status', 'active'),
        'status_label':          cls['label'],
        'status_emoji':          cls['emoji'],
        'is_archived':           cls['archived'],
        'source':                p.get('source', 'manual'),
        'is_billz_draft':        (p.get('source') == 'billz' and not p.get('is_active', True)),
        'deadline':              p.get('deadline', ''),
        'deadline_dt':           p.get('deadline_dt', ''),
        'deadline_seconds_left': deadline_seconds,
        'photos':                photos,
        'variants':              p.get('variants', []),
        'barcode':               p.get('barcode', ''),
        'sku':                   p.get('sku', ''),
        'brand_name':            p.get('brand_name', ''),
        'shop': {
            'name':    p.get('shop_name', ''),
            'channel': p.get('seller_channel', ''),
        },
        'mxik': {
            'code':    p.get('mxik_code'),
            'name':    p.get('mxik_name'),
            'missing': not p.get('mxik_code') and cls['label'] in ('Aktiv', 'Yoqilmagan'),
        },
        'stats': {
            'orders_total':   len(confirmed),
            'revenue':        revenue,
            'wishlist_count': wl_count,
            'first_order':    first_dt.strftime('%d.%m.%Y') if first_dt else '',
            'last_order':     last_dt.strftime('%d.%m.%Y')  if last_dt  else '',
        },
        'channel_post_url':      channel_post_url,
    })

ORDER_STATUS_META = {
    'pending':    {'emoji': '⏳', 'label': "To'lov kutilmoqda"},
    'confirming': {'emoji': '🔄', 'label': "Tasdiqlash kutilmoqda"},
    'confirmed':  {'emoji': '✅', 'label': "Tasdiqlandi"},
    'rejected':   {'emoji': '❌', 'label': "Rad etildi"},
    'cancelled':  {'emoji': '🚫', 'label': "Bekor qilindi"},
}

DELIVERY_LABEL = {'pickup': 'Olib ketish', 'deliver': 'Yetkazib berish'}

def _format_order_item(code, o):
    """Order list yoki detail uchun yagona format."""
    p = products.get(o.get('product_id', ''), {})
    photo_url = p.get('photo_url') or ''
    if not photo_url and p.get('photo_id'):
        photo_url = f"/api/photo/{p['photo_id']}"
    status = o.get('status', 'pending')
    meta = ORDER_STATUS_META.get(status, {'emoji': '?', 'label': status})
    return {
        'code':           code,
        'product_id':     o.get('product_id', ''),
        'product_name':   p.get('name', '')[:60],
        'product_photo':  photo_url,
        'buyer': {
            'user_id':  o.get('user_id', 0),
            'name':     o.get('user_name', ''),
            'phone':    o.get('user_phone', ''),
            'username': o.get('username', ''),
        },
        'amount':         o.get('amount', 0),
        'type':           o.get('type', 'group'),
        'type_label':     'Yakka' if o.get('type') == 'solo' else 'Guruh',
        'variant':        o.get('variant', ''),
        'delivery':       o.get('delivery', 'pickup'),
        'delivery_label': DELIVERY_LABEL.get(o.get('delivery', 'pickup'), '—'),
        'address':        o.get('address', ''),
        'status':         status,
        'status_emoji':   meta['emoji'],
        'status_label':   meta['label'],
        'payment_method': o.get('payment_method', ''),
        'created':        o.get('created', ''),
    }

@app.route('/api/v1/seller/orders', methods=['GET'])
@require_seller
def api_seller_orders():
    """Buyurtmalar list filter va pagination bilan."""
    uid = g.seller_uid
    status = (request.args.get('status', 'confirming') or 'confirming').lower()
    valid_statuses = {'pending', 'confirming', 'confirmed', 'rejected', 'cancelled', 'all'}
    if status not in valid_statuses:
        return jsonify({'error': 'invalid_status'}), 400

    try:
        page = max(0, int(request.args.get('page', 0)))
    except ValueError:
        page = 0
    try:
        limit = min(50, max(1, int(request.args.get('limit', 20))))
    except ValueError:
        limit = 20
    search = (request.args.get('search', '') or '').lower().strip()

    pids = set(_seller_get_pids(uid))

    # Status counts (filter'siz, shu sotuvchi uchun)
    summary = {'pending': 0, 'confirming': 0, 'confirmed': 0, 'rejected': 0}
    for o in orders.values():
        if o.get('product_id') not in pids:
            continue
        st = o.get('status', '')
        if st in summary:
            summary[st] += 1

    # Filtered list
    filtered = []
    for code, o in orders.items():
        if o.get('product_id') not in pids:
            continue
        if status != 'all' and o.get('status') != status:
            continue
        if search:
            hay = (code.lower() + ' ' + (o.get('user_name', '') or '').lower())
            if search not in hay:
                continue
        filtered.append((code, o))

    # Sort by created desc
    filtered.sort(key=lambda x: _parse_order_dt(x[1]) or datetime(2000, 1, 1), reverse=True)

    total = len(filtered)
    pages = max(1, (total + limit - 1) // limit)
    page  = min(page, pages - 1) if total > 0 else 0
    chunk = filtered[page*limit : (page+1)*limit]

    items = [_format_order_item(code, o) for code, o in chunk]

    return jsonify({
        'items':    items,
        'total':    total,
        'page':     page,
        'pages':    pages,
        'has_next': (page + 1) < pages,
        'summary':  summary,
    })

@app.route('/api/v1/seller/categories', methods=['GET'])
@require_seller
def api_seller_categories():
    """Mahsulot kategoriyalari ro'yxati global product count bilan."""
    counts = {}
    for p in products.values():
        if p.get('status') == 'closed':
            continue
        if not p.get('is_active', True):
            continue
        cat = p.get('category', '')
        if cat:
            counts[cat] = counts.get(cat, 0) + 1
    items = []
    for name, icon in CATEGORIES:
        items.append({
            'name':           name,
            'icon':           icon,
            'products_count': counts.get(name, 0),
        })
    return jsonify({'categories': items})

# ─── SPRINT 1.2: Order detail + CRM ─────────────────────────────────

@app.route('/api/v1/seller/orders/<code>', methods=['GET'])
@require_seller
def api_seller_order_detail(code):
    """Buyurtma to'liq detail — product info, buyer lifetime, timeline."""
    uid = g.seller_uid
    o = orders.get(code)
    if not o:
        return jsonify({'error': 'not_found'}), 404
    pid = o.get('product_id', '')
    p = products.get(pid)
    if not p or p.get('seller_id') != uid:
        return jsonify({'error': 'not_found'}), 404

    item = _format_order_item(code, o)
    # Product subset — list ekrandagidan biroz boyroq
    cls = _classify_product_status(p)
    item['product'] = {
        'id':            pid,
        'name':          p.get('name', ''),
        'photo_url':     p.get('photo_url') or (f"/api/photo/{p['photo_id']}" if p.get('photo_id') else ''),
        'original_price':p.get('original_price', 0),
        'group_price':   p.get('group_price', 0),
        'solo_price':    p.get('solo_price', 0),
        'min_group':     p.get('min_group', 0),
        'count':         len(groups.get(pid, [])),
        'status':        p.get('status', 'active'),
        'status_label':  cls['label'],
        'shop_name':     p.get('shop_name', ''),
        'channel':       p.get('seller_channel', ''),
    }
    # Buyer lifetime — CRM'dan
    sid = str(uid)
    cuid = str(o.get('user_id', ''))
    cust = customers.get(sid, {}).get(cuid, {})
    item['buyer']['total_orders']    = cust.get('total_orders', 0)
    item['buyer']['lifetime_value']  = cust.get('total_spent', 0)
    item['buyer']['tags']            = cust.get('tags', [])
    item['buyer']['first_order']     = cust.get('first_order', '')
    item['buyer']['last_order']      = cust.get('last_order', '')

    # Timeline — order state transitions (mavjud field'lardan inferred)
    timeline = [{'event': 'created', 'at': o.get('created', ''), 'meta': {}}]
    payment_id = o.get('telegram_payment_charge_id') or ''
    payment_method = o.get('payment_method') or ''
    if payment_method or payment_id:
        timeline.append({
            'event': 'payment',
            'at':    o.get('created', ''),  # Telegram payment timestamp ayri saqlanmaydi
            'meta':  {'method': payment_method, 'tg_charge_id': payment_id},
        })
    status = o.get('status', 'pending')
    if status == 'confirmed':
        timeline.append({'event': 'confirmed', 'at': o.get('confirmed_at', ''), 'meta': {}})
    elif status == 'rejected':
        timeline.append({'event': 'rejected', 'at': o.get('rejected_at', ''),
                         'meta': {'reason': o.get('reject_reason', '')}})
    elif status == 'cancelled':
        timeline.append({'event': 'cancelled', 'at': '', 'meta': {}})
    item['timeline'] = timeline

    return jsonify(item)

# ─── CRM helpers ────────────────────────────────────────────────────

def _days_since_last(c):
    """last_order'dan beri necha kun. Parse fail bo'lsa 999."""
    try:
        last_dt = datetime.strptime(c.get('last_order', '01.01.2020'), '%d.%m.%Y')
        return (datetime.now() - last_dt).days
    except (ValueError, TypeError):
        return 999

def _classify_customer_activity(c):
    """Returns (key, emoji, label) — 🟢 Faol, 🟡 O'rtacha, 🔴 Yo'qotilgan."""
    d = _days_since_last(c)
    if d < 7:    return ('active',  '🟢', 'Faol')
    if d < 30:   return ('average', '🟡', "O'rtacha")
    return ('lost', '🔴', "Yo'qotilgan")

def _format_customer_brief(cuid, c, rank=None):
    """Customer list item format."""
    act_key, act_emoji, act_label = _classify_customer_activity(c)
    medal = None
    if rank is not None and rank <= 3:
        medal = ['🥇', '🥈', '🥉'][rank - 1]
    return {
        'cuid':            cuid,
        'user_id':         c.get('user_id', 0),
        'name':            c.get('name', '—'),
        'phone':           c.get('phone', ''),
        'username':        c.get('username', ''),
        'total_orders':    c.get('total_orders', 0),
        'total_spent':     c.get('total_spent', 0),
        'first_order':     c.get('first_order', ''),
        'last_order':      c.get('last_order', ''),
        'days_since_last': _days_since_last(c),
        'activity':        act_key,
        'activity_emoji':  act_emoji,
        'activity_label':  act_label,
        'tags':            c.get('tags', []),
        'rank':            rank,
        'medal':           medal,
    }

@app.route('/api/v1/seller/customers', methods=['GET'])
@require_seller
def api_seller_customers():
    """CRM mijozlar ro'yxati filter va pagination bilan."""
    uid = g.seller_uid
    filt = (request.args.get('filter', 'all') or 'all').lower()
    valid_filters = {'all', 'vip', 'active', 'lost', 'new', 'repeat'}
    if filt not in valid_filters:
        return jsonify({'error': 'invalid_filter'}), 400

    try:
        page = max(0, int(request.args.get('page', 0)))
    except ValueError:
        page = 0
    try:
        limit = min(50, max(1, int(request.args.get('limit', 20))))
    except ValueError:
        limit = 20
    search = (request.args.get('search', '') or '').lower().strip()

    sid = str(uid)
    my_custs = customers.get(sid, {})

    # Summary (filter'siz, hammasi)
    total_revenue = sum(v.get('total_spent', 0) for v in my_custs.values())
    summary = {
        'total':         len(my_custs),
        'vip':           sum(1 for v in my_custs.values() if 'vip' in v.get('tags', [])),
        'active':        sum(1 for v in my_custs.values() if _days_since_last(v) < 7),
        'lost':          sum(1 for v in my_custs.values() if _days_since_last(v) >= 30),
        'new':           sum(1 for v in my_custs.values() if v.get('total_orders', 0) == 1),
        'repeat':        sum(1 for v in my_custs.values() if v.get('total_orders', 0) > 1),
        'total_revenue': total_revenue,
    }

    # Filter
    items_filtered = []
    for cuid, c in my_custs.items():
        if filt == 'vip' and 'vip' not in c.get('tags', []):
            continue
        if filt == 'active' and _days_since_last(c) >= 7:
            continue
        if filt == 'lost' and _days_since_last(c) < 30:
            continue
        if filt == 'new' and c.get('total_orders', 0) != 1:
            continue
        if filt == 'repeat' and c.get('total_orders', 0) <= 1:
            continue
        if search:
            hay = (c.get('name', '') or '').lower()
            if search not in hay:
                continue
        items_filtered.append((cuid, c))

    # Sort by total_spent desc
    items_filtered.sort(key=lambda x: x[1].get('total_spent', 0), reverse=True)

    total = len(items_filtered)
    pages = max(1, (total + limit - 1) // limit)
    page  = min(page, pages - 1) if total > 0 else 0
    start_idx = page * limit
    chunk = items_filtered[start_idx : start_idx + limit]

    items = []
    for i, (cuid, c) in enumerate(chunk, start=start_idx + 1):
        items.append(_format_customer_brief(cuid, c, rank=i))

    return jsonify({
        'items':    items,
        'total':    total,
        'page':     page,
        'pages':    pages,
        'has_next': (page + 1) < pages,
        'filter':   filt,
        'summary':  summary,
    })

@app.route('/api/v1/seller/customers/<cuid>', methods=['GET'])
@require_seller
def api_seller_customer_detail(cuid):
    """Bitta mijoz to'liq detail."""
    uid = g.seller_uid
    sid = str(uid)
    c = customers.get(sid, {}).get(cuid)
    if not c:
        return jsonify({'error': 'not_found'}), 404

    avg = (c.get('total_spent', 0) // c.get('total_orders', 1)) if c.get('total_orders') else 0
    act_key, act_emoji, act_label = _classify_customer_activity(c)

    return jsonify({
        'cuid':           cuid,
        'user_id':        c.get('user_id', 0),
        'name':           c.get('name', '—'),
        'phone':          c.get('phone', ''),
        'username':       c.get('username', ''),
        'total_orders':   c.get('total_orders', 0),
        'total_spent':    c.get('total_spent', 0),
        'avg_check':      avg,
        'first_order':    c.get('first_order', ''),
        'last_order':     c.get('last_order', ''),
        'activity':       act_key,
        'activity_emoji': act_emoji,
        'activity_label': act_label,
        'tags':           c.get('tags', []),
        'note':           c.get('note', ''),
        'source':         c.get('source', 'order'),
        'available_tags': [
            {'id': 'vip',     'label': "⭐ VIP"},
            {'id': 'problem', 'label': "🔴 Muammoli"},
            {'id': 'loyal',   'label': "💎 Doimiy"},
        ],
    })

@app.route('/api/v1/seller/customers/<cuid>/history', methods=['GET'])
@require_seller
def api_seller_customer_history(cuid):
    """Mijozning xaridlar tarixi (oxirgi 20 ta saqlangan)."""
    uid = g.seller_uid
    sid = str(uid)
    c = customers.get(sid, {}).get(cuid)
    if not c:
        return jsonify({'error': 'not_found'}), 404

    try:
        page = max(0, int(request.args.get('page', 0)))
    except ValueError:
        page = 0
    try:
        limit = min(50, max(1, int(request.args.get('limit', 20))))
    except ValueError:
        limit = 20

    history = list(reversed(c.get('orders', [])))  # eng yangisi tepada
    total = len(history)
    pages = max(1, (total + limit - 1) // limit)
    page  = min(page, pages - 1) if total > 0 else 0
    chunk = history[page*limit : (page+1)*limit]

    items = []
    for o in chunk:
        items.append({
            'product':  o.get('product', '—'),
            'amount':   o.get('amount', 0),
            'date':     o.get('date', ''),
            'type':     o.get('type', ''),    # bo'sh bo'lishi mumkin (eski yozuvlarda yo'q)
            'status':   o.get('status', ''),  # bo'sh bo'lishi mumkin
        })

    return jsonify({
        'items':       items,
        'total':       total,
        'page':        page,
        'pages':       pages,
        'has_next':    (page + 1) < pages,
        'total_spent': c.get('total_spent', 0),
        'note':        "Oxirgi 20 ta xarid saqlanadi." if total >= 20 else None,
    })

# ─── SPRINT 1.3: Legal + Shops + Integrations ───────────────────────

LEGAL_STATUS_LABEL = {'yatt': 'YaTT', 'mchj': 'MChJ'}

@app.route('/api/v1/seller/legal', methods=['GET'])
@require_seller
def api_seller_legal():
    """Sotuvchining yuridik ma'lumotlari."""
    uid = g.seller_uid
    prof = seller_profiles.get(uid) or seller_profiles.get(str(uid)) or {}
    completed = bool(prof.get('legal_completed_at'))
    if not completed:
        return jsonify({
            'completed':         False,
            'completed_at':      None,
            'legal_status':      None,
            'legal_status_label':None,
            'stir':              None,
            'bank_name':         None,
            'bank_account':      None,
            'bank_account_formatted': None,
            'bank_mfo':          None,
            'director_name':     None,
        })
    acc = prof.get('bank_account') or ''
    return jsonify({
        'completed':              True,
        'completed_at':           prof.get('legal_completed_at'),
        'legal_status':           prof.get('legal_status'),
        'legal_status_label':     LEGAL_STATUS_LABEL.get(prof.get('legal_status'), '—'),
        'stir':                   prof.get('stir'),
        'bank_name':              prof.get('bank_name'),
        'bank_account':           acc,
        'bank_account_formatted': _format_account(acc),
        'bank_mfo':               prof.get('bank_mfo'),
        'director_name':          prof.get('director_name'),
    })

DELIVERY_LABEL_FULL = {
    'pickup':  '🏪 Olib ketish',
    'deliver': '🚚 Yetkazib berish',
    'both':    '🚚🏪 Ikkalasi',
}

def _format_shop_brief(idx, shop, uid):
    """Shop list item format."""
    channel = shop.get('channel', '')
    products_count = sum(
        1 for pid in _seller_get_pids(uid)
        if pid in products
        and products[pid].get('seller_channel') == channel
        and products[pid].get('status') != 'closed'
    )
    return {
        'idx':               idx,
        'name':              shop.get('name', ''),
        'phone':             shop.get('phone', ''),
        'phone2':            shop.get('phone2', ''),
        'address':           shop.get('address', ''),
        'social':            shop.get('social', {}),
        'delivery':          shop.get('delivery', 'pickup'),
        'delivery_label':    DELIVERY_LABEL_FULL.get(shop.get('delivery', 'pickup'), '—'),
        'channel':           channel,
        'channel_verified':  channel in verified_channels,
        'verified':          shop.get('verified', False),
        'onboarding_status': shop.get('onboarding_status', 'active'),
        'products_count':    products_count,
        'billz_connected':   bool(shop.get('billz_secret_token')),
        'billz_shop_name':   shop.get('billz_shop_name', ''),
    }

@app.route('/api/v1/seller/shops', methods=['GET'])
@require_seller
def api_seller_shops():
    """Sotuvchining barcha do'konlari."""
    uid = g.seller_uid
    shops = _seller_get_shops(uid)
    return jsonify({
        'shops': [_format_shop_brief(i, sh, uid) for i, sh in enumerate(shops)],
    })

@app.route('/api/v1/seller/shops/<int:idx>', methods=['GET'])
@require_seller
def api_seller_shop_detail(idx):
    """Bitta shop detail — extra metrics bilan."""
    uid = g.seller_uid
    shops = _seller_get_shops(uid)
    if idx < 0 or idx >= len(shops):
        return jsonify({'error': 'not_found'}), 404
    shop = shops[idx]
    brief = _format_shop_brief(idx, shop, uid)
    # Extra metrics — orders va revenue shu shop kanaligi bo'yicha
    channel = shop.get('channel', '')
    shop_pids = {pid for pid in _seller_get_pids(uid)
                 if pid in products and products[pid].get('seller_channel') == channel}
    shop_orders = [o for o in orders.values() if o.get('product_id') in shop_pids]
    confirmed = [o for o in shop_orders if o.get('status') == 'confirmed']
    last_order_dt = max((_parse_order_dt(o) for o in confirmed if _parse_order_dt(o)), default=None)
    brief.update({
        'orders_total':     len(shop_orders),
        'orders_confirmed': len(confirmed),
        'revenue':          sum(o.get('amount', 0) for o in confirmed),
        'last_order':       last_order_dt.strftime('%d.%m.%Y') if last_order_dt else '',
    })
    return jsonify(brief)

# ─── SPRINT 1.4: chart + billz detail + mxik proxy ──────────────────

@app.route('/api/v1/seller/stats/chart', methods=['GET'])
@require_seller
def api_seller_stats_chart():
    """Kunlik GMV chart data — N kun uchun.
    days: 7 | 14 | 30 | 60 | 90 (default: 14)"""
    uid = g.seller_uid
    try:
        days = int(request.args.get('days', 14))
    except ValueError:
        days = 14
    if days not in (7, 14, 30, 60, 90):
        return jsonify({'error': 'invalid_days', 'reason': 'allowed: 7,14,30,60,90'}), 400

    pids = set(_seller_get_pids(uid))
    confirmed = [o for o in orders.values()
                 if o.get('product_id') in pids and o.get('status') == 'confirmed']

    today  = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    cutoff = today - timedelta(days=days - 1)

    # Bucket'larni nol bilan to'ldirish (continuous chart uchun)
    daily = {}
    for i in range(days):
        day = cutoff + timedelta(days=i)
        daily[day.strftime('%Y-%m-%d')] = {'gmv': 0, 'orders': 0}

    for o in confirmed:
        dt = _parse_order_dt(o)
        if not dt:
            continue
        day_key = dt.strftime('%Y-%m-%d')
        if day_key in daily:
            daily[day_key]['gmv']    += o.get('amount', 0)
            daily[day_key]['orders'] += 1

    data = [{'date': k, 'gmv': v['gmv'], 'orders': v['orders']}
            for k, v in sorted(daily.items())]
    total_gmv = sum(d['gmv'] for d in data)
    avg_daily = total_gmv // days if days else 0

    return jsonify({
        'days':      days,
        'data':      data,
        'total_gmv': total_gmv,
        'avg_daily': avg_daily,
    })

@app.route('/api/v1/seller/integrations/billz/<int:shop_idx>', methods=['GET'])
@require_seller
def api_seller_integration_billz(shop_idx):
    """Billz shop detail per-shop."""
    uid = g.seller_uid
    shops = _seller_get_shops(uid)
    if shop_idx < 0 or shop_idx >= len(shops):
        return jsonify({'error': 'not_found'}), 404
    shop = shops[shop_idx]
    connected = bool(shop.get('billz_secret_token'))

    imported_count = 0
    if connected:
        # Sotuvchining Billz mahsulotlarini sanaymiz (shop bo'yicha aniq filter
        # _billz_make_product_dict'da billz_shop_id saqlanmaydi — umumiy hisob)
        for pid in _seller_get_pids(uid):
            p = products.get(pid)
            if p and p.get('source') == 'billz':
                imported_count += 1

    return jsonify({
        'shop_idx':  shop_idx,
        'shop_name': shop.get('name', ''),
        'billz': {
            'connected':             connected,
            'billz_shop_id':         shop.get('billz_shop_id', '') if connected else '',
            'billz_shop_name':       shop.get('billz_shop_name', '') if connected else '',
            'connected_at':          shop.get('billz_connected_at', '') if connected else '',
            'imported_count':        imported_count,
            'global_solo_discount':  shop.get('billz_global_solo_discount', 10),
            'global_group_discount': shop.get('billz_global_group_discount', 20),
        },
    })

@app.route('/api/v1/seller/mxik/search', methods=['GET'])
@require_seller
def api_seller_mxik_search():
    """MXIK qidiruv proxy — server-side cache va graceful degradation.
    tasnif.soliq.uz Render egress sekin bo'lsa, ok=false bilan xato qaytadi."""
    q = (request.args.get('q', '') or '').strip()
    if len(q) < 2:
        return jsonify({'error': 'invalid_query', 'reason': 'min 2 chars'}), 400
    results, err = mxik_search(q)
    if err:
        # Graceful — 200 with ok:false (Mini App'da xushmuomala xato xabari)
        return jsonify({'ok': False, 'error': err, 'results': []})
    return jsonify({'ok': True, 'results': results or [], 'count': len(results or [])})

@app.route('/api/v1/seller/integrations', methods=['GET'])
@require_seller
def api_seller_integrations():
    """Barcha integratsiyalar status — per-shop bo'yicha."""
    uid = g.seller_uid
    shops = _seller_get_shops(uid)
    items = []
    for entry in INTEGRATIONS:
        info = {
            'id':     entry['id'],
            'name':   entry['name'],
            'icon':   entry['icon'],
            'status': entry['status'],  # 'active' | 'coming_soon'
        }
        if entry['id'] == 'billz' and entry['status'] == 'active':
            shop_statuses = []
            connected = 0
            for i, sh in enumerate(shops):
                is_connected = bool(sh.get('billz_secret_token'))
                if is_connected:
                    connected += 1
                shop_statuses.append({
                    'shop_idx':         i,
                    'shop_name':        sh.get('name', ''),
                    'connected':        is_connected,
                    'billz_shop_name':  sh.get('billz_shop_name', ''),
                })
            info['connected_shops'] = connected
            info['total_shops']     = len(shops)
            info['shop_statuses']   = shop_statuses
        items.append(info)
    return jsonify({'integrations': items})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))

init_db()
load_data()
threading.Thread(target=setup_bot_ui, daemon=True).start()
